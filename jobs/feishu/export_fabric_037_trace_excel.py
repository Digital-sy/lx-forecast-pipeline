#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""
导出 037超绒 面料使用情况与预估用量溯源 Excel。

兼容说明：
- 当前服务器数据库可能存在旧版 `面料预估表`，缺少 `运营预计下单量`、`系统预估下单量` 等新字段。
- 本脚本会动态检查字段是否存在；缺失字段会用 0 或空值兜底，避免导出中断。
- 库存溯源优先读取正式面料预估逻辑使用的 `面料库存台账`。

输出内容：
1. 面料预估结果：来自 `面料预估表`
2. SPU用料关系：来自 `面料核价表`
3. SKU用量溯源：运营预计下单表 + 预测对比表_SKU + 面料核价表
4. 库存台账：面料库存台账
5. 颜色归并：面料颜色归并对照
6. 定制面料参数
7. 检查项：关键异常提示
"""

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

project_root = Path(__file__).resolve().parent.parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

from common import get_logger
from common.database import db_cursor

logger = get_logger('fabric_037_trace_export')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as e:
    raise SystemExit(
        "缺少 openpyxl 依赖，请先执行：\n"
        "  cd /opt/apps/pythondata\n"
        "  ./venv/bin/pip install openpyxl\n"
        "然后重新运行本脚本。"
    ) from e


DEFAULT_KEYWORDS = ['037', '超绒']


def _like_conditions(column_names: List[str], keywords: List[str]) -> Tuple[str, List[str]]:
    parts = []
    params: List[str] = []
    for col in column_names:
        for kw in keywords:
            parts.append(f"`{col}` LIKE %s")
            params.append(f"%{kw}%")
    return '(' + ' OR '.join(parts) + ')', params


def _fetch_all(sql: str, params: Tuple[Any, ...] = ()) -> List[Dict[str, Any]]:
    with db_cursor(dictionary=True) as cursor:
        cursor.execute(sql, params)
        return list(cursor.fetchall())


def _table_exists(table_name: str) -> bool:
    rows = _fetch_all(
        """
        SELECT COUNT(*) AS cnt
        FROM information_schema.TABLES
        WHERE TABLE_SCHEMA = DATABASE()
          AND TABLE_NAME = %s
        """,
        (table_name,),
    )
    return bool(rows and rows[0].get('cnt'))


def _get_columns(table_name: str) -> set:
    if not _table_exists(table_name):
        return set()
    rows = _fetch_all(
        """
        SELECT COLUMN_NAME
        FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA = DATABASE()
          AND TABLE_NAME = %s
        """,
        (table_name,),
    )
    return {r['COLUMN_NAME'] for r in rows}


def _col_expr(existing_cols: set, col_name: str, alias: str = None, default_sql: str = '0') -> str:
    alias = alias or col_name
    if col_name in existing_cols:
        return f"`{col_name}` AS `{alias}`"
    return f"{default_sql} AS `{alias}`"


def _normalize_date(value: Any) -> str:
    if value is None:
        return ''
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m-%d')
    return str(value)[:10]


def _extract_spu_sql(alias: str = 'k') -> str:
    return f"SUBSTRING_INDEX({alias}.SKU, '-', 1)"


def read_fabric_params(keywords: List[str]) -> List[Dict[str, Any]]:
    if not _table_exists('定制面料参数'):
        return []
    cols = _get_columns('定制面料参数')
    where_cols = [c for c in ['面料', '面料编号'] if c in cols]
    if not where_cols:
        return []
    where, params = _like_conditions(where_cols, keywords)
    select_cols = [
        _col_expr(cols, '面料', default_sql="''"),
        _col_expr(cols, '面料编号', default_sql="''"),
        _col_expr(cols, '米数每条', default_sql='0'),
        _col_expr(cols, '公斤数每条', default_sql='0'),
    ]
    return _fetch_all(
        f"""
        SELECT {', '.join(select_cols)}
        FROM `定制面料参数`
        WHERE {where}
        ORDER BY 面料, 面料编号
        """,
        tuple(params),
    )


def read_forecast_result(keywords: List[str]) -> List[Dict[str, Any]]:
    if not _table_exists('面料预估表'):
        return []
    cols = _get_columns('面料预估表')
    where_cols = [c for c in ['面料', '面料编号', '面料颜色编号'] if c in cols]
    if not where_cols:
        return []
    where, params = _like_conditions(where_cols, keywords)

    select_cols = [
        _col_expr(cols, '统计类型', default_sql="''"),
        _col_expr(cols, '月份', default_sql="''"),
        _col_expr(cols, '面料', default_sql="''"),
        _col_expr(cols, '面料编号', default_sql="''"),
        _col_expr(cols, '颜色缩写', default_sql="''"),
        _col_expr(cols, '颜色', default_sql="''"),
        _col_expr(cols, '面料颜色编号', default_sql="''"),
        _col_expr(cols, '统计日期', default_sql='NULL'),
        _col_expr(cols, '运营预计下单量', default_sql='0'),
        _col_expr(cols, '系统预估下单量', default_sql='0'),
        _col_expr(cols, '预计用量/米', default_sql='0'),
        _col_expr(cols, '系统预估用量/米', default_sql='0'),
        _col_expr(cols, '米数每条', default_sql='0'),
        _col_expr(cols, '预计用量/条', default_sql='0'),
        _col_expr(cols, '系统预估用量/条', default_sql='0'),
        _col_expr(cols, '库存量/条', default_sql='0'),
        _col_expr(cols, '库存量/米', default_sql='0'),
        _col_expr(cols, '待到货量/条', default_sql='0'),
        _col_expr(cols, '待到货量/米', default_sql='0'),
        _col_expr(cols, '预计总量/条', default_sql='0'),
        _col_expr(cols, '预计总量/米', default_sql='0'),
        _col_expr(cols, '用量信息缺失SPU', default_sql="''"),
        _col_expr(cols, '更新时间', default_sql='NULL'),
    ]
    rows = _fetch_all(
        f"""
        SELECT {', '.join(select_cols)}
        FROM `面料预估表`
        WHERE {where}
        ORDER BY 统计类型, 统计日期, 面料, 颜色缩写
        """,
        tuple(params),
    )
    for r in rows:
        r['统计日期'] = _normalize_date(r.get('统计日期'))
        r['更新时间'] = _normalize_date(r.get('更新时间'))
        r['系统预计采购条数'] = max(
            0,
            float(r.get('系统预估用量/条') or 0)
            - float(r.get('库存量/条') or 0)
            - float(r.get('待到货量/条') or 0),
        )
    return rows


def read_spu_fabric_usage(keywords: List[str]) -> List[Dict[str, Any]]:
    if not _table_exists('面料核价表'):
        return []
    cols = _get_columns('面料核价表')
    if '面料' not in cols or 'SPU' not in cols:
        return []
    where, params = _like_conditions(['面料'], keywords)
    unit_col = '`单件用量`' if '单件用量' in cols else '0'
    loss_col = '`单件损耗`' if '单件损耗' in cols else '0'
    return _fetch_all(
        f"""
        SELECT
            h.SPU,
            h.面料,
            {unit_col} AS 单件用量,
            {loss_col} AS 单件损耗,
            CASE
                WHEN IFNULL({unit_col}, 0) = (
                    SELECT MAX(IFNULL({unit_col.replace('`', 'h2.`') if unit_col.startswith('`') else unit_col}, 0))
                    FROM `面料核价表` h2
                    WHERE h2.SPU = h.SPU
                ) THEN '主面料'
                ELSE '非主面料'
            END AS 主面料判定
        FROM `面料核价表` h
        WHERE {where}
        ORDER BY h.SPU, 单件用量 DESC
        """,
        tuple(params),
    )


def read_sku_trace(keywords: List[str]) -> List[Dict[str, Any]]:
    if not _table_exists('面料核价表'):
        return []
    fabric_cols = _get_columns('面料核价表')
    if 'SPU' not in fabric_cols or '面料' not in fabric_cols:
        return []

    has_op = _table_exists('运营预计下单表')
    has_sys = _table_exists('预测对比表_SKU')
    if not has_op and not has_sys:
        return []

    keys_parts = []
    if has_op:
        keys_parts.append("""
            SELECT SKU, 统计日期
            FROM `运营预计下单表`
            WHERE SKU IS NOT NULL AND SKU != '' AND 统计日期 IS NOT NULL AND 预计下单量 > 0
        """)
    if has_sys:
        keys_parts.append("""
            SELECT SKU, 统计日期
            FROM `预测对比表_SKU`
            WHERE SKU IS NOT NULL AND SKU != '' AND 统计日期 IS NOT NULL AND 系统预测销量 > 0
        """)
    keys_union = " UNION ".join(keys_parts)

    op_cte = """
        op AS (
            SELECT SKU, 统计日期, SUM(预计下单量) AS 运营预计下单量
            FROM `运营预计下单表`
            GROUP BY SKU, 统计日期
        )
    """ if has_op else "op AS (SELECT NULL AS SKU, NULL AS 统计日期, 0 AS 运营预计下单量)"

    sys_cte = """
        sysf AS (
            SELECT SKU, 统计日期, SUM(系统预测销量) AS 系统预估下单量
            FROM `预测对比表_SKU`
            GROUP BY SKU, 统计日期
        )
    """ if has_sys else "sysf AS (SELECT NULL AS SKU, NULL AS 统计日期, 0 AS 系统预估下单量)"

    where, params = _like_conditions(['面料'], keywords)
    unit_col = 'h.`单件用量`' if '单件用量' in fabric_cols else '0'
    loss_col = 'h.`单件损耗`' if '单件损耗' in fabric_cols else '0'

    sql = f"""
        WITH keys_union AS (
            {keys_union}
        ),
        {op_cte},
        {sys_cte}
        SELECT
            k.SKU,
            {_extract_spu_sql('k')} AS SPU,
            k.统计日期,
            h.面料,
            {unit_col} AS 单件用量,
            {loss_col} AS 单件损耗,
            CASE
                WHEN IFNULL({unit_col}, 0) = (
                    SELECT MAX(IFNULL(h2.`单件用量`, 0))
                    FROM `面料核价表` h2
                    WHERE h2.SPU = h.SPU
                ) THEN '主面料'
                ELSE '非主面料'
            END AS 主面料判定,
            IFNULL(op.运营预计下单量, 0) AS 运营预计下单量,
            IFNULL(sysf.系统预估下单量, 0) AS 系统预估下单量,
            ROUND(
                IFNULL(op.运营预计下单量, 0)
                * IFNULL({unit_col}, 0)
                * CASE WHEN IFNULL({loss_col}, 0) = 0 THEN 1 ELSE {loss_col} END,
                2
            ) AS 运营预计用量_米,
            ROUND(
                IFNULL(sysf.系统预估下单量, 0)
                * IFNULL({unit_col}, 0)
                * CASE WHEN IFNULL({loss_col}, 0) = 0 THEN 1 ELSE {loss_col} END,
                2
            ) AS 系统预估用量_米
        FROM keys_union k
        JOIN `面料核价表` h
            ON h.SPU = {_extract_spu_sql('k')}
        LEFT JOIN op
            ON op.SKU = k.SKU AND op.统计日期 = k.统计日期
        LEFT JOIN sysf
            ON sysf.SKU = k.SKU AND sysf.统计日期 = k.统计日期
        WHERE {where}
        ORDER BY k.统计日期, h.SPU, k.SKU
    """
    rows = _fetch_all(sql, tuple(params))
    for r in rows:
        r['统计日期'] = _normalize_date(r.get('统计日期'))
    return rows


def read_inventory(keywords: List[str]) -> List[Dict[str, Any]]:
    """优先读取正式面料预估逻辑使用的 `面料库存台账`。"""
    if not _table_exists('面料库存台账'):
        return []
    cols = _get_columns('面料库存台账')
    if '面料编号颜色缩写' not in cols:
        return []
    where, params = _like_conditions(['面料编号颜色缩写'], keywords)
    select_cols = [
        _col_expr(cols, '面料编号颜色缩写', default_sql="''"),
        _col_expr(cols, '库存成品数量_条', default_sql='0'),
        _col_expr(cols, '备货中数量_条', default_sql='0'),
        _col_expr(cols, '现有胚布数量_条', default_sql='0'),
        _col_expr(cols, '面料', default_sql="''"),
        _col_expr(cols, '颜色', default_sql="''"),
        _col_expr(cols, '更新时间', default_sql='NULL'),
    ]
    rows = _fetch_all(
        f"""
        SELECT {', '.join(select_cols)}
        FROM `面料库存台账`
        WHERE {where}
        ORDER BY 面料编号颜色缩写
        """,
        tuple(params),
    )
    for r in rows:
        r['更新时间'] = _normalize_date(r.get('更新时间'))
    return rows


def read_color_merge(keywords: List[str]) -> List[Dict[str, Any]]:
    if not _table_exists('面料颜色归并对照'):
        return []
    cols = _get_columns('面料颜色归并对照')
    if '面料编号' not in cols:
        return []
    where, params = _like_conditions(['面料编号'], keywords)
    select_cols = [
        _col_expr(cols, '面料编号', default_sql="''"),
        _col_expr(cols, '原始颜色缩写', default_sql="''"),
        _col_expr(cols, '归并颜色缩写', default_sql="''"),
        _col_expr(cols, '是否启用', default_sql='0'),
    ]
    return _fetch_all(
        f"""
        SELECT {', '.join(select_cols)}
        FROM `面料颜色归并对照`
        WHERE {where}
        ORDER BY 面料编号, 原始颜色缩写
        """,
        tuple(params),
    )


def build_check_items(
    params_rows: List[Dict[str, Any]],
    result_rows: List[Dict[str, Any]],
    spu_rows: List[Dict[str, Any]],
    sku_rows: List[Dict[str, Any]],
    inventory_rows: List[Dict[str, Any]],
    color_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    checks = []

    def add(item: str, status: str, detail: str):
        checks.append({'检查项': item, '状态': status, '说明': detail})

    add('定制面料参数', '正常' if params_rows else '异常', f'命中 {len(params_rows)} 条')
    add('面料核价表/SPU用料关系', '正常' if spu_rows else '异常', f'命中 {len(spu_rows)} 个 SPU-面料关系')
    add('SKU用量溯源', '正常' if sku_rows else '异常', f'命中 {len(sku_rows)} 条 SKU+月份 来源明细')
    add('面料预估表最终结果', '正常' if result_rows else '异常', f'命中 {len(result_rows)} 条最终预估结果')
    add('面料库存台账', '正常' if inventory_rows else '提醒', f'命中 {len(inventory_rows)} 条库存台账；若为0，检查面料编号颜色缩写是否包含037')
    add('颜色归并', '正常' if color_rows else '提醒', f'命中 {len(color_rows)} 条颜色归并；没有不一定异常')

    zero_usage_spu = [r for r in spu_rows if float(r.get('单件用量') or 0) <= 0]
    add('单件用量为0', '异常' if zero_usage_spu else '正常', f'{len(zero_usage_spu)} 条 SPU 用量为0')

    zero_roll = [r for r in params_rows if float(r.get('米数每条') or 0) <= 0]
    add('米数每条为0', '异常' if zero_roll else '正常', f'{len(zero_roll)} 条定制面料参数 米数每条<=0')

    missing_usage = [r for r in result_rows if r.get('用量信息缺失SPU')]
    add('用量信息缺失SPU', '异常' if missing_usage else '正常', f'{len(missing_usage)} 条结果存在用量信息缺失SPU')

    return checks


def _safe_sheet_name(name: str) -> str:
    invalid = r'[]:*?/\\'
    for ch in invalid:
        name = name.replace(ch, '_')
    return name[:31]


def write_sheet(wb: Workbook, title: str, rows: List[Dict[str, Any]], freeze: str = 'A2'):
    ws = wb.create_sheet(_safe_sheet_name(title))
    if not rows:
        ws.append(['提示'])
        ws.append(['无数据'])
        ws.freeze_panes = freeze
        return ws

    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])

    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='D9E2F3')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center')

    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions

    end_row = ws.max_row
    end_col = ws.max_column
    if end_row >= 2 and end_col >= 1:
        ref = f"A1:{get_column_letter(end_col)}{end_row}"
        table_name = re.sub(r'[^A-Za-z0-9_]', '_', f"tbl_{title}")[:250]
        if not re.match(r'^[A-Za-z_]', table_name):
            table_name = f'T_{table_name}'
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name='TableStyleMedium2',
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        try:
            ws.add_table(table)
        except Exception:
            pass

    for idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 8
        for cell in col_cells:
            val = '' if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 42)

    return ws


def write_summary(wb: Workbook, summary: Dict[str, Any], checks: List[Dict[str, Any]]):
    ws = wb.active
    ws.title = '汇总说明'
    ws.append(['037超绒面料使用情况与预估用量溯源'])
    ws.append(['生成时间', summary['生成时间']])
    ws.append(['关键词', ', '.join(summary['关键词'])])
    ws.append(['输出说明', '用于核查037超绒的最终预估、SPU用料、SKU来源、库存台账和颜色归并。'])
    ws.append([])
    ws.append(['数据量汇总'])
    for key, value in summary['数据量汇总'].items():
        ws.append([key, value])
    ws.append([])
    ws.append(['检查项', '状态', '说明'])
    for r in checks:
        ws.append([r['检查项'], r['状态'], r['说明']])

    ws.merge_cells('A1:C1')
    ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='1F4E78')
    ws['A1'].alignment = Alignment(horizontal='center')

    for row in range(2, ws.max_row + 1):
        for col in range(1, 4):
            ws.cell(row=row, column=col).alignment = Alignment(vertical='center')
    for cell in ws[6]:
        cell.font = Font(bold=True)
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == '检查项':
            for cell in ws[row]:
                cell.fill = PatternFill('solid', fgColor='D9EAF7')
                cell.font = Font(bold=True)
            break
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 90
    ws.freeze_panes = 'A2'


def export_excel(output_path: str, keywords: List[str]):
    logger.info('开始读取037超绒相关数据...')
    params_rows = read_fabric_params(keywords)
    result_rows = read_forecast_result(keywords)
    spu_rows = read_spu_fabric_usage(keywords)
    sku_rows = read_sku_trace(keywords)
    inventory_rows = read_inventory(keywords)
    color_rows = read_color_merge(keywords)

    checks = build_check_items(params_rows, result_rows, spu_rows, sku_rows, inventory_rows, color_rows)
    summary = {
        '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        '关键词': keywords,
        '数据量汇总': {
            '面料预估结果': len(result_rows),
            'SPU用料关系': len(spu_rows),
            'SKU用量溯源': len(sku_rows),
            '库存台账': len(inventory_rows),
            '颜色归并': len(color_rows),
            '定制面料参数': len(params_rows),
        }
    }

    wb = Workbook()
    write_summary(wb, summary, checks)
    write_sheet(wb, '01_面料预估结果', result_rows)
    write_sheet(wb, '02_SPU用料关系', spu_rows)
    write_sheet(wb, '03_SKU用量溯源', sku_rows)
    write_sheet(wb, '04_库存台账', inventory_rows)
    write_sheet(wb, '05_颜色归并', color_rows)
    write_sheet(wb, '06_定制面料参数', params_rows)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    logger.info(f'✓ Excel 已生成: {output}')
    logger.info('数据量汇总: ' + ', '.join(f'{k}={v}' for k, v in summary['数据量汇总'].items()))


def parse_args():
    parser = argparse.ArgumentParser(description='导出037超绒面料使用情况与预估用量溯源Excel')
    parser.add_argument(
        '--fabric-keyword',
        action='append',
        default=None,
        help='面料匹配关键词，可重复传入。默认：037、超绒',
    )
    parser.add_argument(
        '--output',
        default=None,
        help='输出Excel路径，默认 exports/037超绒面料溯源_YYYYMMDD_HHMMSS.xlsx',
    )
    return parser.parse_args()


def main():
    args = parse_args()
    keywords = args.fabric_keyword or DEFAULT_KEYWORDS
    output_path = args.output or f"exports/037超绒面料溯源_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    export_excel(output_path, keywords)


if __name__ == '__main__':
    main()
