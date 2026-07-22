#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""导出颜色体系感知的采购建议 Excel。"""
from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from common import get_logger
from common.database import db_cursor
from jobs.feishu.procurement_color_logic import future_months

logger = get_logger("export_procurement_excel_color_system")

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
THIN = Side(style="thin", color="D9E1F2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def read_order_data(month_order: Sequence[str]) -> List[Dict[str, Any]]:
    with db_cursor() as cursor:
        cursor.execute("""
            SELECT * FROM `建议下单量表`
            ORDER BY 面料类型 DESC, SPU, 颜色体系, 颜色缩写, 店铺
        """)
        return list(cursor.fetchall())


def read_fabric_data() -> List[Dict[str, Any]]:
    with db_cursor() as cursor:
        cursor.execute("""
            SELECT 面料, SPU数量, 建议下单量合计,
                   `单件用量(米)`, `原始单耗加权均值`,
                   `预计用量(米)`, 计算口径
            FROM `面料预计用量表`
            ORDER BY `预计用量(米)` DESC
        """)
        return list(cursor.fetchall())


def read_color_exceptions() -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with db_cursor() as cursor:
        cursor.execute("""
            SELECT SPU, 颜色体系, 颜色缩写, 颜色汇总代码, 店铺,
                   建议下单量, 'SKU颜色体系待定' AS 异常类型
            FROM `建议下单量表`
            WHERE 颜色体系='待定'
            ORDER BY 建议下单量 DESC, SPU, 颜色缩写
        """)
        rows.extend(cursor.fetchall())

        cursor.execute("""
            SELECT '' AS SPU, 颜色体系, 颜色缩写, 颜色汇总代码,
                   '' AS 店铺, 0 AS 建议下单量,
                   CONCAT('面料库存：', 库存归属状态) AS 异常类型
            FROM `面料预估表`
            WHERE 统计类型='带颜色'
              AND 库存归属状态 LIKE '跨颜色体系冲突%'
            ORDER BY 面料, 颜色体系, 颜色缩写
        """)
        rows.extend(cursor.fetchall())
    return rows


def style_header(ws, row: int, headers: Sequence[str]) -> None:
    for column, header in enumerate(headers, 1):
        cell = ws.cell(row, column, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def autosize(ws, minimum: int = 10, maximum: int = 32) -> None:
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(length + 2, minimum), maximum)


def build_order_sheet(wb: Workbook, rows: Sequence[Dict[str, Any]], months: Sequence[str]) -> None:
    ws = wb.active
    ws.title = "建议下单量"
    headers = [
        "SPU", "颜色体系", "颜色缩写", "颜色汇总代码", "店铺", "工厂",
        "面料类型", "覆盖月数", "库存", "待到货",
    ]
    for month in months:
        headers.extend([f"{month}运营预计", f"{month}建议下单"])
    headers.extend(["运营预计合计", "建议下单合计"])
    style_header(ws, 1, headers)

    for row_index, row in enumerate(rows, 2):
        values: List[Any] = [
            row.get("SPU", ""), row.get("颜色体系", ""), row.get("颜色缩写", ""),
            row.get("颜色汇总代码", ""), row.get("店铺", ""), row.get("工厂", ""),
            row.get("面料类型", ""), int(row.get("覆盖月数", 0) or 0),
            int(row.get("库存", 0) or 0), int(row.get("待到货", 0) or 0),
        ]
        for month in months:
            values.extend([
                int(row.get(f"{month}运营预计", 0) or 0),
                int(row.get(f"{month}建议下单", 0) or 0),
            ])
        values.extend([
            int(row.get("运营预计合计", 0) or 0),
            int(row.get("建议下单合计", 0) or 0),
        ])
        for column, value in enumerate(values, 1):
            cell = ws.cell(row_index, column, value)
            cell.border = BORDER
            cell.alignment = RIGHT if isinstance(value, (int, float)) else LEFT
            if isinstance(value, int):
                cell.number_format = "#,##0"
        if row.get("颜色体系") == "待定":
            for cell in ws[row_index]:
                cell.fill = PatternFill("solid", fgColor="FFF2CC")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize(ws)


def build_fabric_sheet(wb: Workbook, rows: Sequence[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("面料预计用量")
    headers = [
        "面料", "SPU数量", "建议下单量合计", "有效加权单耗(米)",
        "原始单耗加权均值", "预计用量(米)", "计算口径",
    ]
    style_header(ws, 1, headers)
    for row_index, row in enumerate(rows, 2):
        values = [
            row.get("面料", ""), int(row.get("SPU数量", 0) or 0),
            int(row.get("建议下单量合计", 0) or 0),
            float(row.get("单件用量(米)", 0) or 0),
            float(row.get("原始单耗加权均值", 0) or 0),
            float(row.get("预计用量(米)", 0) or 0), row.get("计算口径", ""),
        ]
        for column, value in enumerate(values, 1):
            cell = ws.cell(row_index, column, value)
            cell.border = BORDER
            cell.alignment = RIGHT if isinstance(value, (int, float)) else LEFT
            if column in (4, 5):
                cell.number_format = "0.000"
            elif column == 6:
                cell.number_format = "#,##0.00"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize(ws)


def build_exception_sheet(wb: Workbook, rows: Sequence[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("颜色体系待处理")
    headers = ["异常类型", "SPU", "颜色体系", "颜色缩写", "颜色汇总代码", "店铺", "建议下单量"]
    style_header(ws, 1, headers)
    for row_index, row in enumerate(rows, 2):
        values = [
            row.get("异常类型", ""), row.get("SPU", ""), row.get("颜色体系", ""),
            row.get("颜色缩写", ""), row.get("颜色汇总代码", ""), row.get("店铺", ""),
            int(row.get("建议下单量", 0) or 0),
        ]
        for column, value in enumerate(values, 1):
            cell = ws.cell(row_index, column, value)
            cell.border = BORDER
            cell.alignment = RIGHT if isinstance(value, int) else LEFT
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize(ws)


def main() -> Path:
    now = datetime.now()
    months = [label for _, label in future_months(now, 4)]
    order_rows = read_order_data(months)
    fabric_rows = read_fabric_data()
    exception_rows = read_color_exceptions()

    workbook = Workbook()
    build_order_sheet(workbook, order_rows, months)
    build_fabric_sheet(workbook, fabric_rows)
    build_exception_sheet(workbook, exception_rows)

    export_dir = Path(os.getenv("PROCUREMENT_EXPORT_DIR", "/opt/apps/pythondata/exports"))
    export_dir.mkdir(parents=True, exist_ok=True)
    output = export_dir / f"采购建议报告_颜色体系_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    workbook.save(output)
    logger.info(
        f"✓ 导出完成：{output}；建议下单 {len(order_rows)} 行，"
        f"面料 {len(fabric_rows)} 行，颜色待处理 {len(exception_rows)} 行"
    )
    return output


if __name__ == "__main__":
    main()
