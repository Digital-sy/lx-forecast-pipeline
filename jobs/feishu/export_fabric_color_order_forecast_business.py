#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""按业务固定表头导出最终“面料预估表”，并可同步写入飞书。

计算完全复用 ``export_fabric_color_order_forecast_final.build_final_rows``，只调整
Excel / 飞书展示结构，不改变颜色匹配、A2023/B2024 判定、库存归属和面料用量口径。

正式业务口径使用方案 B：SPU 只要配置了目标面料，该面料就按当前 SKU 的最终飞书颜色拆分。
方案 A（仅主面料拆颜色）每天保留一份 MySQL 快照，仅用于历史对照，不写飞书。

主表固定为 21 列：
SKU、面料颜色编号、面料、颜色、领星颜色、库存量/米、待到货量/米、统计类型、
面料编号、颜色缩写、库存量/条、待到货量/条、用量信息缺失SPU、当月完整预估/米、
当月已下单消耗/米、当月剩余预估/米、T+1月预估/米、T+2月预估/米、
运营当月预估/米、运营T+1月预估/米、运营T+2月预估/米。

月份标题随运行月份动态滚动。例如 2026-08 运行时显示 8月、9月、10月。
A2023/B2024 仍严格来自当前 SKU 自身明确标签，飞书颜色和人工映射绝不反推体系。
"""
from __future__ import annotations

import argparse
import asyncio
import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import Workbook

from common import get_logger
from common.database import db_cursor
from common.feishu import FeishuClient
from jobs.feishu import fabric_color_stocking as stocking
from jobs.feishu import export_fabric_color_order_forecast_final as final_export
from jobs.feishu import generate_procurement_report as procurement_base
from jobs.feishu.fabric_color_stocking_spu import DEFAULT_SPU_MANUAL_MAPPING_PATH

logger = get_logger("export_fabric_color_order_forecast_business")

DEFAULT_FEISHU_TABLE_NAME = os.getenv(
    "FABRIC_FORECAST_FEISHU_TABLE_NAME", "面料预估明细"
)
SCHEME_A_HISTORY_TABLE = os.getenv(
    "FABRIC_FORECAST_SCHEME_A_HISTORY_TABLE", "面料预估方案A历史"
)


def _month_number(label: str) -> str:
    """26年8月 -> 8月；无法解析时原样返回。"""
    match = re.search(r"(\d{1,2})月", str(label or ""))
    return f"{int(match.group(1))}月" if match else str(label or "")


def _find_month_labels(rows: Sequence[Mapping[str, Any]]) -> list[str]:
    """从最终计算结果字段名提取 T/T+1/T+2/T+3 月份标签。"""
    labels: list[str] = []
    if rows:
        for key in rows[0].keys():
            text = str(key)
            match = re.fullmatch(r"(.+月)完整预估/米", text)
            if match:
                labels.append(match.group(1))
                break
        current = labels[0] if labels else ""
        following: list[str] = []
        for key in rows[0].keys():
            text = str(key)
            match = re.fullmatch(r"(.+月)预估/米", text)
            if not match or text.startswith("运营"):
                continue
            label = match.group(1)
            if label != current and label not in following:
                following.append(label)
        labels.extend(following)
    return labels[:4]


async def _load_lingxing_name_by_record_id() -> dict[str, str]:
    """读取飞书字段“领星新颜色名称”，仅用于展示列“领星颜色”。"""
    client = FeishuClient(
        app_token=os.getenv(
            "FABRIC_COLOR_CATALOG_BASE_TOKEN", stocking.DEFAULT_BASE_TOKEN
        ),
        table_id=os.getenv(
            "FABRIC_COLOR_CATALOG_TABLE_ID", stocking.DEFAULT_CATALOG_TABLE_ID
        ),
        view_id=os.getenv(
            "FABRIC_COLOR_CATALOG_VIEW_ID", stocking.DEFAULT_CATALOG_VIEW_ID
        ) or None,
    )
    field_map = await client.get_table_fields()
    actual_fields = set(field_map.values())
    target_field = "领星新颜色名称"
    if target_field not in actual_fields:
        logger.warning("飞书清单缺少字段“领星新颜色名称”，主表“领星颜色”将为空")
        return {}

    records = await client.read_records(page_size=500)
    result: dict[str, str] = {}
    for record in records:
        record_id = str(record.get("record_id") or "")
        values = stocking._texts((record.get("fields") or {}).get(target_field))
        if record_id and values:
            result[record_id] = values[0]
    return result


def _lingxing_name(row: Mapping[str, Any], by_record_id: Mapping[str, str]) -> str:
    ids = [
        value.strip()
        for value in str(row.get("飞书记录ID") or "").split("、")
        if value.strip()
    ]
    values = [by_record_id[record_id] for record_id in ids if record_id in by_record_id]
    return "、".join(dict.fromkeys(values))


def _business_headers(month_labels: Sequence[str]) -> list[str]:
    m0 = _month_number(month_labels[0]) if len(month_labels) > 0 else "当月"
    m1 = _month_number(month_labels[1]) if len(month_labels) > 1 else "T+1月"
    m2 = _month_number(month_labels[2]) if len(month_labels) > 2 else "T+2月"
    return [
        "SKU",
        "面料颜色编号",
        "面料",
        "颜色",
        "领星颜色",
        "库存量/米",
        "待到货量/米",
        "统计类型",
        "面料编号",
        "颜色缩写",
        "库存量/条",
        "待到货量/条",
        "用量信息缺失SPU",
        f"{m0}完整预估/米",
        f"{m0}已下单消耗/米",
        f"{m0}剩余预估/米",
        f"{m1}预估/米",
        f"{m2}预估/米",
        f"运营{m0}预估/米",
        f"运营{m1}预估/米",
        f"运营{m2}预估/米",
    ]


def _demand_value(row: Mapping[str, Any], label: str, suffix: str) -> Any:
    return row.get(f"{label}{suffix}", 0)


def _to_business_color_row(
    row: Mapping[str, Any],
    labels: Sequence[str],
    lingxing_names: Mapping[str, str],
) -> dict[str, Any]:
    m0 = labels[0]
    m1 = labels[1]
    m2 = labels[2]
    return {
        # 当前最终预计下单表是“面料+飞书颜色”聚合结果，一行会包含多个 SKU；
        # 为保持旧预估表兼容，聚合行 SKU 留空，不拼接超长 SKU 列表。
        "SKU": "",
        "面料颜色编号": row.get("库存匹配键") or "",
        "面料": row.get("面料") or "",
        "颜色": row.get("最终飞书颜色") or "",
        "领星颜色": _lingxing_name(row, lingxing_names),
        "库存量/米": row.get("库存量/米") or 0,
        "待到货量/米": row.get("待到货量/米") or 0,
        "统计类型": "带颜色",
        "面料编号": row.get("面料编号") or "",
        "颜色缩写": row.get("领星新颜色缩写") or "",
        "库存量/条": row.get("库存量/条") or 0,
        "待到货量/条": row.get("待到货量/条") or 0,
        "用量信息缺失SPU": row.get("用量信息缺失SPU") or "",
        f"{_month_number(m0)}完整预估/米": _demand_value(row, m0, "完整预估/米"),
        f"{_month_number(m0)}已下单消耗/米": _demand_value(row, m0, "已下单消耗/米"),
        f"{_month_number(m0)}剩余预估/米": _demand_value(row, m0, "剩余预估/米"),
        f"{_month_number(m1)}预估/米": _demand_value(row, m1, "预估/米"),
        f"{_month_number(m2)}预估/米": _demand_value(row, m2, "预估/米"),
        f"运营{_month_number(m0)}预估/米": _demand_value(row, f"运营{m0}", "预估/米"),
        f"运营{_month_number(m1)}预估/米": _demand_value(row, f"运营{m1}", "预估/米"),
        f"运营{_month_number(m2)}预估/米": _demand_value(row, f"运营{m2}", "预估/米"),
    }


def _to_business_total_row(
    row: Mapping[str, Any],
    labels: Sequence[str],
) -> dict[str, Any]:
    m0 = labels[0]
    m1 = labels[1]
    m2 = labels[2]
    return {
        "SKU": "",
        "面料颜色编号": "",
        "面料": row.get("面料") or "",
        "颜色": "",
        "领星颜色": "",
        "库存量/米": row.get("库存量/米") or 0,
        "待到货量/米": row.get("待到货量/米") or 0,
        "统计类型": "总量",
        "面料编号": row.get("面料编号") or "",
        "颜色缩写": "",
        "库存量/条": row.get("库存量/条") or 0,
        "待到货量/条": row.get("待到货量/条") or 0,
        "用量信息缺失SPU": row.get("用量信息缺失SPU") or "",
        f"{_month_number(m0)}完整预估/米": _demand_value(row, m0, "完整预估/米"),
        f"{_month_number(m0)}已下单消耗/米": _demand_value(row, m0, "已下单消耗/米"),
        f"{_month_number(m0)}剩余预估/米": _demand_value(row, m0, "剩余预估/米"),
        f"{_month_number(m1)}预估/米": _demand_value(row, m1, "预估/米"),
        f"{_month_number(m2)}预估/米": _demand_value(row, m2, "预估/米"),
        f"运营{_month_number(m0)}预估/米": _demand_value(row, f"运营{m0}", "预估/米"),
        f"运营{_month_number(m1)}预估/米": _demand_value(row, f"运营{m1}", "预估/米"),
        f"运营{_month_number(m2)}预估/米": _demand_value(row, f"运营{m2}", "预估/米"),
    }


def build_business_rows(
    color_rows: Sequence[Mapping[str, Any]],
    total_rows: Sequence[Mapping[str, Any]],
    lingxing_names: Mapping[str, str],
) -> tuple[list[str], list[dict[str, Any]]]:
    """把最终计算结果转换成业务固定 21 列。"""
    labels = _find_month_labels(color_rows or total_rows)
    if len(labels) < 3:
        raise RuntimeError(f"无法识别连续3个月份字段，实际识别到: {labels}")

    headers = _business_headers(labels)
    business_rows: list[dict[str, Any]] = []
    colors_by_fabric: dict[str, list[Mapping[str, Any]]] = {}
    for row in color_rows:
        colors_by_fabric.setdefault(str(row.get("面料") or ""), []).append(row)

    # 每个面料先放总量，再放该面料各颜色，便于业务查看。
    for total in total_rows:
        fabric = str(total.get("面料") or "")
        business_rows.append(_to_business_total_row(total, labels))
        for color in colors_by_fabric.get(fabric, ()):
            business_rows.append(_to_business_color_row(color, labels, lingxing_names))
    return headers, business_rows


def export_business_workbook(
    color_rows: Sequence[Mapping[str, Any]],
    total_rows: Sequence[Mapping[str, Any]],
    pending_rows: Sequence[Mapping[str, Any]],
    metrics: Mapping[str, Any],
    lingxing_names: Mapping[str, str],
    output_dir: Path,
) -> Path:
    headers, business_rows = build_business_rows(
        color_rows, total_rows, lingxing_names
    )

    workbook = Workbook()
    ws = workbook.active
    ws.title = "面料预估表"
    final_export._write_rows(ws, headers, business_rows)

    # 待确认和核对摘要继续保留，避免格式收口后丢失治理信息。
    pending_ws = workbook.create_sheet("待确认颜色")
    pending_headers = list(pending_rows[0].keys()) if pending_rows else [
        "面料", "颜色体系", "原始颜色编码", "未确认原因"
    ]
    final_export._write_rows(pending_ws, pending_headers, pending_rows, pending=True)

    summary_ws = workbook.create_sheet("核对摘要")
    final_export._style_header(summary_ws, ["指标", "值"])
    summary_metrics = dict(metrics)
    summary_metrics["business_sheet_headers"] = headers
    summary_metrics["business_sheet_row_count"] = len(business_rows)
    summary_metrics["business_sheet_month_policy"] = "主表仅展示当月、T+1、T+2三个月；月份标题随运行月份动态滚动"
    row_index = 2
    for key, value in summary_metrics.items():
        if isinstance(value, (dict, list, tuple)):
            value = json.dumps(value, ensure_ascii=False, sort_keys=True)
        summary_ws.cell(row_index, 1, key)
        summary_ws.cell(row_index, 2, value)
        for cell in summary_ws[row_index]:
            cell.border = final_export.BORDER
            cell.alignment = final_export.LEFT
        row_index += 1
    summary_ws.freeze_panes = "A2"
    final_export._autosize(summary_ws, maximum=80)

    output_dir.mkdir(parents=True, exist_ok=True)
    output = output_dir / f"面料预估表_最终版_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    workbook.save(output)
    logger.info(
        "业务表头预计下单表完成：%s；主表 %d 行，颜色 %d 行，总量 %d 行，待确认 %d 行，覆盖率 %.2f%%",
        output,
        len(business_rows),
        len(color_rows),
        len(total_rows),
        len(pending_rows),
        float(metrics.get("confirmed_coverage_pct") or 0),
    )
    return output


def _field_list(headers: Sequence[str]) -> list[dict[str, Any]]:
    """飞书 21 列字段定义；米数/条数字段均使用数字类型。"""
    result: list[dict[str, Any]] = []
    for header in headers:
        is_number = (
            header.endswith("/米")
            or header.endswith("/条")
            or header in {"库存量/米", "待到货量/米", "库存量/条", "待到货量/条"}
        )
        if is_number:
            result.append({"name": header, "type": "number", "precision": 2})
        else:
            result.append({"name": header, "type": "text"})
    return result


async def write_business_rows_to_feishu(
    headers: Sequence[str],
    business_rows: Sequence[Mapping[str, Any]],
    table_name: str = DEFAULT_FEISHU_TABLE_NAME,
) -> int:
    """用最终业务 21 列全量覆盖飞书“面料预估明细”。"""
    if not business_rows:
        raise RuntimeError("最终业务面料预估结果为空，拒绝清空飞书表")

    client = await procurement_base._get_or_create_table(
        procurement_base.FEISHU_APP_TOKEN,
        table_name,
        _field_list(headers),
        remove_extra=True,
    )
    old_count = await client.delete_all_records()
    logger.info("飞书%s：已清空旧记录 %d 条", table_name, old_count)
    written = await client.write_records(list(business_rows), batch_size=500)
    if written != len(business_rows):
        raise RuntimeError(
            f"飞书写入数量不一致：应写 {len(business_rows)} 条，实际 {written} 条"
        )
    logger.info("✓ 飞书%s写入最终业务面料预估 %d 条", table_name, written)
    return written


def _validated_table_name(table_name: str) -> str:
    """表名来自环境变量；只接受中文/字母/数字/下划线，避免动态 SQL 注入。"""
    value = str(table_name or "").strip()
    if not value or not re.fullmatch(r"[\w\u4e00-\u9fff]+", value):
        raise ValueError(f"非法方案A历史表名: {table_name!r}")
    return value


def archive_scheme_a_to_mysql(
    color_rows: Sequence[Mapping[str, Any]],
    total_rows: Sequence[Mapping[str, Any]],
    lingxing_names: Mapping[str, str],
    table_name: str = SCHEME_A_HISTORY_TABLE,
) -> int:
    """保存方案A的每日业务21列快照；同一天重跑时覆盖当天快照。"""
    labels = _find_month_labels(color_rows or total_rows)
    if len(labels) < 3:
        raise RuntimeError(f"方案A无法识别连续3个月份字段，拒绝覆盖历史快照: {labels}")

    _, business_rows = build_business_rows(color_rows, total_rows, lingxing_names)
    if not business_rows:
        raise RuntimeError("方案A业务结果为空，拒绝删除当天数据库快照")

    table_name = _validated_table_name(table_name)
    now = datetime.now()
    business_date = now.date()
    m0, m1, m2 = labels[:3]
    h0, h1, h2 = _month_number(m0), _month_number(m1), _month_number(m2)

    create_sql = f"""
        CREATE TABLE IF NOT EXISTS `{table_name}` (
            `id` BIGINT UNSIGNED NOT NULL AUTO_INCREMENT,
            `业务日期` DATE NOT NULL,
            `生成时间` DATETIME NOT NULL,
            `方案` VARCHAR(8) NOT NULL DEFAULT 'A',
            `当月标签` VARCHAR(16) NOT NULL,
            `T1标签` VARCHAR(16) NOT NULL,
            `T2标签` VARCHAR(16) NOT NULL,
            `SKU` TEXT,
            `面料颜色编号` VARCHAR(255),
            `面料` VARCHAR(255),
            `颜色` VARCHAR(255),
            `领星颜色` VARCHAR(255),
            `库存量米` DECIMAL(18,2) NOT NULL DEFAULT 0,
            `待到货量米` DECIMAL(18,2) NOT NULL DEFAULT 0,
            `统计类型` VARCHAR(32),
            `面料编号` VARCHAR(255),
            `颜色缩写` VARCHAR(255),
            `库存量条` DECIMAL(18,2) NOT NULL DEFAULT 0,
            `待到货量条` DECIMAL(18,2) NOT NULL DEFAULT 0,
            `用量信息缺失SPU` TEXT,
            `当月完整预估米` DECIMAL(18,2) NOT NULL DEFAULT 0,
            `当月已下单消耗米` DECIMAL(18,2) NOT NULL DEFAULT 0,
            `当月剩余预估米` DECIMAL(18,2) NOT NULL DEFAULT 0,
            `T1预估米` DECIMAL(18,2) NOT NULL DEFAULT 0,
            `T2预估米` DECIMAL(18,2) NOT NULL DEFAULT 0,
            `运营当月预估米` DECIMAL(18,2) NOT NULL DEFAULT 0,
            `运营T1预估米` DECIMAL(18,2) NOT NULL DEFAULT 0,
            `运营T2预估米` DECIMAL(18,2) NOT NULL DEFAULT 0,
            PRIMARY KEY (`id`),
            KEY `idx_scheme_a_date` (`业务日期`, `方案`),
            KEY `idx_scheme_a_fabric_color` (`面料`, `颜色`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """

    insert_sql = f"""
        INSERT INTO `{table_name}` (
            `业务日期`, `生成时间`, `方案`, `当月标签`, `T1标签`, `T2标签`,
            `SKU`, `面料颜色编号`, `面料`, `颜色`, `领星颜色`,
            `库存量米`, `待到货量米`, `统计类型`, `面料编号`, `颜色缩写`,
            `库存量条`, `待到货量条`, `用量信息缺失SPU`,
            `当月完整预估米`, `当月已下单消耗米`, `当月剩余预估米`,
            `T1预估米`, `T2预估米`, `运营当月预估米`, `运营T1预估米`, `运营T2预估米`
        ) VALUES (
            %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
            %s,%s,%s,%s,%s,%s,%s,%s
        )
    """

    values = []
    for row in business_rows:
        values.append((
            business_date,
            now,
            "A",
            m0,
            m1,
            m2,
            str(row.get("SKU") or ""),
            str(row.get("面料颜色编号") or ""),
            str(row.get("面料") or ""),
            str(row.get("颜色") or ""),
            str(row.get("领星颜色") or ""),
            float(row.get("库存量/米") or 0),
            float(row.get("待到货量/米") or 0),
            str(row.get("统计类型") or ""),
            str(row.get("面料编号") or ""),
            str(row.get("颜色缩写") or ""),
            float(row.get("库存量/条") or 0),
            float(row.get("待到货量/条") or 0),
            str(row.get("用量信息缺失SPU") or ""),
            float(row.get(f"{h0}完整预估/米") or 0),
            float(row.get(f"{h0}已下单消耗/米") or 0),
            float(row.get(f"{h0}剩余预估/米") or 0),
            float(row.get(f"{h1}预估/米") or 0),
            float(row.get(f"{h2}预估/米") or 0),
            float(row.get(f"运营{h0}预估/米") or 0),
            float(row.get(f"运营{h1}预估/米") or 0),
            float(row.get(f"运营{h2}预估/米") or 0),
        ))

    with db_cursor(dictionary=False) as cur:
        cur.execute(create_sql)
        cur.execute(
            f"DELETE FROM `{table_name}` WHERE `业务日期`=%s AND `方案`='A'",
            (business_date,),
        )
        cur.executemany(insert_sql, values)
        cur.execute(
            f"SELECT COUNT(*) FROM `{table_name}` WHERE `业务日期`=%s AND `方案`='A'",
            (business_date,),
        )
        saved_count = int(cur.fetchone()[0])
        if saved_count != len(values):
            raise RuntimeError(
                f"方案A数据库快照行数不一致：应保存 {len(values)}，实际 {saved_count}"
            )

    logger.info(
        "✓ 方案A每日快照已保存到 MySQL `%s`：%s，%d 行；月份=%s/%s/%s",
        table_name,
        business_date,
        saved_count,
        m0,
        m1,
        m2,
    )
    return saved_count


async def run(
    output_dir: Path,
    manual_mapping_path: Path,
    spu_manual_mapping_path: Path,
    write_feishu: bool = False,
    feishu_table_name: str = DEFAULT_FEISHU_TABLE_NAME,
    scheme_a_history_table: str = SCHEME_A_HISTORY_TABLE,
) -> Path:
    # 正式口径：方案B。默认参数也是B，这里显式传入，避免未来默认值变更造成静默回退。
    color_rows, total_rows, pending_rows, metrics = await final_export.build_final_rows(
        manual_mapping_path=manual_mapping_path,
        spu_manual_mapping_path=spu_manual_mapping_path,
        color_split_mode="B",
    )

    # 对照口径：方案A。只用于数据库每日快照，不写入飞书。
    a_color_rows, a_total_rows, _, _ = await final_export.build_final_rows(
        manual_mapping_path=manual_mapping_path,
        spu_manual_mapping_path=spu_manual_mapping_path,
        color_split_mode="A",
    )

    lingxing_names = await _load_lingxing_name_by_record_id()
    archive_scheme_a_to_mysql(
        a_color_rows,
        a_total_rows,
        lingxing_names,
        table_name=scheme_a_history_table,
    )

    output = export_business_workbook(
        color_rows,
        total_rows,
        pending_rows,
        metrics,
        lingxing_names,
        output_dir,
    )
    if write_feishu:
        headers, business_rows = build_business_rows(
            color_rows, total_rows, lingxing_names
        )
        await write_business_rows_to_feishu(
            headers,
            business_rows,
            table_name=feishu_table_name,
        )
    return output


def main() -> Path:
    parser = argparse.ArgumentParser(description="按业务21列表头生成最终面料预估表：B写飞书，A留MySQL")
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path(os.getenv("PROCUREMENT_EXPORT_DIR", "/opt/apps/pythondata/exports")),
    )
    parser.add_argument(
        "--manual-mapping",
        type=Path,
        default=stocking.DEFAULT_MANUAL_MAPPING_PATH,
        help="历史四字段人工映射 CSV",
    )
    parser.add_argument(
        "--spu-manual-mapping",
        type=Path,
        default=DEFAULT_SPU_MANUAL_MAPPING_PATH,
        help="SPU级人工映射 CSV；键=面料名+原始颜色编码+SPU",
    )
    parser.add_argument(
        "--write-feishu",
        action="store_true",
        help="将方案B最终业务21列表全量覆盖写入飞书面料预估明细",
    )
    parser.add_argument(
        "--feishu-table-name",
        default=DEFAULT_FEISHU_TABLE_NAME,
        help="目标飞书表名，默认面料预估明细",
    )
    parser.add_argument(
        "--scheme-a-history-table",
        default=SCHEME_A_HISTORY_TABLE,
        help="方案A每日快照MySQL表名，默认面料预估方案A历史",
    )
    args = parser.parse_args()
    return asyncio.run(
        run(
            output_dir=args.output_dir,
            manual_mapping_path=args.manual_mapping,
            spu_manual_mapping_path=args.spu_manual_mapping,
            write_feishu=args.write_feishu,
            feishu_table_name=args.feishu_table_name,
            scheme_a_history_table=args.scheme_a_history_table,
        )
    )


if __name__ == "__main__":
    main()
