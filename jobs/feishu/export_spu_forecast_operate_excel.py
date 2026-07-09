#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""
导出 SPU 未来4个月预估下单量（运营版）。

字段固定：
SPU、年份、季节、店铺数、库存、生产中、25年1月~26年6月历史销量、
未来4个月系统/运营预估合计、差异、差异率、26年7月~26年10月系统/运营预估。

运行：
  cd /opt/apps/pythondata
  ./venv/bin/python -m jobs.feishu.export_spu_forecast_operate_excel
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

project_root = Path(__file__).resolve().parent.parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

from common import get_logger
from common.database import db_cursor

logger = get_logger('export_spu_forecast_operate_excel')

EXCLUDED_SHOPS = {
    'TEMU半托管-A店', 'TEMU半托管-C店', 'TEMU半托管-M店',
    'TEMU半托管-P店', 'TEMU半托管-V店', 'TEMU半托管-本土店-R店',
    'TK本土店-1店', 'TK跨境店-2店', 'CY-US', 'DX-US', 'MT-CA'
}

HISTORY_MONTHS: List[Tuple[date, str]] = [
    (date(2025, m, 1), f'25年{m}月历史销量') for m in range(1, 13)
] + [
    (date(2026, m, 1), f'26年{m}月历史销量') for m in range(1, 7)
]

FORECAST_MONTHS: List[Tuple[date, str]] = [
    (date(2026, 7, 1), '26年7月'),
    (date(2026, 8, 1), '26年8月'),
    (date(2026, 9, 1), '26年9月'),
    (date(2026, 10, 1), '26年10月'),
]


def _fetch_all(sql: str, params: Tuple[Any, ...] = ()) -> List[Dict[str, Any]]:
    with db_cursor(dictionary=True) as cursor:
        cursor.execute(sql, params)
        return list(cursor.fetchall())


def _table_exists(table_name: str) -> bool:
    rows = _fetch_all(
        """
        SELECT COUNT(*) AS cnt
        FROM information_schema.TABLES
        WHERE TABLE_SCHEMA = DATABASE()
          AND TABLE_NAME = %s
        """,
        (table_name,),
    )
    return bool(rows and rows[0].get('cnt'))


def _get_columns(table_name: str) -> set:
    if not _table_exists(table_name):
        return set()
    rows = _fetch_all(
        """
        SELECT COLUMN_NAME
        FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA = DATABASE()
          AND TABLE_NAME = %s
        """,
        (table_name,),
    )
    return {r['COLUMN_NAME'] for r in rows}


def _first_day(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return date(value.year, value.month, 1)
    if isinstance(value, date):
        return date(value.year, value.month, 1)
    try:
        d = datetime.strptime(str(value)[:10], '%Y-%m-%d')
        return date(d.year, d.month, 1)
    except Exception:
        return None


def _extract_spu_from_sku(sku: str) -> str:
    if not sku:
        return ''
    sku = re.sub(r'\d+(?:PSC|PCS)', '', str(sku), flags=re.IGNORECASE)
    sku = re.sub(r'-+', '-', sku).strip('-')
    idx = sku.find('-')
    return sku[:idx] if idx > 0 else sku


def _infer_year_season_from_spu(spu: str) -> Tuple[str, str]:
    """兜底从 SPU 文本推断年份和季节。识别不到返回空。"""
    text = str(spu or '').upper()
    year = ''
    season = ''
    m = re.search(r'(20)?(25|26|27|28)\s*(SS|AW|FW|SP|SU)?', text)
    if m:
        year = '20' + m.group(2)
        raw = m.group(3) or ''
        if raw == 'SS':
            season = '春夏'
        elif raw in ('AW', 'FW'):
            season = '秋冬'
        elif raw == 'SP':
            season = '春季'
        elif raw == 'SU':
            season = '夏季'
    return year, season


def read_spu_attrs() -> Dict[str, Dict[str, str]]:
    """读取 SPU 年份/季节。

    优先尝试常见表和字段；如果不存在则后续用 SPU 文本兜底推断。
    """
    candidates = [
        ('SPU基础信息表', 'SPU'),
        ('spu基础信息表', 'SPU'),
        ('商品企划表', 'SPU'),
        ('lxpm_feishu_product_info', 'spu'),
        ('lxpm_product_category_snapshot', 'spu'),
    ]
    year_names = ['年份', 'year', '开发年份', '季节年份']
    season_names = ['季节', 'season', '销售季节', '开发季节']

    result: Dict[str, Dict[str, str]] = {}
    for table, spu_col in candidates:
        if not _table_exists(table):
            continue
        cols = _get_columns(table)
        if spu_col not in cols:
            continue
        year_col = next((c for c in year_names if c in cols), None)
        season_col = next((c for c in season_names if c in cols), None)
        if not year_col and not season_col:
            continue
        sql = f"""
            SELECT `{spu_col}` AS SPU,
                   {f'`{year_col}`' if year_col else "''"} AS 年份,
                   {f'`{season_col}`' if season_col else "''"} AS 季节
            FROM `{table}`
            WHERE `{spu_col}` IS NOT NULL AND `{spu_col}` != ''
        """
        rows = _fetch_all(sql)
        for r in rows:
            spu = (r.get('SPU') or '').strip()
            if not spu:
                continue
            result.setdefault(spu, {})
            if r.get('年份') and not result[spu].get('年份'):
                result[spu]['年份'] = str(r.get('年份') or '').strip()
            if r.get('季节') and not result[spu].get('季节'):
                result[spu]['季节'] = str(r.get('季节') or '').strip()
        if result:
            logger.info(f'已从 {table} 读取 SPU 年份/季节：{len(result)} 个')
            return result

    logger.warning('未找到可用的 SPU 年份/季节来源，将从 SPU 文本兜底推断')
    return result


def read_forecast_data() -> List[Dict[str, Any]]:
    if not _table_exists('预测对比表'):
        raise RuntimeError('预测对比表不存在，请先运行采购预测流水线')
    month_dates = [d.strftime('%Y-%m-%d') for d, _ in FORECAST_MONTHS]
    placeholders = ','.join(['%s'] * len(month_dates))
    rows = _fetch_all(
        f"""
        SELECT SPU, 店铺, 统计日期,
               SUM(系统预测销量) AS 系统预测销量,
               SUM(运营预计下单量) AS 运营预计下单量
        FROM `预测对比表`
        WHERE 统计日期 IN ({placeholders})
          AND SPU IS NOT NULL AND SPU != ''
          AND 店铺 IS NOT NULL AND 店铺 != ''
        GROUP BY SPU, 店铺, 统计日期
        """,
        tuple(month_dates),
    )

    label_by_date = {d.strftime('%Y-%m-%d'): label for d, label in FORECAST_MONTHS}
    by_spu: Dict[str, Dict[str, Any]] = defaultdict(lambda: {'SPU': '', '店铺数': set()})
    for r in rows:
        spu = (r.get('SPU') or '').strip()
        shop = (r.get('店铺') or '').strip()
        if not spu or not shop or shop in EXCLUDED_SHOPS:
            continue
        d = _first_day(r.get('统计日期'))
        date_key = d.strftime('%Y-%m-%d') if d else ''
        label = label_by_date.get(date_key)
        if not label:
            continue
        sys_qty = int(r.get('系统预测销量') or 0)
        op_qty = int(r.get('运营预计下单量') or 0)

        row = by_spu[spu]
        row['SPU'] = spu
        row['店铺数'].add(shop)
        row[f'{label}_系统预估销量'] = row.get(f'{label}_系统预估销量', 0) + sys_qty
        row[f'{label}_运营预估下单量'] = row.get(f'{label}_运营预估下单量', 0) + op_qty

    result = []
    for spu, row in by_spu.items():
        out = dict(row)
        out['店铺数'] = len(row.get('店铺数') or [])
        sys_total = sum(int(out.get(f'{label}_系统预估销量', 0) or 0) for _, label in FORECAST_MONTHS)
        op_total = sum(int(out.get(f'{label}_运营预估下单量', 0) or 0) for _, label in FORECAST_MONTHS)
        out['未来4个月系统预估合计'] = sys_total
        out['未来4个月运营预估合计'] = op_total
        out['运营-系统差异'] = op_total - sys_total
        out['运营/系统差异率'] = round((op_total - sys_total) / sys_total, 4) if sys_total else None
        result.append(out)
    result.sort(key=lambda x: (-(x.get('未来4个月运营预估合计') or 0), x.get('SPU') or ''))
    logger.info(f'预测数据读取完成：{len(result)} 个 SPU')
    return result


def read_history_sales() -> Dict[str, Dict[str, int]]:
    table = '销量统计_msku月度'
    if not _table_exists(table):
        logger.warning('销量统计_msku月度 不存在，历史销量为空')
        return {}
    cols = _get_columns(table)
    has_spu = 'SPU' in cols

    start = HISTORY_MONTHS[0][0].strftime('%Y-%m-%d')
    end = date(2026, 7, 1).strftime('%Y-%m-%d')
    sql = f"""
        SELECT SKU, 店铺, 统计日期, SUM(销量) AS 销量 {', SPU' if has_spu else ''}
        FROM `{table}`
        WHERE 统计日期 >= %s AND 统计日期 < %s
          AND SKU IS NOT NULL AND SKU != ''
          AND 店铺 IS NOT NULL AND 店铺 != ''
        GROUP BY SKU, 店铺, 统计日期 {', SPU' if has_spu else ''}
    """
    rows = _fetch_all(sql, (start, end))
    label_by_date = {d.strftime('%Y-%m-%d'): label for d, label in HISTORY_MONTHS}
    result: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in rows:
        sku = (r.get('SKU') or '').strip()
        shop = (r.get('店铺') or '').strip()
        if not sku or not shop or shop in EXCLUDED_SHOPS:
            continue
        spu = (r.get('SPU') or '').strip() if has_spu else ''
        spu = spu or _extract_spu_from_sku(sku)
        d = _first_day(r.get('统计日期'))
        if not spu or not d:
            continue
        label = label_by_date.get(d.strftime('%Y-%m-%d'))
        if label:
            result[spu][label] += int(r.get('销量') or 0)
    logger.info(f'历史销量读取完成：{len(result)} 个 SPU')
    return {k: dict(v) for k, v in result.items()}


def read_inventory_metrics() -> Dict[str, Dict[str, int]]:
    """库存=FBA可售+本地可用量；生产中=FBA在途+本地待到货。"""
    result: Dict[str, Dict[str, int]] = defaultdict(lambda: {'库存': 0, '生产中': 0})
    if _table_exists('库存预估表'):
        rows = _fetch_all("""
            SELECT SKU, SPU, 店铺, 库存状态, SUM(数量) AS 数量
            FROM `库存预估表`
            WHERE SKU IS NOT NULL AND SKU != ''
              AND 店铺 IS NOT NULL AND 店铺 != ''
            GROUP BY SKU, SPU, 店铺, 库存状态
        """)
        for r in rows:
            shop = (r.get('店铺') or '').strip()
            if not shop or shop in EXCLUDED_SHOPS:
                continue
            sku = (r.get('SKU') or '').strip()
            spu = (r.get('SPU') or '').strip() or _extract_spu_from_sku(sku)
            status = (r.get('库存状态') or '').strip()
            qty = int(r.get('数量') or 0)
            if not spu:
                continue
            if status in ('FBA可售', '本地可用量'):
                result[spu]['库存'] += qty
            elif status in ('FBA在途', '本地待到货'):
                result[spu]['生产中'] += qty
        logger.info(f'库存预估表读取完成：{len(result)} 个 SPU')
        return {k: dict(v) for k, v in result.items()}

    logger.warning('库存预估表不存在，库存/生产中为空')
    return {}


def merge_rows() -> List[Dict[str, Any]]:
    forecast_rows = read_forecast_data()
    history = read_history_sales()
    inventory = read_inventory_metrics()
    attrs = read_spu_attrs()

    all_spus = {r['SPU'] for r in forecast_rows} | set(history.keys()) | set(inventory.keys())
    forecast_by_spu = {r['SPU']: r for r in forecast_rows}

    output = []
    for spu in sorted(all_spus):
        base = dict(forecast_by_spu.get(spu, {'SPU': spu, '店铺数': 0}))
        attr = attrs.get(spu, {})
        year = attr.get('年份') or ''
        season = attr.get('季节') or ''
        if not year or not season:
            infer_year, infer_season = _infer_year_season_from_spu(spu)
            year = year or infer_year
            season = season or infer_season
        inv = inventory.get(spu, {})

        row = {
            'SPU': spu,
            '年份': year,
            '季节': season,
            '店铺数': int(base.get('店铺数') or 0),
            '库存': int(inv.get('库存') or 0),
            '生产中': int(inv.get('生产中') or 0),
        }
        for _, label in HISTORY_MONTHS:
            row[label] = int(history.get(spu, {}).get(label, 0) or 0)
        for key in ['未来4个月系统预估合计', '未来4个月运营预估合计', '运营-系统差异', '运营/系统差异率']:
            row[key] = base.get(key, None if key.endswith('差异率') else 0)
        for _, label in FORECAST_MONTHS:
            row[f'{label}_系统预估销量'] = int(base.get(f'{label}_系统预估销量', 0) or 0)
            row[f'{label}_运营预估下单量'] = int(base.get(f'{label}_运营预估下单量', 0) or 0)
        output.append(row)

    output.sort(key=lambda x: (-(x.get('未来4个月运营预估合计') or 0), x.get('SPU') or ''))
    return output


def ordered_columns() -> List[str]:
    cols = ['SPU', '年份', '季节', '店铺数', '库存', '生产中']
    cols += [label for _, label in HISTORY_MONTHS]
    cols += ['未来4个月系统预估合计', '未来4个月运营预估合计', '运营-系统差异', '运营/系统差异率']
    for _, label in FORECAST_MONTHS:
        cols += [f'{label}_系统预估销量', f'{label}_运营预估下单量']
    return cols


def write_sheet(wb: Workbook, rows: List[Dict[str, Any]]) -> None:
    ws = wb.active
    ws.title = 'SPU未来4个月预估下单量'
    columns = ordered_columns()
    ws.append(columns)
    for r in rows:
        ws.append([r.get(c, '') for c in columns])

    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(bold=True, color='FFFFFF')
    thin = Side(style='thin', color='D9E2F3')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = 'G2'
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if isinstance(cell.value, (int, float)):
                header = str(ws.cell(1, cell.column).value)
                cell.number_format = '0.00%' if '差异率' in header else '#,##0'

    if rows:
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        table = Table(displayName='tbl_spu_forecast_operate', ref=ref)
        table.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True, showColumnStripes=False)
        try:
            ws.add_table(table)
        except Exception:
            pass

    for idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(len(str(c.value or '')) for c in col_cells)
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 22)


def write_note_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet('字段说明')
    notes = [
        ('历史销量', '固定读取 2025-01 至 2026-06，按 SKU 解析 SPU 后汇总。'),
        ('未来4个月', '固定读取 2026-07 至 2026-10。'),
        ('库存', '库存预估表中 FBA可售 + 本地可用量。'),
        ('生产中', '库存预估表中 FBA在途 + 本地待到货，即采购中未回货/待入库口径。'),
        ('系统预估销量', '预测对比表.系统预测销量。'),
        ('运营预估下单量', '预测对比表.运营预计下单量。'),
        ('运营-系统差异', '未来4个月运营预估合计 - 未来4个月系统预估合计。'),
        ('运营/系统差异率', '运营-系统差异 / 未来4个月系统预估合计；系统为0时为空。'),
        ('年份/季节', '优先从常见 SPU 属性表读取；没有来源时尝试从 SPU 文本推断，识别不到则为空。'),
    ]
    ws.append(['字段', '说明'])
    for n in notes:
        ws.append(n)
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 100
    for c in ws[1]:
        c.fill = PatternFill('solid', fgColor='1F4E78')
        c.font = Font(bold=True, color='FFFFFF')
    for row in ws.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical='top')


def export_excel(output_path: str) -> str:
    rows = merge_rows()
    wb = Workbook()
    write_sheet(wb, rows)
    write_note_sheet(wb)
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    logger.info(f'✓ 已生成：{out}')
    logger.info(f'SPU行数：{len(rows)}')
    return str(out)


def parse_args():
    parser = argparse.ArgumentParser(description='导出 SPU未来4个月预估下单量（运营版）')
    parser.add_argument('--output', default=None, help='输出路径')
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output = args.output or f"exports/SPU未来4个月预估下单量_运营版_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    export_excel(output)


if __name__ == '__main__':
    main()
