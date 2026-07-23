#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""导出最新版“面料-颜色预计下单表”。

颜色名称严格来自 A2023/B2024 颜色编制表，展示格式为“中文名称｜颜色体系”。
"""
from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from common import get_logger
from common.database import db_cursor
from jobs.feishu.color_mapping_catalog import ColorMappingCatalog

logger = get_logger("export_fabric_color_order_forecast")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
PENDING_FILL = PatternFill("solid", fgColor="FFF2CC")
CONFLICT_FILL = PatternFill("solid", fgColor="F4CCCC")
THIN = Side(style="thin", color="D9E1F2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def read_rows() -> List[Dict[str, Any]]:
    with db_cursor() as cursor:
        cursor.execute("""
            SELECT 统计类型, 面料, 面料编号, 颜色体系, 颜色缩写, 颜色汇总代码,
                   面料颜色编号, 库存归属状态,
                   `库存量/条`, `库存量/米`, `待到货量/条`, `待到货量/米`,
                   `当月已下单消耗/米`, `当月完整预估/米`, `当月剩余预估/米`, 当月月份,
                   `T+1月预估/米`, `T+1月份`, `T+2月预估/米`, `T+2月份`,
                   `T+3月预估/米`, `T+3月份`,
                   `运营当月预估/米`, `运营T+1月预估/米`, `运营T+2月预估/米`,
                   `运营T+3月预估/米`, 用量信息缺失SPU
            FROM `面料预估表`
            ORDER BY 统计类型, 面料, 颜色体系, 颜色缩写
        """)
        return list(cursor.fetchall())


def style_header(ws, row: int, headers: Sequence[str]) -> None:
    for index, header in enumerate(headers, 1):
        cell = ws.cell(row, index, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def autosize(ws, minimum: int = 10, maximum: int = 38) -> None:
    for cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in cells)
        ws.column_dimensions[get_column_letter(cells[0].column)].width = min(
            max(length + 2, minimum), maximum
        )


def month_labels(rows: Sequence[Dict[str, Any]]) -> list[str]:
    row = next((item for item in rows if item.get("统计类型") == "带颜色"), rows[0] if rows else {})
    return [
        str(row.get("当月月份") or "T月"),
        str(row.get("T+1月份") or "T+1月"),
        str(row.get("T+2月份") or "T+2月"),
        str(row.get("T+3月份") or "T+3月"),
    ]


def enriched_row(row: Dict[str, Any], catalog: ColorMappingCatalog) -> Dict[str, Any]:
    system = str(row.get("颜色体系") or "")
    code = str(row.get("颜色缩写") or "")
    if row.get("统计类型") == "带颜色":
        info = catalog.describe(system, code)
    else:
        info = {
            "中文颜色名称": "",
            "颜色显示名称": "",
            "颜色映射状态": "总量行",
            "A2023中文候选": "",
            "B2024中文候选": "",
        }
    return {**row, **info}


def build_color_sheet(
    workbook: Workbook,
    rows: Sequence[Dict[str, Any]],
    labels: Sequence[str],
    title: str = "面料-颜色预计下单",
) -> None:
    ws = workbook.active if title == "面料-颜色预计下单" else workbook.create_sheet(title)
    ws.title = title
    headers = [
        "面料", "面料编号", "颜色体系", "颜色代码", "中文颜色名称", "颜色名称+体系",
        "A2023中文候选", "B2024中文候选", "颜色映射状态", "颜色汇总代码",
        "面料颜色编号", "库存归属状态", "库存量/条", "库存量/米",
        "待到货量/条", "待到货量/米", f"{labels[0]}已下单消耗/米",
        f"{labels[0]}完整预估/米", f"{labels[0]}剩余预估/米",
        f"{labels[1]}预估/米", f"{labels[2]}预估/米", f"{labels[3]}预估/米",
        f"运营{labels[0]}预估/米", f"运营{labels[1]}预估/米",
        f"运营{labels[2]}预估/米", f"运营{labels[3]}预估/米", "用量信息缺失SPU",
    ]
    style_header(ws, 1, headers)
    keys = [
        "面料", "面料编号", "颜色体系", "颜色缩写", "中文颜色名称", "颜色显示名称",
        "A2023中文候选", "B2024中文候选", "颜色映射状态", "颜色汇总代码",
        "面料颜色编号", "库存归属状态", "库存量/条", "库存量/米",
        "待到货量/条", "待到货量/米", "当月已下单消耗/米", "当月完整预估/米",
        "当月剩余预估/米", "T+1月预估/米", "T+2月预估/米", "T+3月预估/米",
        "运营当月预估/米", "运营T+1月预估/米", "运营T+2月预估/米",
        "运营T+3月预估/米", "用量信息缺失SPU",
    ]
    for row_index, row in enumerate(rows, 2):
        for column, key in enumerate(keys, 1):
            value = row.get(key, "")
            cell = ws.cell(row_index, column, value)
            cell.border = BORDER
            cell.alignment = RIGHT if isinstance(value, (int, float)) else LEFT
            if isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"
        if row.get("颜色体系") == "待定":
            for cell in ws[row_index]:
                cell.fill = PENDING_FILL
        if str(row.get("库存归属状态") or "").startswith("跨颜色体系冲突"):
            for cell in ws[row_index]:
                cell.fill = CONFLICT_FILL
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize(ws)


def build_total_sheet(workbook: Workbook, rows: Sequence[Dict[str, Any]], labels: Sequence[str]) -> None:
    ws = workbook.create_sheet("面料总量")
    headers = [
        "面料", "面料编号", "库存量/条", "库存量/米", "待到货量/条", "待到货量/米",
        f"{labels[0]}已下单消耗/米", f"{labels[0]}完整预估/米", f"{labels[0]}剩余预估/米",
        f"{labels[1]}预估/米", f"{labels[2]}预估/米", f"{labels[3]}预估/米",
        f"运营{labels[0]}预估/米", f"运营{labels[1]}预估/米",
        f"运营{labels[2]}预估/米", f"运营{labels[3]}预估/米", "用量信息缺失SPU",
    ]
    style_header(ws, 1, headers)
    keys = [
        "面料", "面料编号", "库存量/条", "库存量/米", "待到货量/条", "待到货量/米",
        "当月已下单消耗/米", "当月完整预估/米", "当月剩余预估/米",
        "T+1月预估/米", "T+2月预估/米", "T+3月预估/米",
        "运营当月预估/米", "运营T+1月预估/米", "运营T+2月预估/米",
        "运营T+3月预估/米", "用量信息缺失SPU",
    ]
    for row_index, row in enumerate(rows, 2):
        for column, key in enumerate(keys, 1):
            value = row.get(key, "")
            cell = ws.cell(row_index, column, value)
            cell.border = BORDER
            cell.alignment = RIGHT if isinstance(value, (int, float)) else LEFT
            if isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize(ws)


def main() -> Path:
    raw_rows = read_rows()
    if not raw_rows:
        raise RuntimeError("面料预估表为空，请先运行采购建议与面料预估任务")
    catalog = ColorMappingCatalog.from_runtime()
    rows = [enriched_row(row, catalog) for row in raw_rows]
    labels = month_labels(rows)
    color_rows = [row for row in rows if row.get("统计类型") == "带颜色"]
    total_rows = [row for row in rows if row.get("统计类型") == "总量"]
    pending_rows = [row for row in color_rows if row.get("颜色体系") == "待定"]

    workbook = Workbook()
    build_color_sheet(workbook, color_rows, labels)
    build_total_sheet(workbook, total_rows, labels)
    if pending_rows:
        build_color_sheet(workbook, pending_rows, labels, title="待定颜色")

    export_dir = Path(os.getenv("PROCUREMENT_EXPORT_DIR", "/opt/apps/pythondata/exports"))
    export_dir.mkdir(parents=True, exist_ok=True)
    output = export_dir / f"面料-颜色预计下单表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    workbook.save(output)
    logger.info(
        f"✓ 导出完成：{output}；面料颜色 {len(color_rows)} 行，"
        f"面料总量 {len(total_rows)} 行，待定颜色 {len(pending_rows)} 行"
    )
    return output


if __name__ == "__main__":
    main()
