#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""
导出所有 SPU 未来4个月预估下单量 Excel。

包含：
1. SPU汇总：按 SPU 汇总所有店铺，展示未来4个月系统预估、运营预估、过往销量。
2. SPU店铺明细：按 SPU+店铺 展示未来4个月系统预估、运营预估、过往销量。
3. 算法说明：说明系统预估、运营预估、过往销量字段的来源和口径。

数据来源：
- 预测对比表：SPU+店铺+月份，包含系统预测销量、运营预计下单量。
- 销量统计_msku月度：历史销量，按 SKU 解析 SPU 后汇总。

运行：
  cd /opt/apps/pythondata
  ./venv/bin/python -m jobs.feishu.export_spu_forecast_4m_excel
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

project_root = Path(__file__).resolve().parent.parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

from common import get_logger
from common.database import db_cursor

logger = get_logger('export_spu_forecast_4m_excel')

EXCLUDED_SHOPS = {
    'TEMU半托管-A店', 'TEMU半托管-C店', 'TEMU半托管-M店',
    'TEMU半托管-P店', 'TEMU半托管-V店', 'TEMU半托管-本土店-R店',
    'TK本土店-1店', 'TK跨境店-2店', 'CY-US', 'DX-US', 'MT-CA'
}


def _fetch_all(sql: str, params: Tuple[Any, ...] = ()) -> List[Dict[str, Any]]:
    with db_cursor(dictionary=True) as cursor:
        cursor.execute(sql, params)
        return list(cursor.fetchall())


def _table_exists(table_name: str) -> bool:
    rows = _fetch_all(
        """
        SELECT COUNT(*) AS cnt
        FROM information_schema.TABLES
        WHERE TABLE_SCHEMA = DATABASE()
          AND TABLE_NAME = %s
        """,
        (table_name,),
    )
    return bool(rows and rows[0].get('cnt'))


def _get_columns(table_name: str) -> set:
    if not _table_exists(table_name):
        return set()
    rows = _fetch_all(
        """
        SELECT COLUMN_NAME
        FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA = DATABASE()
          AND TABLE_NAME = %s
        """,
        (table_name,),
    )
    return {r['COLUMN_NAME'] for r in rows}


def _first_day(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return date(value.year, value.month, 1)
    if isinstance(value, date):
        return date(value.year, value.month, 1)
    s = str(value)[:10]
    try:
        d = datetime.strptime(s, '%Y-%m-%d')
        return date(d.year, d.month, 1)
    except Exception:
        return None


def _month_label(d: date) -> str:
    return f"{str(d.year)[-2:]}年{d.month}月"


def _extract_spu_from_sku(sku: str) -> str:
    if not sku:
        return ''
    sku = re.sub(r'\d+(?:PSC|PCS)', '', str(sku), flags=re.IGNORECASE)
    sku = re.sub(r'-+', '-', sku).strip('-')
    idx = sku.find('-')
    return sku[:idx] if idx > 0 else sku


def _month_add(d: date, delta: int) -> date:
    y = d.year
    m = d.month + delta
    while m > 12:
        m -= 12
        y += 1
    while m < 1:
        m += 12
        y -= 1
    return date(y, m, 1)


def read_forecast_months() -> List[Tuple[date, str]]:
    """优先以预测对比表中本月起的月份为准，取未来4个月。"""
    today = datetime.now().date()
    current_month = date(today.year, today.month, 1)
    if not _table_exists('预测对比表'):
        raise RuntimeError('预测对比表不存在，请先运行 generate_forecast_comparison.py')

    rows = _fetch_all(
        """
        SELECT DISTINCT 统计日期, 月份
        FROM `预测对比表`
        WHERE 统计日期 >= %s
        ORDER BY 统计日期
        LIMIT 4
        """,
        (current_month.strftime('%Y-%m-%d'),),
    )
    months: List[Tuple[date, str]] = []
    for r in rows:
        d = _first_day(r.get('统计日期'))
        lbl = (r.get('月份') or '').strip()
        if d:
            months.append((d, lbl or _month_label(d)))

    if len(months) < 4:
        existing = {d for d, _ in months}
        d = current_month
        while len(months) < 4:
            if d not in existing:
                months.append((d, _month_label(d)))
            d = _month_add(d, 1)

    logger.info(f"未来4个月：{[lbl for _, lbl in months]}")
    return months[:4]


def read_forecast_data(months: List[Tuple[date, str]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """读取预测对比表，返回 SPU汇总 和 SPU+店铺明细。"""
    month_dates = [d.strftime('%Y-%m-%d') for d, _ in months]
    placeholders = ','.join(['%s'] * len(month_dates))
    rows = _fetch_all(
        f"""
        SELECT SPU, 店铺, 月份, 统计日期,
               SUM(系统预测销量) AS 系统预测销量,
               SUM(运营预计下单量) AS 运营预计下单量
        FROM `预测对比表`
        WHERE 统计日期 IN ({placeholders})
          AND SPU IS NOT NULL AND SPU != ''
          AND 店铺 IS NOT NULL AND 店铺 != ''
        GROUP BY SPU, 店铺, 月份, 统计日期
        """,
        tuple(month_dates),
    )

    by_spu: Dict[str, Dict[str, Any]] = defaultdict(lambda: {'SPU': '', '店铺数': set()})
    by_spu_shop: Dict[Tuple[str, str], Dict[str, Any]] = defaultdict(lambda: {'SPU': '', '店铺': ''})

    label_by_date = {d.strftime('%Y-%m-%d'): lbl for d, lbl in months}
    for r in rows:
        spu = (r.get('SPU') or '').strip()
        shop = (r.get('店铺') or '').strip()
        if not spu or not shop or shop in EXCLUDED_SHOPS:
            continue
        d = _first_day(r.get('统计日期'))
        date_key = d.strftime('%Y-%m-%d') if d else ''
        lbl = label_by_date.get(date_key, (r.get('月份') or '').strip())
        sys_qty = int(r.get('系统预测销量') or 0)
        op_qty = int(r.get('运营预计下单量') or 0)

        srow = by_spu[spu]
        srow['SPU'] = spu
        srow['店铺数'].add(shop)
        srow[f'{lbl}_系统预估销量'] = srow.get(f'{lbl}_系统预估销量', 0) + sys_qty
        srow[f'{lbl}_运营预估下单量'] = srow.get(f'{lbl}_运营预估下单量', 0) + op_qty

        drow = by_spu_shop[(spu, shop)]
        drow['SPU'] = spu
        drow['店铺'] = shop
        drow[f'{lbl}_系统预估销量'] = drow.get(f'{lbl}_系统预估销量', 0) + sys_qty
        drow[f'{lbl}_运营预估下单量'] = drow.get(f'{lbl}_运营预估下单量', 0) + op_qty

    summary_rows = []
    for spu, row in by_spu.items():
        out = dict(row)
        out['店铺数'] = len(row.get('店铺数') or [])
        sys_total = 0
        op_total = 0
        for _, lbl in months:
            sys_total += int(out.get(f'{lbl}_系统预估销量', 0) or 0)
            op_total += int(out.get(f'{lbl}_运营预估下单量', 0) or 0)
        out['未来4个月系统预估合计'] = sys_total
        out['未来4个月运营预估合计'] = op_total
        out['运营-系统差异'] = op_total - sys_total
        out['运营/系统差异率'] = round((op_total - sys_total) / sys_total, 4) if sys_total else None
        summary_rows.append(out)

    detail_rows = []
    for _, row in by_spu_shop.items():
        out = dict(row)
        sys_total = 0
        op_total = 0
        for _, lbl in months:
            sys_total += int(out.get(f'{lbl}_系统预估销量', 0) or 0)
            op_total += int(out.get(f'{lbl}_运营预估下单量', 0) or 0)
        out['未来4个月系统预估合计'] = sys_total
        out['未来4个月运营预估合计'] = op_total
        out['运营-系统差异'] = op_total - sys_total
        out['运营/系统差异率'] = round((op_total - sys_total) / sys_total, 4) if sys_total else None
        detail_rows.append(out)

    summary_rows.sort(key=lambda x: (-(x.get('未来4个月运营预估合计') or 0), x.get('SPU') or ''))
    detail_rows.sort(key=lambda x: (x.get('SPU') or '', x.get('店铺') or ''))
    return summary_rows, detail_rows


def read_sales_history(months: List[Tuple[date, str]]) -> Tuple[Dict[str, Dict[str, int]], Dict[Tuple[str, str], Dict[str, int]], List[str]]:
    """读取过往销量，返回 SPU 汇总、SPU+店铺明细、历史月份标签。"""
    table = '销量统计_msku月度'
    if not _table_exists(table):
        logger.warning('销量统计_msku月度 不存在，过往销量为空')
        return {}, {}, []

    current_month = months[0][0]
    hist_months = [_month_add(current_month, -i) for i in range(1, 13)]
    hist_months = sorted(hist_months)
    start = hist_months[0]
    end = current_month
    cols = _get_columns(table)
    has_spu = 'SPU' in cols

    sql = f"""
        SELECT SKU, 店铺, 统计日期, SUM(销量) AS 销量 {', SPU' if has_spu else ''}
        FROM `{table}`
        WHERE 统计日期 >= %s AND 统计日期 < %s
          AND SKU IS NOT NULL AND SKU != ''
          AND 店铺 IS NOT NULL AND 店铺 != ''
        GROUP BY SKU, 店铺, 统计日期 {', SPU' if has_spu else ''}
    """
    rows = _fetch_all(sql, (start.strftime('%Y-%m-%d'), end.strftime('%Y-%m-%d')))

    by_spu: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    by_spu_shop: Dict[Tuple[str, str], Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    month_labels = [_month_label(d) for d in hist_months]

    for r in rows:
        sku = (r.get('SKU') or '').strip()
        shop = (r.get('店铺') or '').strip()
        if not sku or not shop or shop in EXCLUDED_SHOPS:
            continue
        spu = (r.get('SPU') or '').strip() if has_spu else ''
        spu = spu or _extract_spu_from_sku(sku)
        if not spu:
            continue
        d = _first_day(r.get('统计日期'))
        if not d:
            continue
        lbl = _month_label(d)
        qty = int(r.get('销量') or 0)
        by_spu[spu][f'{lbl}_历史销量'] += qty
        by_spu_shop[(spu, shop)][f'{lbl}_历史销量'] += qty

    return {k: dict(v) for k, v in by_spu.items()}, {k: dict(v) for k, v in by_spu_shop.items()}, month_labels


def attach_history(rows: List[Dict[str, Any]], hist_map: Dict[Any, Dict[str, int]], month_labels: List[str], key_fields: List[str]) -> None:
    last3 = month_labels[-3:]
    last6 = month_labels[-6:]
    last12 = month_labels[-12:]
    for row in rows:
        if key_fields == ['SPU']:
            key = row.get('SPU')
        else:
            key = (row.get('SPU'), row.get('店铺'))
        hist = hist_map.get(key, {})
        for lbl in month_labels:
            row[f'{lbl}_历史销量'] = int(hist.get(f'{lbl}_历史销量', 0) or 0)
        row['近3月销量合计'] = sum(int(hist.get(f'{lbl}_历史销量', 0) or 0) for lbl in last3)
        row['近6月销量合计'] = sum(int(hist.get(f'{lbl}_历史销量', 0) or 0) for lbl in last6)
        row['近12月销量合计'] = sum(int(hist.get(f'{lbl}_历史销量', 0) or 0) for lbl in last12)
        row['近3月月均销量'] = round(row['近3月销量合计'] / 3, 2)
        row['近6月月均销量'] = round(row['近6月销量合计'] / 6, 2)
        row['近12月月均销量'] = round(row['近12月销量合计'] / 12, 2)


def ordered_columns(months: List[Tuple[date, str]], hist_labels: List[str], detail: bool = False) -> List[str]:
    fixed = ['SPU'] + (['店铺'] if detail else ['店铺数'])
    hist_summary = ['近3月销量合计', '近6月销量合计', '近12月销量合计', '近3月月均销量', '近6月月均销量', '近12月月均销量']
    forecast_total = ['未来4个月系统预估合计', '未来4个月运营预估合计', '运营-系统差异', '运营/系统差异率']
    forecast_month_cols = []
    for _, lbl in months:
        forecast_month_cols.extend([f'{lbl}_系统预估销量', f'{lbl}_运营预估下单量'])
    hist_cols = [f'{lbl}_历史销量' for lbl in hist_labels]
    return fixed + hist_summary + forecast_total + forecast_month_cols + hist_cols


def write_sheet(wb: Workbook, title: str, rows: List[Dict[str, Any]], columns: List[str], freeze: str = 'A2') -> None:
    ws = wb.create_sheet(title)
    ws.freeze_panes = freeze
    ws.append(columns)
    for r in rows:
        ws.append([r.get(c, '') for c in columns])

    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(bold=True, color='FFFFFF')
    thin = Side(style='thin', color='D9E2F3')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if isinstance(cell.value, (int, float)):
                if '差异率' in str(ws.cell(1, cell.column).value):
                    cell.number_format = '0.00%'
                else:
                    cell.number_format = '#,##0.00' if isinstance(cell.value, float) and not float(cell.value).is_integer() else '#,##0'

    ws.auto_filter.ref = ws.dimensions
    if rows:
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        table_name = re.sub(r'[^A-Za-z0-9_]', '_', f'tbl_{title}')[:250]
        if not re.match(r'^[A-Za-z_]', table_name):
            table_name = 'T_' + table_name
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True, showColumnStripes=False)
        try:
            ws.add_table(table)
        except Exception:
            pass

    for idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(len(str(c.value or '')) for c in col_cells)
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 28)


def write_algorithm_sheet(wb: Workbook, months: List[Tuple[date, str]], hist_labels: List[str]) -> None:
    ws = wb.create_sheet('算法说明')
    rows = [
        ('报表口径', '所有 SPU 未来4个月预估下单量，包含 SPU 汇总和 SPU+店铺明细。'),
        ('未来4个月', '、'.join(lbl for _, lbl in months)),
        ('过往销量月份', '、'.join(hist_labels)),
        ('系统预估销量来源', '预测对比表.系统预测销量。该表由 generate_forecast_comparison.py 生成。'),
        ('系统预估销量算法', '从 销量统计_msku月度 读取历史销量，按 SKU 调用 forecast_sales_improved.compute_forecast_for_shop 计算未来月份预测，再聚合到 SPU+店铺+月份。算法输入包含近3个月销量、去年同期及前后缓冲月份，并加载 SPU 季节映射。'),
        ('运营预估下单量来源', '运营预计下单表，经 generate_forecast_comparison.py 按 SKU 解析 SPU 后聚合到 SPU+店铺+月份，写入 预测对比表.运营预计下单量。'),
        ('过往销量来源', '销量统计_msku月度，按 SKU 解析 SPU，并按 SPU 或 SPU+店铺聚合。'),
        ('近3/6/12月销量', '以当前预测起始月的前一个月为最近完整月，向前滚动 3/6/12 个月汇总。'),
        ('运营-系统差异', '未来4个月运营预估合计 - 未来4个月系统预估合计。'),
        ('运营/系统差异率', '(未来4个月运营预估合计 - 未来4个月系统预估合计) / 未来4个月系统预估合计；系统为0时为空。'),
        ('注意', '本报表是预估口径，不扣库存和待到货；库存扣减后的建议下单量仍以采购建议报告为准。'),
    ]
    ws.append(['项目', '说明'])
    for row in rows:
        ws.append(row)
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 120
    for cell in ws[1]:
        cell.fill = PatternFill('solid', fgColor='1F4E78')
        cell.font = Font(bold=True, color='FFFFFF')
    for row in ws.iter_rows(min_row=2):
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical='top')


def export_excel(output_path: str) -> str:
    months = read_forecast_months()
    summary_rows, detail_rows = read_forecast_data(months)
    hist_spu, hist_spu_shop, hist_labels = read_sales_history(months)
    attach_history(summary_rows, hist_spu, hist_labels, ['SPU'])
    attach_history(detail_rows, hist_spu_shop, hist_labels, ['SPU', '店铺'])

    wb = Workbook()
    # 删除默认sheet后按顺序创建
    default = wb.active
    wb.remove(default)
    write_sheet(wb, 'SPU汇总', summary_rows, ordered_columns(months, hist_labels, detail=False), freeze='A2')
    write_sheet(wb, 'SPU店铺明细', detail_rows, ordered_columns(months, hist_labels, detail=True), freeze='A2')
    write_algorithm_sheet(wb, months, hist_labels)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    logger.info(f"✓ 已生成：{output}")
    logger.info(f"SPU汇总 {len(summary_rows)} 行，SPU店铺明细 {len(detail_rows)} 行")
    return str(output)


def parse_args():
    parser = argparse.ArgumentParser(description='导出所有SPU未来4个月预估下单量Excel')
    parser.add_argument('--output', default=None, help='输出路径，默认 exports/SPU未来4个月预估下单量_YYYYMMDD_HHMM.xlsx')
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output = args.output or f"exports/SPU未来4个月预估下单量_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    export_excel(output)


if __name__ == '__main__':
    main()
