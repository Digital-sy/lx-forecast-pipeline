#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""生成最终版“面料-颜色预计下单表”。

最终颜色口径统一使用飞书当前有效颜色：

1. SPU 级人工映射（面料名+原始颜色编码+SPU）优先；
2. 未命中时复用 ``fabric_color_stocking`` 的确定性规则；
3. 再未命中时复用历史四字段人工映射；
4. 模糊候选绝不进入最终预计下单表，只进入“待确认颜色”；
5. A2023/B2024 仅来自 SKU 自身显式主数据，任何飞书颜色或人工映射都不得反推颜色体系；
6. 颜色库存按“面料编号+领星新颜色缩写”精确匹配，同一飞书颜色只扣一次库存。

需求口径沿用现有生产面料预估：建议下单量按 SKU 预测颜色/尺码比例分摊，
米数=数量×单件用量×单件损耗；单耗缺失仍沿用现有生产表的平均单耗兜底并显式标记。

颜色拆分口径（方案B）：只要 SPU 在《面料核价表》中配置了目标面料，该面料就按当前 SKU
解析出的最终飞书颜色拆分，不再要求该面料是该 SPU 单件用量最大的“主面料”。

本模块只生成 Excel，不回写 MySQL/飞书。
"""
from __future__ import annotations

import argparse
import asyncio
import json
import os
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Mapping, MutableMapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from common import get_logger
from jobs.feishu import fabric_color_stocking as stocking
from jobs.feishu import generate_fabric_forecast as fabric_base
from jobs.feishu import generate_fabric_forecast_color_system as color_system
from jobs.feishu.color_system_resolver import ColorSystemResolver, normalize_sku
from jobs.feishu.fabric_color_stocking_spu import (
    DEFAULT_SPU_MANUAL_MAPPING_PATH,
    SpuManualMappingCatalog,
    load_spu_manual_mapping_catalog,
)

logger = get_logger("export_fabric_color_order_forecast_final")

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
PENDING_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="D9E1F2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _meters(value: float) -> float:
    return round(float(value or 0), 2)


def _quantity(value: float) -> int | float:
    value = round(float(value or 0), 4)
    return int(value) if value.is_integer() else value


def _empty_bucket() -> dict[str, Any]:
    return {
        "purchase_m": 0.0,
        "sys_month_m": [0.0, 0.0, 0.0, 0.0],
        "op_month_m": [0.0, 0.0, 0.0, 0.0],
        "缺失SPU": set(),
        "methods": set(),
        "systems": set(),
        "codes": set(),
        "spus": set(),
        "skus": set(),
    }


def _month_delta(stat_date: Any, current_date: datetime) -> int | None:
    try:
        if isinstance(stat_date, str):
            value = datetime.strptime(stat_date[:10], "%Y-%m-%d")
        else:
            value = stat_date
        for delta in range(4):
            year, month = color_system.add_months(
                current_date.year, current_date.month, delta
            )
            if value.year == year and value.month == month:
                return delta
    except Exception:
        pass
    return None


def _fallback_forecast(
    sku: str,
    forecast_by_sku: Mapping[str, stocking.ForecastSku],
    resolver: ColorSystemResolver,
    snapshot_by_sku: Mapping[str, Mapping[str, Any]],
    governance_catalog: stocking.ColorMappingCatalog,
) -> stocking.ForecastSku:
    normalized = normalize_sku(sku)
    if normalized in forecast_by_sku:
        return forecast_by_sku[normalized]

    identity = resolver.resolve(normalized)
    snapshot = snapshot_by_sku.get(normalized, {})
    product_name = str(snapshot.get("product_name") or "")
    color_name = stocking.parse_lingxing_color(product_name)
    if not color_name and identity.color_system in stocking.SUPPORTED_SYSTEMS:
        entry = governance_catalog.lookup(identity.color_system, identity.color_code)
        color_name = entry.chinese if entry else ""

    return stocking.ForecastSku(
        sku=normalized,
        spu=str(snapshot.get("spu") or identity.spu or "").strip(),
        product_name=product_name,
        color_code=str(identity.color_code or "").strip(),
        color_name=str(color_name or "").strip(),
        color_system=str(identity.color_system or stocking.UNKNOWN_SYSTEM).strip()
        or stocking.UNKNOWN_SYSTEM,
        forecast_qty=0,
    )


def _resolve_final_color(
    forecast: stocking.ForecastSku,
    fabric_name: str,
    index: stocking.CatalogIndex,
    governance_catalog: stocking.ColorMappingCatalog,
    manual_catalog: stocking.ManualMappingCatalog,
    spu_catalog: SpuManualMappingCatalog,
) -> stocking.MatchDecision:
    """SPU 人工规则是优先级 0；无规则时完全复用既有确定性匹配。"""
    spu_decision = spu_catalog.decision(forecast, fabric_name, index)
    if spu_decision is not None:
        return spu_decision
    return stocking.match_catalog_row(
        forecast,
        fabric_name,
        index,
        governance_catalog,
        manual_catalog=manual_catalog,
    )


def _style_header(ws: Any, headers: Sequence[str]) -> None:
    for column, header in enumerate(headers, 1):
        cell = ws.cell(1, column, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _autosize(ws: Any, minimum: int = 10, maximum: int = 45) -> None:
    for cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in cells)
        ws.column_dimensions[get_column_letter(cells[0].column)].width = min(
            max(length + 2, minimum), maximum
        )


def _write_rows(
    ws: Any,
    headers: Sequence[str],
    rows: Sequence[Mapping[str, Any]],
    pending: bool = False,
) -> None:
    _style_header(ws, headers)
    for row_index, row in enumerate(rows, 2):
        for column, header in enumerate(headers, 1):
            value = row.get(header, "")
            cell = ws.cell(row_index, column, value)
            cell.border = BORDER
            cell.alignment = RIGHT if isinstance(value, (int, float)) else LEFT
            if isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"
            if pending:
                cell.fill = PENDING_FILL
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autosize(ws)


async def build_final_rows(
    manual_mapping_path: Path,
    spu_manual_mapping_path: Path,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    current_date = datetime.now()
    target_fabrics = tuple(stocking.TARGET_FABRICS)
    target_set = set(target_fabrics)

    governance_catalog = stocking.ColorMappingCatalog.from_runtime(strict=True)
    catalog_rows, catalog_source_audit = await stocking.load_catalog_from_feishu(
        base_token=os.getenv(
            "FABRIC_COLOR_CATALOG_BASE_TOKEN", stocking.DEFAULT_BASE_TOKEN
        ),
        table_id=os.getenv(
            "FABRIC_COLOR_CATALOG_TABLE_ID", stocking.DEFAULT_CATALOG_TABLE_ID
        ),
        view_id=os.getenv(
            "FABRIC_COLOR_CATALOG_VIEW_ID", stocking.DEFAULT_CATALOG_VIEW_ID
        ),
    )
    target_catalog = [row for row in catalog_rows if row.fabric_name in target_set]
    index = stocking.CatalogIndex(
        target_catalog,
        source_record_count=catalog_source_audit.get("source_record_count"),
    )

    manual_catalog = stocking.load_manual_mapping_catalog(manual_mapping_path)
    spu_catalog = load_spu_manual_mapping_catalog(spu_manual_mapping_path)

    forecasts, forecast_audit = stocking.load_forecast_skus(governance_catalog)
    forecast_by_sku = {row.sku: row for row in forecasts}
    snapshot_rows = stocking._load_snapshot_rows()
    snapshot_by_sku, _ = stocking._snapshot_index(snapshot_rows)
    resolver = ColorSystemResolver(snapshot_rows)

    fabric_params = fabric_base.get_fabric_params()
    fabric_usage = fabric_base.get_fabric_price_data()
    purchase_order_data = fabric_base.get_purchase_order_data()
    system_forecast_data = fabric_base.get_system_forecast_data()
    suggest_data = color_system.get_suggest_order_data_color(
        resolver, current_date=current_date
    )
    operation_forecast_data = fabric_base.get_forecast_order_data()
    effective_qty, _ = color_system._effective_sku_quantities(
        resolver,
        system_forecast_data,
        suggest_data,
        current_date,
    )

    legacy_merge_map = fabric_base.get_fabric_color_merge_mapping()
    inventory_data, pending_data = fabric_base.get_inventory_data(legacy_merge_map)
    inv_by_fabric, pend_by_fabric = fabric_base.get_inventory_by_fabric(
        inventory_data,
        pending_data,
        fabric_params,
    )

    usage_by_spu: MutableMapping[str, list[tuple[str, Mapping[str, Any]]]] = defaultdict(list)
    for (spu, fabric_name), usage_data in fabric_usage.items():
        if fabric_name in target_set:
            usage_by_spu[str(spu).strip()].append((fabric_name, usage_data))

    total_agg: MutableMapping[str, dict[str, Any]] = defaultdict(_empty_bucket)
    color_agg: MutableMapping[tuple[str, str, str], dict[str, Any]] = defaultdict(_empty_bucket)
    pending_agg: MutableMapping[tuple[str, str, str, str], dict[str, Any]] = defaultdict(_empty_bucket)

    match_method_counts: MutableMapping[str, int] = defaultdict(int)
    unmatched_reason_counts: MutableMapping[str, int] = defaultdict(int)

    def add_usage(sku: str, qty: int, target: str, delta: int | None = None) -> None:
        if qty <= 0:
            return
        forecast = _fallback_forecast(
            sku,
            forecast_by_sku,
            resolver,
            snapshot_by_sku,
            governance_catalog,
        )
        spu = str(forecast.spu or "").strip()
        if not spu:
            return

        for fabric_name, usage_data in usage_by_spu.get(spu, ()):
            meters, missing = fabric_base._calc_usage_meters(
                qty,
                usage_data.get("单件用量"),
                usage_data.get("单件损耗"),
                fabric_name,
                spu,
                fabric_usage,
            )
            if meters <= 0:
                continue

            total_bucket = total_agg[fabric_name]
            if target == "purchase":
                total_bucket["purchase_m"] += meters
            else:
                total_bucket[target][delta] += meters
            if missing:
                total_bucket["缺失SPU"].add(spu)

            # 方案B：只要 SPU 配置了该面料，就按当前 SKU 的最终飞书颜色拆分。
            # 不再以“单件用量最大”作为是否允许拆色的条件。
            decision = _resolve_final_color(
                forecast,
                fabric_name,
                index,
                governance_catalog,
                manual_catalog,
                spu_catalog,
            )
            if decision.row:
                row = decision.row
                bucket = color_agg[row.identity]
                bucket["methods"].add(decision.method)
                bucket["systems"].add(forecast.color_system)
                bucket["codes"].add(forecast.color_code)
                bucket["spus"].add(spu)
                bucket["skus"].add(forecast.sku)
                if target == "purchase":
                    bucket["purchase_m"] += meters
                else:
                    bucket[target][delta] += meters
                if missing:
                    bucket["缺失SPU"].add(spu)
                match_method_counts[decision.method] += 1
                continue

            reason = str(decision.reason or decision.reason_code or "待确认")
            key = (
                fabric_name,
                forecast.color_system,
                forecast.color_code,
                reason,
            )
            bucket = pending_agg[key]
            bucket["systems"].add(forecast.color_system)
            bucket["codes"].add(forecast.color_code)
            bucket["spus"].add(spu)
            bucket["skus"].add(forecast.sku)
            if target == "purchase":
                bucket["purchase_m"] += meters
            else:
                bucket[target][delta] += meters
            if missing:
                bucket["缺失SPU"].add(spu)
            unmatched_reason_counts[reason] += 1

    for sku, qty in purchase_order_data.items():
        add_usage(sku, int(qty), "purchase")

    for (sku, delta), qty in effective_qty.items():
        add_usage(sku, int(qty), "sys_month_m", delta)

    for (sku, stat_date), qty in operation_forecast_data.items():
        delta = _month_delta(stat_date, current_date)
        if delta is not None:
            add_usage(sku, int(qty), "op_month_m", delta)

    month_labels = [label for _, label in color_system.future_months(current_date, 4)]

    def demand_fields(bucket: Mapping[str, Any]) -> dict[str, Any]:
        purchase_m = _meters(bucket["purchase_m"])
        system = [_meters(value) for value in bucket["sys_month_m"]]
        operation = [_meters(value) for value in bucket["op_month_m"]]
        return {
            f"{month_labels[0]}已下单消耗/米": purchase_m,
            f"{month_labels[0]}完整预估/米": system[0],
            f"{month_labels[0]}剩余预估/米": _meters(max(0.0, system[0] - purchase_m)),
            f"{month_labels[1]}预估/米": system[1],
            f"{month_labels[2]}预估/米": system[2],
            f"{month_labels[3]}预估/米": system[3],
            f"运营{month_labels[0]}预估/米": operation[0],
            f"运营{month_labels[1]}预估/米": operation[1],
            f"运营{month_labels[2]}预估/米": operation[2],
            f"运营{month_labels[3]}预估/米": operation[3],
            "未来4月系统预估合计/米": _meters(sum(system)),
            "用量信息缺失SPU": "、".join(sorted(bucket["缺失SPU"])),
        }

    final_rows: list[dict[str, Any]] = []
    for row in target_catalog:
        bucket = color_agg.get(row.identity, _empty_bucket())
        fabric_name, final_color, lingxing_color = row.identity
        params = fabric_params.get(fabric_name, {})
        fabric_code = str(params.get("面料编号") or "").strip()
        meters_per_roll = float(params.get("米数每条") or 0)
        inventory_key = f"{fabric_code}-{lingxing_color}".upper() if fabric_code and lingxing_color else ""
        inventory_rolls = int(inventory_data.get(inventory_key, 0) or 0) if inventory_key else 0
        pending_rolls = int(pending_data.get(inventory_key, 0) or 0) if inventory_key else 0
        final_rows.append({
            "面料": fabric_name,
            "颜色": final_color,
            "领星颜色": lingxing_color,
            "库存量/条": inventory_rolls,
            "库存量/米": _meters(inventory_rolls * meters_per_roll),
            "待到货量/条": pending_rolls,
            "待到货量/米": _meters(pending_rolls * meters_per_roll),
            **demand_fields(bucket),
            "匹配方式": "、".join(sorted(bucket["methods"])),
            "颜色体系": "、".join(sorted(bucket["systems"])),
            "原颜色编码": "、".join(sorted(bucket["codes"])),
            "SPU": "、".join(sorted(bucket["spus"])),
            "SKU": "、".join(sorted(bucket["skus"])),
        })

    final_rows.sort(key=lambda row: (
        target_fabrics.index(str(row["面料"])) if str(row["面料"]) in target_fabrics else 999,
        str(row["颜色"]),
    ))

    total_rows: list[dict[str, Any]] = []
    for fabric_name in target_fabrics:
        if fabric_name not in total_agg:
            continue
        bucket = total_agg[fabric_name]
        params = fabric_params.get(fabric_name, {})
        meters_per_roll = float(params.get("米数每条") or 0)
        inventory_rolls = int(inv_by_fabric.get(fabric_name, 0) or 0)
        pending_rolls = int(pend_by_fabric.get(fabric_name, 0) or 0)
        total_rows.append({
            "面料": fabric_name,
            "库存量/条": inventory_rolls,
            "库存量/米": _meters(inventory_rolls * meters_per_roll),
            "待到货量/条": pending_rolls,
            "待到货量/米": _meters(pending_rolls * meters_per_roll),
            **demand_fields(bucket),
        })

    pending_rows: list[dict[str, Any]] = []
    for (fabric_name, system, code, reason), bucket in pending_agg.items():
        row = {
            "面料": fabric_name,
            "颜色体系": system,
            "原颜色编码": code,
            "待确认原因": reason,
            **demand_fields(bucket),
            "SPU": "、".join(sorted(bucket["spus"])),
            "SKU": "、".join(sorted(bucket["skus"])),
        }
        pending_rows.append(row)
    pending_rows.sort(key=lambda row: (str(row["面料"]), str(row["颜色体系"]), str(row["原颜色编码"])))

    audit = {
        "生成时间": current_date.isoformat(timespec="seconds"),
        "颜色拆分口径": "方案B：SPU配置目标面料即按SKU最终飞书颜色拆分，不限制主面料",
        "飞书颜色目录审计": catalog_source_audit,
        "SKU主数据审计": forecast_audit,
        "最终颜色行数": len(final_rows),
        "面料总量行数": len(total_rows),
        "待确认颜色行数": len(pending_rows),
        "SPU人工规则数": len(spu_catalog.rows),
        "匹配方式计数": dict(sorted(match_method_counts.items())),
        "待确认原因计数": dict(sorted(unmatched_reason_counts.items())),
    }
    return final_rows, total_rows, pending_rows, audit


def export_workbook(
    final_rows: Sequence[Mapping[str, Any]],
    total_rows: Sequence[Mapping[str, Any]],
    pending_rows: Sequence[Mapping[str, Any]],
    audit: Mapping[str, Any],
    output_dir: Path,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    output = output_dir / f"最终面料颜色预计下单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "最终面料颜色预计下单"
    final_headers = list(final_rows[0].keys()) if final_rows else ["面料", "颜色", "领星颜色"]
    _write_rows(ws, final_headers, final_rows)

    ws_total = wb.create_sheet("面料总量")
    total_headers = list(total_rows[0].keys()) if total_rows else ["面料"]
    _write_rows(ws_total, total_headers, total_rows)

    ws_pending = wb.create_sheet("待确认颜色")
    pending_headers = list(pending_rows[0].keys()) if pending_rows else ["面料", "颜色体系", "原颜色编码", "待确认原因"]
    _write_rows(ws_pending, pending_headers, pending_rows, pending=True)

    ws_audit = wb.create_sheet("审计")
    _style_header(ws_audit, ["项目", "值"])
    for row_index, (key, value) in enumerate(audit.items(), 2):
        ws_audit.cell(row_index, 1, key).border = BORDER
        ws_audit.cell(row_index, 1).alignment = LEFT
        if isinstance(value, (dict, list, tuple, set)):
            value = json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
        ws_audit.cell(row_index, 2, value).border = BORDER
        ws_audit.cell(row_index, 2).alignment = LEFT
    _autosize(ws_audit)

    wb.save(output)
    logger.info("最终面料颜色预计下单表已生成: %s", output)
    return output


async def run(
    output_dir: Path,
    manual_mapping_path: Path,
    spu_manual_mapping_path: Path,
) -> Path:
    rows = await build_final_rows(
        manual_mapping_path=manual_mapping_path,
        spu_manual_mapping_path=spu_manual_mapping_path,
    )
    return export_workbook(*rows, output_dir=output_dir)


def main() -> None:
    parser = argparse.ArgumentParser(description="生成最终版面料-颜色预计下单表（方案B）")
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path(os.getenv("FABRIC_COLOR_FORECAST_OUTPUT_DIR", "/opt/apps/pythondata/exports")),
    )
    parser.add_argument(
        "--manual-mapping",
        type=Path,
        default=Path(os.getenv(
            "FABRIC_COLOR_MANUAL_MAPPING_PATH",
            "/opt/apps/pythondata/shared_config/fabric_color_manual_mapping.csv",
        )),
    )
    parser.add_argument(
        "--spu-manual-mapping",
        type=Path,
        default=Path(os.getenv(
            "FABRIC_COLOR_SPU_MANUAL_MAPPING_PATH",
            str(DEFAULT_SPU_MANUAL_MAPPING_PATH),
        )),
    )
    args = parser.parse_args()
    asyncio.run(run(
        output_dir=args.output_dir,
        manual_mapping_path=args.manual_mapping,
        spu_manual_mapping_path=args.spu_manual_mapping,
    ))


if __name__ == "__main__":
    main()
