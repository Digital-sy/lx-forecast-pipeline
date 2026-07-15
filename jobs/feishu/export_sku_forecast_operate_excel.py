#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""
导出 SKU 未来4个月预估下单量（运营版）。

排除 SKU（忽略大小写）：
- X00 开头
- B0 开头
- AMA 开头

历史销量固定：2025-01 ~ 2026-06
未来预测固定：2026-07 ~ 2026-10
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

project_root = Path(__file__).resolve().parent.parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

from common import get_logger
from common.database import db_cursor

logger = get_logger('export_sku_forecast_operate_excel')

EXCLUDED_SHOPS = {
    'TEMU半托管-A店', 'TEMU半托管-C店', 'TEMU半托管-M店',
    'TEMU半托管-P店', 'TEMU半托管-V店', 'TEMU半托管-本土店-R店',
    'TK本土店-1店', 'TK跨境店-2店', 'CY-US', 'DX-US', 'MT-CA'
}

EXCLUDED_SKU_PREFIXES = ('X00', 'B0', 'AMA')

HISTORY_MONTHS: List[Tuple[date, str]] = [
    (date(2025, m, 1), f'25年{m}月历史销量') for m in range(1, 13)
] + [
    (date(2026, m, 1), f'26年{m}月历史销量') for m in range(1, 7)
]

FORECAST_MONTHS: List[Tuple[date, str]] = [
    (date(2026, 7, 1), '26年7月'),
    (date(2026, 8, 1), '26年8月'),
    (date(2026, 9, 1), '26年9月'),
    (date(2026, 10, 1), '26年10月'),
]


def fetch_all(sql: str, params: Tuple[Any, ...] = ()) -> List[Dict[str, Any]]:
    with db_cursor(dictionary=True) as cursor:
        cursor.execute(sql, params)
        return list(cursor.fetchall())


def table_exists(table_name: str) -> bool:
    rows = fetch_all(
        """
        SELECT COUNT(*) AS cnt
        FROM information_schema.TABLES
        WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = %s
        """,
        (table_name,),
    )
    return bool(rows and rows[0].get('cnt'))


def get_columns(table_name: str) -> set:
    if not table_exists(table_name):
        return set()
    rows = fetch_all(
        """
        SELECT COLUMN_NAME
        FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = %s
        """,
        (table_name,),
    )
    return {r['COLUMN_NAME'] for r in rows}


def first_day(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return date(value.year, value.month, 1)
    if isinstance(value, date):
        return date(value.year, value.month, 1)
    try:
        d = datetime.strptime(str(value)[:10], '%Y-%m-%d')
        return date(d.year, d.month, 1)
    except Exception:
        return None


def clean_sku(value: Any) -> str:
    return str(value or '').strip()


def should_exclude_sku(sku: str) -> bool:
    upper = clean_sku(sku).upper()
    return not upper or upper.startswith(EXCLUDED_SKU_PREFIXES)


def extract_spu(sku: str) -> str:
    sku = re.sub(r'\d+(?:PSC|PCS)', '', clean_sku(sku), flags=re.IGNORECASE)
    sku = re.sub(r'-+', '-', sku).strip('-')
    return sku.split('-', 1)[0] if '-' in sku else sku


def infer_year_season(spu: str) -> Tuple[str, str]:
    text = str(spu or '').upper()
    m = re.search(r'(20)?(25|26|27|28)\s*(SS|AW|FW|SP|SU)?', text)
    if not m:
        return '', ''
    year = '20' + m.group(2)
    season_map = {'SS': '春夏', 'AW': '秋冬', 'FW': '秋冬', 'SP': '春季', 'SU': '夏季'}
    return year, season_map.get(m.group(3) or '', '')


def read_spu_attrs() -> Dict[str, Dict[str, str]]:
    candidates = [
        ('SPU基础信息表', 'SPU'),
        ('spu基础信息表', 'SPU'),
        ('商品企划表', 'SPU'),
        ('lxpm_feishu_product_info', 'spu'),
        ('lxpm_product_category_snapshot', 'spu'),
    ]
    year_names = ['年份', 'year', '开发年份', '季节年份']
    season_names = ['季节', 'season', '销售季节', '开发季节']

    result: Dict[str, Dict[str, str]] = {}
    for table, spu_col in candidates:
        if not table_exists(table):
            continue
        cols = get_columns(table)
        if spu_col not in cols:
            continue
        year_col = next((c for c in year_names if c in cols), None)
        season_col = next((c for c in season_names if c in cols), None)
        if not year_col and not season_col:
            continue
        rows = fetch_all(f"""
            SELECT `{spu_col}` AS SPU,
                   {f'`{year_col}`' if year_col else "''"} AS 年份,
                   {f'`{season_col}`' if season_col else "''"} AS 季节
            FROM `{table}`
            WHERE `{spu_col}` IS NOT NULL AND `{spu_col}` != ''
        """)
        for r in rows:
            spu = str(r.get('SPU') or '').strip()
            if spu:
                result[spu] = {
                    '年份': str(r.get('年份') or '').strip(),
                    '季节': str(r.get('季节') or '').strip(),
                }
        if result:
            logger.info(f'已从 {table} 读取年份/季节：{len(result)} 个SPU')
            break
    return result


def read_system_forecast() -> Dict[str, Dict[str, Any]]:
    if not table_exists('预测对比表_SKU'):
        raise RuntimeError('预测对比表_SKU 不存在，请先运行 generate_forecast_comparison.py')
    dates = [d.strftime('%Y-%m-%d') for d, _ in FORECAST_MONTHS]
    placeholders = ','.join(['%s'] * len(dates))
    rows = fetch_all(f"""
        SELECT SKU, SPU, 店铺, 统计日期, SUM(系统预测销量) AS 系统预测销量
        FROM `预测对比表_SKU`
        WHERE 统计日期 IN ({placeholders})
          AND SKU IS NOT NULL AND SKU != ''
          AND 店铺 IS NOT NULL AND 店铺 != ''
        GROUP BY SKU, SPU, 店铺, 统计日期
    """, tuple(dates))

    label_by_date = {d.strftime('%Y-%m-%d'): label for d, label in FORECAST_MONTHS}
    result: Dict[str, Dict[str, Any]] = defaultdict(lambda: {'SPU': '', '店铺数': set()})
    for r in rows:
        sku = clean_sku(r.get('SKU'))
        shop = str(r.get('店铺') or '').strip()
        if should_exclude_sku(sku) or not shop or shop in EXCLUDED_SHOPS:
            continue
        spu = str(r.get('SPU') or '').strip() or extract_spu(sku)
        d = first_day(r.get('统计日期'))
        if not d:
            continue
        label = label_by_date.get(d.strftime('%Y-%m-%d'))
        if not label:
            continue
        result[sku]['SPU'] = spu
        result[sku]['店铺数'].add(shop)
        result[sku][f'{label}_系统预估销量'] = result[sku].get(f'{label}_系统预估销量', 0) + int(r.get('系统预测销量') or 0)
    return result


def read_operation_forecast() -> Dict[str, Dict[str, int]]:
    if not table_exists('运营预计下单表'):
        logger.warning('运营预计下单表不存在，运营预估全部为0')
        return {}
    dates = [d.strftime('%Y-%m-%d') for d, _ in FORECAST_MONTHS]
    placeholders = ','.join(['%s'] * len(dates))
    rows = fetch_all(f"""
        SELECT SKU, 店铺, 统计日期, SUM(预计下单量) AS 预计下单量
        FROM `运营预计下单表`
        WHERE 统计日期 IN ({placeholders})
          AND SKU IS NOT NULL AND SKU != ''
          AND 店铺 IS NOT NULL AND 店铺 != ''
        GROUP BY SKU, 店铺, 统计日期
    """, tuple(dates))

    label_by_date = {d.strftime('%Y-%m-%d'): label for d, label in FORECAST_MONTHS}
    result: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in rows:
        sku = clean_sku(r.get('SKU'))
        shop = str(r.get('店铺') or '').strip()
        if should_exclude_sku(sku) or not shop or shop in EXCLUDED_SHOPS:
            continue
        d = first_day(r.get('统计日期'))
        if not d:
            continue
        label = label_by_date.get(d.strftime('%Y-%m-%d'))
        if label:
            result[sku][f'{label}_运营预估下单量'] += int(r.get('预计下单量') or 0)
    return {k: dict(v) for k, v in result.items()}


def read_history_sales() -> Dict[str, Dict[str, int]]:
    table = '销量统计_msku月度'
    if not table_exists(table):
        return {}
    rows = fetch_all("""
        SELECT SKU, 店铺, 统计日期, SUM(销量) AS 销量
        FROM `销量统计_msku月度`
        WHERE 统计日期 >= '2025-01-01' AND 统计日期 < '2026-07-01'
          AND SKU IS NOT NULL AND SKU != ''
          AND 店铺 IS NOT NULL AND 店铺 != ''
        GROUP BY SKU, 店铺, 统计日期
    """)
    label_by_date = {d.strftime('%Y-%m-%d'): label for d, label in HISTORY_MONTHS}
    result: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in rows:
        sku = clean_sku(r.get('SKU'))
        shop = str(r.get('店铺') or '').strip()
        if should_exclude_sku(sku) or not shop or shop in EXCLUDED_SHOPS:
            continue
        d = first_day(r.get('统计日期'))
        if not d:
            continue
        label = label_by_date.get(d.strftime('%Y-%m-%d'))
        if label:
            result[sku][label] += int(r.get('销量') or 0)
    return {k: dict(v) for k, v in result.items()}


def read_inventory() -> Dict[str, Dict[str, int]]:
    result: Dict[str, Dict[str, int]] = defaultdict(lambda: {'库存': 0, '生产中': 0})
    if not table_exists('库存预估表'):
        logger.warning('库存预估表不存在，库存和生产中为空')
        return {}
    rows = fetch_all("""
        SELECT SKU, 店铺, 库存状态, SUM(数量) AS 数量
        FROM `库存预估表`
        WHERE SKU IS NOT NULL AND SKU != ''
          AND 店铺 IS NOT NULL AND 店铺 != ''
        GROUP BY SKU, 店铺, 库存状态
    """)
    for r in rows:
        sku = clean_sku(r.get('SKU'))
        shop = str(r.get('店铺') or '').strip()
        if should_exclude_sku(sku) or not shop or shop in EXCLUDED_SHOPS:
            continue
        status = str(r.get('库存状态') or '').strip()
        qty = int(r.get('数量') or 0)
        if status in ('FBA可售', '本地可用量'):
            result[sku]['库存'] += qty
        elif status in ('FBA在途', '本地待到货'):
            result[sku]['生产中'] += qty
    return {k: dict(v) for k, v in result.items()}


def build_rows() -> List[Dict[str, Any]]:
    system = read_system_forecast()
    operation = read_operation_forecast()
    history = read_history_sales()
    inventory = read_inventory()
    attrs = read_spu_attrs()

    all_skus = set(system) | set(operation) | set(history) | set(inventory)
    rows = []
    for sku in sorted(all_skus):
        if should_exclude_sku(sku):
            continue
        sys_row = system.get(sku, {})
        spu = str(sys_row.get('SPU') or '').strip() or extract_spu(sku)
        attr = attrs.get(spu, {})
        year = attr.get('年份') or ''
        season = attr.get('季节') or ''
        if not year or not season:
            inferred_year, inferred_season = infer_year_season(spu)
            year = year or inferred_year
            season = season or inferred_season

        row: Dict[str, Any] = {
            'SKU': sku,
            'SPU': spu,
            '年份': year,
            '季节': season,
            '店铺数': len(sys_row.get('店铺数') or []),
            '库存': int(inventory.get(sku, {}).get('库存', 0) or 0),
            '生产中': int(inventory.get(sku, {}).get('生产中', 0) or 0),
        }
        for _, label in HISTORY_MONTHS:
            row[label] = int(history.get(sku, {}).get(label, 0) or 0)

        sys_total = 0
        op_total = 0
        for _, label in FORECAST_MONTHS:
            sys_qty = int(sys_row.get(f'{label}_系统预估销量', 0) or 0)
            op_qty = int(operation.get(sku, {}).get(f'{label}_运营预估下单量', 0) or 0)
            row[f'{label}_系统预估销量'] = sys_qty
            row[f'{label}_运营预估下单量'] = op_qty
            sys_total += sys_qty
            op_total += op_qty
        row['未来4个月系统预估合计'] = sys_total
        row['未来4个月运营预估合计'] = op_total
        row['运营-系统差异'] = op_total - sys_total
        row['运营/系统差异率'] = round((op_total - sys_total) / sys_total, 4) if sys_total else None
        rows.append(row)

    rows.sort(key=lambda x: (-(x.get('未来4个月运营预估合计') or 0), x.get('SKU') or ''))
    logger.info(f'SKU行数：{len(rows)}')
    return rows


def columns() -> List[str]:
    cols = ['SKU', 'SPU', '年份', '季节', '店铺数', '库存', '生产中']
    cols += [label for _, label in HISTORY_MONTHS]
    cols += ['未来4个月系统预估合计', '未来4个月运营预估合计', '运营-系统差异', '运营/系统差异率']
    for _, label in FORECAST_MONTHS:
        cols += [f'{label}_系统预估销量', f'{label}_运营预估下单量']
    return cols


def write_excel(output_path: str) -> str:
    rows = build_rows()
    wb = Workbook()
    ws = wb.active
    ws.title = 'SKU未来4个月预估下单量'
    headers = columns()
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, '') for h in headers])

    fill = PatternFill('solid', fgColor='1F4E78')
    font = Font(bold=True, color='FFFFFF')
    thin = Side(style='thin', color='D9E2F3')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.freeze_panes = 'H2'
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if isinstance(cell.value, (int, float)):
                header = str(ws.cell(1, cell.column).value)
                cell.number_format = '0.00%' if '差异率' in header else '#,##0'

    if rows:
        ref = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'
        table = Table(displayName='tbl_sku_forecast_operate', ref=ref)
        table.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True, showColumnStripes=False)
        try:
            ws.add_table(table)
        except Exception:
            pass

    for idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(len(str(c.value or '')) for c in col_cells)
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 22)

    note = wb.create_sheet('字段说明')
    notes = [
        ('过滤规则', '排除以 X00、B0、AMA 开头的 SKU，忽略大小写。'),
        ('历史销量', '固定读取 2025-01 至 2026-06。'),
        ('系统预测', '来自预测对比表_SKU。'),
        ('运营预估', '来自运营预计下单表，按 SKU 汇总。'),
        ('库存', 'FBA可售 + 本地可用量。'),
        ('生产中', 'FBA在途 + 本地待到货，即采购中未回货/待入库。'),
    ]
    note.append(['字段', '说明'])
    for item in notes:
        note.append(item)
    note.column_dimensions['A'].width = 20
    note.column_dimensions['B'].width = 100

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    logger.info(f'✓ 已生成：{out}')
    return str(out)


def parse_args():
    parser = argparse.ArgumentParser(description='导出SKU未来4个月预估下单量（运营版）')
    parser.add_argument('--output', default=None)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output = args.output or f"exports/SKU未来4个月预估下单量_运营版_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    write_excel(output)


if __name__ == '__main__':
    main()
