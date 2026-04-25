#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""
采购建议报告导出 Excel
从数据库读取 建议下单量表 + 面料预计用量表，导出为两个 Sheet 的 Excel 文件

运行方式：
    cd /opt/apps/pythondata
    source venv/bin/activate
    python -m jobs.feishu.export_procurement_excel

输出路径：/opt/apps/pythondata/exports/采购建议报告_YYYYMMDD.xlsx
"""

import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

project_root = Path(__file__).resolve().parent.parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

from common import get_logger
from common.database import db_cursor

logger = get_logger('export_procurement_excel')

# ── 颜色常量 ──────────────────────────────────────────────────────────────
C_HEADER_BG     = 'FF2C2C2A'
C_MONTH_OP      = 'FFD6EFD8'
C_MONTH_SYS     = 'FFE8E4F8'
C_MONTH_HDR_OP  = 'FF0F6E56'
C_MONTH_HDR_SYS = 'FF3C3489'
C_WHITE         = 'FFFFFFFF'
C_LIGHT_GRAY    = 'FFF5F5F5'
C_CUSTOM        = 'FFEEEDFE'
C_STOCK         = 'FFE1F5EE'
C_RED_TEXT      = 'FFA32D2D'
C_AMBER_TEXT    = 'FF854F0B'

thin = Side(style='thin',   color='FFCCCCCC')
med  = Side(style='medium', color='FF888888')
bdr     = Border(left=thin, right=thin, top=thin, bottom=thin)
bdr_mb  = Border(left=med,  right=thin, top=thin, bottom=thin)

def hfont(color=C_WHITE, size=10):
    return Font(name='Arial', bold=True, color=color, size=size)

def dfont(bold=False, size=10, color='FF333333'):
    return Font(name='Arial', bold=bold, size=size, color=color)

def fl(c):
    return PatternFill('solid', fgColor=c)

def ctr():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def rgt():
    return Alignment(horizontal='right', vertical='center')

def lft():
    return Alignment(horizontal='left', vertical='center', wrap_text=False)


# ────────────────────────────────────────────────────────────────────────────
# 读取数据
# ────────────────────────────────────────────────────────────────────────────

def read_order_data():
    """读取建议下单量表，并识别月份列"""
    with db_cursor() as cursor:
        cursor.execute("SHOW COLUMNS FROM `建议下单量表`")
        all_cols = [r['Field'] for r in cursor.fetchall()]
        cursor.execute("""
            SELECT * FROM `建议下单量表`
            ORDER BY 面料类型 DESC, SPU, 店铺
        """)
        rows = cursor.fetchall()

    # 从列名里提取月份（取"XX运营预计"的月份前缀，去重保序）
    months = []
    seen = set()
    for c in all_cols:
        if c.endswith('运营预计') and c != '运营预计合计':
            m = c[:-4]  # 去掉"运营预计"4个字
            if m not in seen:
                months.append(m)
                seen.add(m)

    logger.info(f"建议下单量表：{len(rows)} 行，月份：{months}")
    return rows, months


def read_fabric_data():
    """读取面料预计用量表"""
    with db_cursor() as cursor:
        cursor.execute("""
            SELECT 面料, SPU数量, 建议下单量合计, `单件用量(米)`, `预计用量(米)`
            FROM `面料预计用量表`
            ORDER BY `预计用量(米)` DESC
        """)
        rows = cursor.fetchall()
    logger.info(f"面料预计用量表：{len(rows)} 行")
    return rows


# ────────────────────────────────────────────────────────────────────────────
# Sheet1：建议下单量（生产经理）
# ────────────────────────────────────────────────────────────────────────────

def build_sheet1(wb: Workbook, rows, months):
    ws = wb.active
    ws.title = '建议下单量（生产经理）'
    ws.freeze_panes = 'D3'
    ws.sheet_view.zoomScale = 90

    total_cols = 3 + len(months) * 2 + 2

    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    tc = ws.cell(1, 1, f'建议下单量报表（生产经理）  —  生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}')
    tc.font = Font(name='Arial', bold=True, size=12, color=C_WHITE)
    tc.fill = fl(C_HEADER_BG)
    tc.alignment = ctr()
    ws.row_dimensions[1].height = 26

    # 第2行：月份合并标题
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 36

    # 固定列（2-3行合并）
    for ci, h in enumerate(['SPU', '工厂', '面料类型'], 1):
        ws.merge_cells(start_row=2, start_column=ci, end_row=3, end_column=ci)
        c = ws.cell(2, ci, h)
        c.font = hfont()
        c.fill = fl(C_HEADER_BG)
        c.alignment = ctr()
        c.border = bdr

    # 月份双层列头
    col = 4
    for m in months:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)
        mc = ws.cell(2, col, m)
        mc.font = hfont()
        mc.fill = fl(C_HEADER_BG)
        mc.alignment = ctr()
        mc.border = bdr

        oh = ws.cell(3, col, '运营预估')
        oh.font = hfont(size=9)
        oh.fill = fl(C_MONTH_HDR_OP)
        oh.alignment = ctr()
        oh.border = bdr_mb

        sh = ws.cell(3, col + 1, '系统建议')
        sh.font = hfont(size=9)
        sh.fill = fl(C_MONTH_HDR_SYS)
        sh.alignment = ctr()
        sh.border = bdr
        col += 2

    # 合计列（2-3行合并）
    for h in ['建议下单合计', '运营预计合计']:
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
        c = ws.cell(2, col, h)
        c.font = hfont(size=9)
        c.fill = fl(C_HEADER_BG)
        c.alignment = ctr()
        c.border = bdr
        col += 1

    # 数据行
    for ri, row in enumerate(rows, 4):
        ws.row_dimensions[ri].height = 18
        shade = C_LIGHT_GRAY if ri % 2 == 0 else C_WHITE
        ftype = str(row.get('面料类型', '') or '')
        type_bg = C_CUSTOM if ftype == '定制面料' else C_STOCK

        # SPU
        c = ws.cell(ri, 1, str(row.get('SPU', '') or ''))
        c.font = dfont(bold=True)
        c.fill = fl(shade)
        c.alignment = lft()
        c.border = bdr

        # 工厂
        c = ws.cell(ri, 2, str(row.get('工厂', '') or ''))
        c.font = dfont(size=9, color='FF666666')
        c.fill = fl(shade)
        c.alignment = lft()
        c.border = bdr

        # 面料类型
        c = ws.cell(ri, 3, ftype)
        c.font = dfont(size=9)
        c.fill = fl(type_bg)
        c.alignment = ctr()
        c.border = bdr

        # 月份数据
        col = 4
        for m in months:
            op_v = int(row.get(f'{m}运营预计', 0) or 0)
            sy_v = int(row.get(f'{m}建议下单', 0) or 0)

            oc = ws.cell(ri, col, op_v if op_v > 0 else None)
            oc.font = dfont()
            oc.fill = fl(C_MONTH_OP)
            oc.alignment = rgt()
            oc.border = bdr_mb
            oc.number_format = '#,##0;-;-'

            sc = ws.cell(ri, col + 1, sy_v if sy_v > 0 else None)
            sc_color = C_RED_TEXT if sy_v > 300 else C_AMBER_TEXT if sy_v > 50 else 'FF333333'
            sc.font = dfont(bold=sy_v > 50, color=sc_color)
            sc.fill = fl(C_MONTH_SYS)
            sc.alignment = rgt()
            sc.border = bdr
            sc.number_format = '#,##0;-;-'
            col += 2

        # 合计
        for key in ['建议下单合计', '运营预计合计']:
            v = int(row.get(key, 0) or 0)
            c = ws.cell(ri, col, v if v > 0 else None)
            c.font = dfont(bold=True)
            c.fill = fl(shade)
            c.alignment = rgt()
            c.border = bdr
            c.number_format = '#,##0;-;-'
            col += 1

    # 列宽
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 10
    col = 4
    for _ in months:
        ws.column_dimensions[get_column_letter(col)].width = 9
        ws.column_dimensions[get_column_letter(col + 1)].width = 9
        col += 2
    for i in range(2):
        ws.column_dimensions[get_column_letter(col + i)].width = 12

    logger.info(f"Sheet1 写入完成：{len(rows)} 行数据")


# ────────────────────────────────────────────────────────────────────────────
# Sheet2：面料用量（产品经理）
# ────────────────────────────────────────────────────────────────────────────

def build_sheet2(wb: Workbook, rows):
    ws = wb.create_sheet('面料用量（产品经理）')
    ws.freeze_panes = 'A3'
    ws.sheet_view.zoomScale = 95

    # 标题
    ws.merge_cells('A1:E1')
    tc = ws.cell(1, 1, f'定制面料预计用量报表（产品经理）  —  生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}')
    tc.font = Font(name='Arial', bold=True, size=12, color=C_WHITE)
    tc.fill = fl(C_HEADER_BG)
    tc.alignment = ctr()
    ws.row_dimensions[1].height = 26

    # 表头
    headers = ['面料', 'SPU数量', '建议下单量合计', '单件用量(米)', '预计用量(米)']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(2, ci, h)
        c.font = hfont()
        c.fill = fl(C_HEADER_BG)
        c.alignment = ctr()
        c.border = bdr
    ws.row_dimensions[2].height = 22

    # 数据行
    for ri, row in enumerate(rows, 3):
        ws.row_dimensions[ri].height = 18
        shade = C_LIGHT_GRAY if ri % 2 == 0 else C_WHITE

        vals  = [
            str(row['面料']),
            int(row['SPU数量'] or 0),
            int(row['建议下单量合计'] or 0),
            float(row['单件用量(米)'] or 0),
            float(row['预计用量(米)'] or 0),
        ]
        fmts  = [None, '#,##0', '#,##0', '0.00', '#,##0.0']
        alns  = [lft(), rgt(), rgt(), rgt(), rgt()]
        bolds = [True, False, False, False, True]

        for ci, (v, fmt, aln, bold) in enumerate(zip(vals, fmts, alns, bolds), 1):
            c = ws.cell(ri, ci, v)
            c.font = dfont(bold=bold)
            c.fill = fl(shade)
            c.alignment = aln
            c.border = bdr
            if fmt:
                c.number_format = fmt

    # 合计行
    total_row = len(rows) + 3
    ws.row_dimensions[total_row].height = 20
    total_data = [
        ('合计', None, lft()),
        ('', None, rgt()),
        (f'=SUM(C3:C{total_row-1})', '#,##0', rgt()),
        ('', None, rgt()),
        (f'=SUM(E3:E{total_row-1})', '#,##0.0', rgt()),
    ]
    for ci, (v, fmt, aln) in enumerate(total_data, 1):
        c = ws.cell(total_row, ci, v)
        c.font = Font(name='Arial', bold=True, size=10, color=C_WHITE)
        c.fill = fl(C_HEADER_BG)
        c.alignment = aln
        c.border = bdr
        if fmt:
            c.number_format = fmt

    # 列宽
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 9
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14

    logger.info(f"Sheet2 写入完成：{len(rows)} 种面料")


# ────────────────────────────────────────────────────────────────────────────
# 主函数
# ────────────────────────────────────────────────────────────────────────────

def main():
    logger.info("=" * 60)
    logger.info("采购建议报告 Excel 导出")
    logger.info("=" * 60)

    order_rows, months = read_order_data()
    fabric_rows = read_fabric_data()

    wb = Workbook()
    build_sheet1(wb, order_rows, months)
    build_sheet2(wb, fabric_rows)

    # 输出目录
    output_dir = project_root / 'exports'
    output_dir.mkdir(exist_ok=True)
    filename = f'采购建议报告_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    output_path = output_dir / filename

    wb.save(str(output_path))

    logger.info(f"✓ 已导出：{output_path}")
    logger.info(f"  建议下单量：{len(order_rows)} 行，{len(months)} 个月份")
    logger.info(f"  定制面料用量：{len(fabric_rows)} 种")
    logger.info("=" * 60)

    return str(output_path)


if __name__ == '__main__':
    main()
