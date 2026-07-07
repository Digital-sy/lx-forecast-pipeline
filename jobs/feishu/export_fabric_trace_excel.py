#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""
通用面料溯源 Excel 导出。

用途：按面料关键词导出面料预估结果、SPU用料关系、SKU/月度来源明细、颜色贡献。

示例：
  cd /opt/apps/pythondata
  ./venv/bin/python -m jobs.feishu.export_fabric_trace_excel \
    --fabric "013仿棉拉架-优化" \
    --fabric "7525锦双磨"
"""
from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

project_root = Path(__file__).resolve().parent.parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

from common import get_logger
from common.database import db_cursor

logger = get_logger('fabric_trace_export')


def fetch_all(sql: str, params: Tuple[Any, ...] = ()) -> List[Dict[str, Any]]:
    with db_cursor(dictionary=True) as cursor:
        cursor.execute(sql, params)
        return list(cursor.fetchall())


def table_exists(table_name: str) -> bool:
    rows = fetch_all(
        """
        SELECT COUNT(*) AS cnt
        FROM information_schema.TABLES
        WHERE TABLE_SCHEMA = DATABASE()
          AND TABLE_NAME = %s
        """,
        (table_name,),
    )
    return bool(rows and rows[0].get('cnt'))


def get_columns(table_name: str) -> set:
    if not table_exists(table_name):
        return set()
    rows = fetch_all(
        """
        SELECT COLUMN_NAME
        FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA = DATABASE()
          AND TABLE_NAME = %s
        """,
        (table_name,),
    )
    return {r['COLUMN_NAME'] for r in rows}


def norm_date(value: Any) -> str:
    if value is None:
        return ''
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m-%d')
    return str(value)[:10]


def safe_name(name: str, max_len: int = 80) -> str:
    s = re.sub(r'[\\/:*?"<>|\[\]]+', '_', str(name or '').strip())
    return s[:max_len] or 'fabric'


def extract_spu(sku: str) -> str:
    if not sku:
        return ''
    sku = re.sub(r'\d+(?:PSC|PCS)', '', str(sku), flags=re.IGNORECASE)
    sku = re.sub(r'-+', '-', sku).strip('-')
    return sku.split('-', 1)[0] if '-' in sku else sku


def parse_lx_color(product_name: str) -> str:
    text = (product_name or '').strip()
    m = re.search(r'(\d+\s*#\s*[^#，,;/|]+)', text)
    return re.sub(r'\s+', '', m.group(1).strip()) if m else ''


def read_lx_color_map() -> Dict[str, str]:
    if not table_exists('lxpm_product_category_snapshot'):
        return {}
    rows = fetch_all("""
        SELECT sku, product_name
        FROM `lxpm_product_category_snapshot`
        WHERE sku IS NOT NULL AND sku != ''
          AND product_name IS NOT NULL AND product_name != ''
    """)
    out = {}
    for r in rows:
        color = parse_lx_color(r.get('product_name') or '')
        if color:
            out[(r.get('sku') or '').strip()] = color
    return out


def like_condition(columns: Iterable[str], keyword: str) -> Tuple[str, List[str]]:
    parts, params = [], []
    for col in columns:
        parts.append(f"`{col}` LIKE %s")
        params.append(f"%{keyword}%")
    return '(' + ' OR '.join(parts) + ')', params


def read_forecast_result(keyword: str, lx_color_map: Dict[str, str]) -> List[Dict[str, Any]]:
    table = '面料预估表'
    if not table_exists(table):
        return []
    cols = get_columns(table)
    where_cols = [c for c in ['面料', '面料编号', '面料颜色编号'] if c in cols]
    if not where_cols:
        return []
    where, params = like_condition(where_cols, keyword)
    select_cols = [
        '统计类型', '月份', '面料', '面料编号', '颜色缩写', '颜色', '面料颜色编号',
        '库存量/条', '库存量/米', '待到货量/条', '待到货量/米',
        '当月已下单消耗/米', '当月完整预估/米', '当月剩余预估/米',
        'T+1月预估/米', 'T+2月预估/米', 'T+3月预估/米',
        '运营当月预估/米', '运营T+1月预估/米', '运营T+2月预估/米', '运营T+3月预估/米',
        '用量信息缺失SPU', '更新时间'
    ]
    exprs = []
    for c in select_cols:
        if c in cols:
            exprs.append(f'`{c}`')
        else:
            default = 'NULL' if c == '更新时间' else "''" if c in ['统计类型','月份','面料','面料编号','颜色缩写','颜色','面料颜色编号','用量信息缺失SPU'] else '0'
            exprs.append(f'{default} AS `{c}`')
    rows = fetch_all(f"SELECT {', '.join(exprs)} FROM `{table}` WHERE {where} ORDER BY 统计类型, 面料, 颜色缩写", tuple(params))
    for r in rows:
        code = (r.get('面料颜色编号') or '').strip()
        r['颜色-领星'] = lx_color_map.get(code, '')
        r['更新时间'] = norm_date(r.get('更新时间'))
    return rows


def read_spu_usage(keyword: str) -> List[Dict[str, Any]]:
    table = '面料核价表'
    if not table_exists(table):
        return []
    cols = get_columns(table)
    if 'SPU' not in cols or '面料' not in cols:
        return []
    usage_expr = '`单件用量`' if '单件用量' in cols else '0'
    loss_expr = '`单件损耗`' if '单件损耗' in cols else '0'
    rows = fetch_all(f"""
        SELECT SPU, 面料, {usage_expr} AS 单件用量, {loss_expr} AS 单件损耗
        FROM `{table}`
        WHERE 面料 LIKE %s
        ORDER BY SPU, 单件用量 DESC
    """, (f'%{keyword}%',))
    return rows


def read_source_detail(keyword: str, lx_color_map: Dict[str, str]) -> List[Dict[str, Any]]:
    """按 SKU+月份 溯源：运营预计、系统预测、单件用量、推算用量。"""
    spu_rows = read_spu_usage(keyword)
    usage_by_spu: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for r in spu_rows:
        usage_by_spu[(r.get('SPU') or '').strip()].append(r)

    op_map: Dict[Tuple[str, str], int] = defaultdict(int)
    if table_exists('运营预计下单表'):
        cols = get_columns('运营预计下单表')
        if {'SKU', '统计日期', '预计下单量'} <= cols:
            rows = fetch_all("""
                SELECT SKU, 统计日期, SUM(预计下单量) AS qty
                FROM `运营预计下单表`
                WHERE SKU IS NOT NULL AND SKU != '' AND 统计日期 IS NOT NULL
                GROUP BY SKU, 统计日期
            """)
            for r in rows:
                op_map[((r.get('SKU') or '').strip(), norm_date(r.get('统计日期')))] += int(r.get('qty') or 0)

    sys_map: Dict[Tuple[str, str], int] = defaultdict(int)
    if table_exists('预测对比表_SKU'):
        cols = get_columns('预测对比表_SKU')
        if {'SKU', '统计日期', '系统预测销量'} <= cols:
            rows = fetch_all("""
                SELECT SKU, 统计日期, SUM(系统预测销量) AS qty
                FROM `预测对比表_SKU`
                WHERE SKU IS NOT NULL AND SKU != '' AND 统计日期 IS NOT NULL
                GROUP BY SKU, 统计日期
            """)
            for r in rows:
                sys_map[((r.get('SKU') or '').strip(), norm_date(r.get('统计日期')))] += int(r.get('qty') or 0)

    keys = sorted(set(op_map.keys()) | set(sys_map.keys()), key=lambda x: (x[1], x[0]))
    out = []
    for sku, stat_date in keys:
        spu = extract_spu(sku)
        for u in usage_by_spu.get(spu, []):
            unit = float(u.get('单件用量') or 0)
            loss = float(u.get('单件损耗') or 0)
            factor = loss if loss else 1
            op_qty = int(op_map.get((sku, stat_date), 0) or 0)
            sys_qty = int(sys_map.get((sku, stat_date), 0) or 0)
            if op_qty == 0 and sys_qty == 0:
                continue
            color_code = '-'.join(str(sku).split('-')[:2]) if '-' in str(sku) else str(sku)
            out.append({
                'SKU': sku,
                'SPU': spu,
                '统计日期': stat_date,
                '面料': u.get('面料') or '',
                '颜色款号': color_code,
                '颜色-领星': lx_color_map.get(color_code, ''),
                '单件用量': unit,
                '单件损耗': loss,
                '运营预计下单量': op_qty,
                '系统预估销量': sys_qty,
                '运营预计用量_米': round(op_qty * unit * factor, 2),
                '系统预估用量_米': round(sys_qty * unit * factor, 2),
            })
    return out


def summarize_spu(source_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    agg = defaultdict(lambda: defaultdict(float))
    for r in source_rows:
        k = (r['SPU'], r['面料'])
        agg[k]['运营预计下单量'] += float(r.get('运营预计下单量') or 0)
        agg[k]['系统预估销量'] += float(r.get('系统预估销量') or 0)
        agg[k]['运营预计用量_米'] += float(r.get('运营预计用量_米') or 0)
        agg[k]['系统预估用量_米'] += float(r.get('系统预估用量_米') or 0)
    rows = []
    total = sum(v['系统预估用量_米'] for v in agg.values()) or 1
    for (spu, fabric), v in agg.items():
        rows.append({
            'SPU': spu,
            '面料': fabric,
            '运营预计下单量': int(v['运营预计下单量']),
            '系统预估销量': int(v['系统预估销量']),
            '运营预计用量_米': round(v['运营预计用量_米'], 2),
            '系统预估用量_米': round(v['系统预估用量_米'], 2),
            '系统用量占比': round(v['系统预估用量_米'] / total, 4),
        })
    rows.sort(key=lambda x: -x['系统预估用量_米'])
    return rows


def summarize_color(source_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    agg = defaultdict(lambda: defaultdict(float))
    for r in source_rows:
        k = (r.get('颜色款号') or '', r.get('颜色-领星') or '')
        agg[k]['运营预计用量_米'] += float(r.get('运营预计用量_米') or 0)
        agg[k]['系统预估用量_米'] += float(r.get('系统预估用量_米') or 0)
        agg[k]['运营预计下单量'] += float(r.get('运营预计下单量') or 0)
        agg[k]['系统预估销量'] += float(r.get('系统预估销量') or 0)
    rows = []
    for (code, color), v in agg.items():
        rows.append({
            '颜色款号': code,
            '颜色-领星': color,
            '运营预计下单量': int(v['运营预计下单量']),
            '系统预估销量': int(v['系统预估销量']),
            '运营预计用量_米': round(v['运营预计用量_米'], 2),
            '系统预估用量_米': round(v['系统预估用量_米'], 2),
        })
    rows.sort(key=lambda x: -x['系统预估用量_米'])
    return rows


def write_sheet(wb: Workbook, title: str, rows: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet(title[:31])
    if not rows:
        ws.append(['提示'])
        ws.append(['无数据'])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, '') for h in headers])
    fill = PatternFill('solid', fgColor='1F4E78')
    font = Font(bold=True, color='FFFFFF')
    thin = Side(style='thin', color='D9E2F3')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in ws[1]:
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical='center')
            if isinstance(c.value, (int, float)):
                c.number_format = '0.00%' if '占比' in str(ws.cell(1, c.column).value) else '#,##0.00'
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    if ws.max_row >= 2:
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        table_name = re.sub(r'[^A-Za-z0-9_]', '_', f'tbl_{title}')[:200]
        if not re.match(r'^[A-Za-z_]', table_name):
            table_name = 'T_' + table_name
        try:
            table = Table(displayName=table_name, ref=ref)
            table.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True, showColumnStripes=False)
            ws.add_table(table)
        except Exception:
            pass
    for idx, col in enumerate(ws.columns, start=1):
        width = min(max(max(len(str(c.value or '')) for c in col) + 2, 10), 42)
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_summary(wb: Workbook, fabric: str, result_rows, spu_rows, source_rows, color_rows):
    ws = wb.active
    ws.title = '汇总说明'
    ws.append([f'{fabric} 面料溯源报告'])
    ws.append(['生成时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    ws.append(['面料预估结果行数', len(result_rows)])
    ws.append(['SPU用料关系行数', len(spu_rows)])
    ws.append(['SKU来源明细行数', len(source_rows)])
    ws.append(['颜色贡献行数', len(color_rows)])
    ws.append(['说明', '系统/运营来源明细按 SKU+月份 追溯，单件用量来自面料核价表；最终面料预估结果来自面料预估表。'])
    ws.merge_cells('A1:B1')
    ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='1F4E78')
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 100


def export_one(fabric: str, output_dir: str, lx_color_map: Dict[str, str]) -> str:
    logger.info(f'开始导出面料溯源：{fabric}')
    result_rows = read_forecast_result(fabric, lx_color_map)
    spu_rows = read_spu_usage(fabric)
    source_rows = read_source_detail(fabric, lx_color_map)
    spu_summary = summarize_spu(source_rows)
    color_summary = summarize_color(source_rows)

    wb = Workbook()
    write_summary(wb, fabric, result_rows, spu_rows, source_rows, color_summary)
    write_sheet(wb, '01_面料预估结果', result_rows)
    write_sheet(wb, '02_SPU用料关系', spu_rows)
    write_sheet(wb, '03_SPU贡献汇总', spu_summary)
    write_sheet(wb, '04_颜色贡献汇总', color_summary)
    write_sheet(wb, '05_SKU月份溯源明细', source_rows)

    out = Path(output_dir) / f"{safe_name(fabric)}_面料溯源_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    logger.info(f'✓ 已生成：{out}')
    return str(out)


def parse_args():
    parser = argparse.ArgumentParser(description='通用面料溯源Excel导出')
    parser.add_argument('--fabric', action='append', required=True, help='面料关键词，可重复传入')
    parser.add_argument('--output-dir', default='exports', help='输出目录，默认 exports')
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    lx_color_map = read_lx_color_map()
    logger.info(f'颜色-领星映射：{len(lx_color_map)} 条')
    for fabric in args.fabric:
        export_one(fabric, args.output_dir, lx_color_map)


if __name__ == '__main__':
    main()
