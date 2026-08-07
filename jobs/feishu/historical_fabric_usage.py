#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""历史采购实际面料用量：纯计算与 Excel 导出逻辑。"""
from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

LOSS_FACTOR = 1.10
VOID_STATUSES = {-1, 124}


def normalize_text(value: Any) -> str:
    return '' if value is None else str(value).strip()


def parse_datetime(value: Any) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    text = normalize_text(value)
    if not text:
        return None
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
        try:
            return datetime.strptime(text[:19] if ' ' in fmt else text[:10], fmt)
        except ValueError:
            continue
    return None


def month_range(start_date: date, end_date: date) -> List[str]:
    if start_date > end_date:
        raise ValueError('start_date 不能晚于 end_date')
    result: List[str] = []
    year, month = start_date.year, start_date.month
    while (year, month) <= (end_date.year, end_date.month):
        result.append(f'{year}-{month:02d}')
        month += 1
        if month == 13:
            month = 1
            year += 1
    return result


def fallback_spu_from_sku(sku: str) -> str:
    sku = normalize_text(sku)
    if not sku:
        return ''
    return sku.split('-', 1)[0].strip()


def normalize_fabric_map(rows: Iterable[Mapping[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    result: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    seen = set()
    for row in rows:
        spu = normalize_text(row.get('SPU'))
        fabric = normalize_text(row.get('面料'))
        try:
            usage = float(row.get('单件用量') or 0)
        except (TypeError, ValueError):
            usage = 0.0
        if not spu or not fabric:
            continue
        key = (spu, fabric)
        if key in seen:
            continue
        seen.add(key)
        result[spu].append({'面料': fabric, '单件用量': usage})
    return dict(result)


def expand_purchase_orders(
    orders: Sequence[Mapping[str, Any]],
    analysis_start: date,
    analysis_end: date,
    fabric_map: Mapping[str, Sequence[Mapping[str, Any]]],
    custom_fabrics: Iterable[str],
    loss_factor: float = LOSS_FACTOR,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """将采购单展开到 采购明细 × SPU面料 粒度。"""
    custom_set = {normalize_text(x) for x in custom_fabrics if normalize_text(x)}
    detail_rows: List[Dict[str, Any]] = []
    unmapped_rows: List[Dict[str, Any]] = []

    for order in orders:
        status = order.get('status')
        try:
            status_int = int(status) if status is not None and str(status).strip() else None
        except (TypeError, ValueError):
            status_int = None
        if status_int in VOID_STATUSES:
            continue

        order_time = parse_datetime(order.get('order_time'))
        if not order_time:
            continue
        if order_time.date() < analysis_start or order_time.date() > analysis_end:
            continue

        order_sn = normalize_text(order.get('order_sn'))
        month = order_time.strftime('%Y-%m')
        item_list = order.get('item_list') or []
        if not isinstance(item_list, list):
            continue

        for item in item_list:
            if not isinstance(item, Mapping):
                continue
            try:
                is_delete = int(item.get('is_delete') or 0)
            except (TypeError, ValueError):
                is_delete = 0
            if is_delete == 1:
                continue

            sku = normalize_text(item.get('sku'))
            spu = normalize_text(item.get('spu')) or fallback_spu_from_sku(sku)
            spu_source = '接口SPU' if normalize_text(item.get('spu')) else ('SKU推导' if spu else '缺失')
            try:
                qty_real = float(item.get('quantity_real') or 0)
            except (TypeError, ValueError):
                qty_real = 0.0
            try:
                qty_entry = float(item.get('quantity_entry') or 0)
            except (TypeError, ValueError):
                qty_entry = 0.0
            try:
                qty_receive = float(item.get('quantity_receive') or 0)
            except (TypeError, ValueError):
                qty_receive = 0.0

            fabrics = list(fabric_map.get(spu, [])) if spu else []
            valid_fabrics = []
            for fabric_info in fabrics:
                fabric = normalize_text(fabric_info.get('面料'))
                try:
                    unit_usage = float(fabric_info.get('单件用量') or 0)
                except (TypeError, ValueError):
                    unit_usage = 0.0
                if fabric and unit_usage > 0:
                    valid_fabrics.append((fabric, unit_usage))

            if not valid_fabrics:
                reason = '缺少款号' if not spu else ('面料核价表无款号映射' if not fabrics else '面料单件用量为空或<=0')
                unmapped_rows.append({
                    '采购单号': order_sn,
                    '采购明细ID': item.get('id'),
                    'SKU': sku,
                    '款号': spu,
                    '款号来源': spu_source,
                    '下单时间': order_time.strftime('%Y-%m-%d %H:%M:%S'),
                    '采购月份': month,
                    '实际采购量': qty_real,
                    '到货入库量': qty_entry,
                    '待到货量': qty_receive,
                    '异常原因': reason,
                })
                continue

            for fabric, unit_usage in valid_fabrics:
                usage_m = qty_entry * unit_usage * loss_factor
                detail_rows.append({
                    '采购单号': order_sn,
                    '采购明细ID': item.get('id'),
                    'SKU': sku,
                    '款号': spu,
                    '款号来源': spu_source,
                    '下单时间': order_time.strftime('%Y-%m-%d %H:%M:%S'),
                    '采购月份': month,
                    '采购状态': normalize_text(order.get('status_text')),
                    '到货状态': normalize_text(order.get('status_shipped_text')),
                    '实际采购量': qty_real,
                    '到货入库量': qty_entry,
                    '待到货量': qty_receive,
                    '面料': fabric,
                    '面料种类': '定制' if fabric in custom_set else '现货',
                    '单件用量(M)': unit_usage,
                    '损耗系数': loss_factor,
                    '含损耗单件用量(M)': unit_usage * loss_factor,
                    '本批面料用量(M)': usage_m,
                    '到货量大于采购量': '是' if qty_entry > qty_real and qty_real > 0 else '否',
                })

    return detail_rows, unmapped_rows


def aggregate_usage(
    detail_rows: Sequence[Mapping[str, Any]],
    months: Sequence[str],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    fabric_monthly: Dict[Tuple[str, str], Dict[str, float]] = defaultdict(lambda: defaultdict(float))
    spu_fabric_monthly: Dict[Tuple[str, str, str, float], Dict[str, float]] = defaultdict(lambda: defaultdict(float))

    for row in detail_rows:
        month = normalize_text(row.get('采购月份'))
        if month not in months:
            continue
        fabric = normalize_text(row.get('面料'))
        fabric_type = normalize_text(row.get('面料种类'))
        spu = normalize_text(row.get('款号'))
        try:
            unit_usage = float(row.get('单件用量(M)') or 0)
            qty_m = float(row.get('本批面料用量(M)') or 0)
        except (TypeError, ValueError):
            continue
        fabric_monthly[(fabric, fabric_type)][month] += qty_m
        spu_fabric_monthly[(spu, fabric, fabric_type, unit_usage)][month] += qty_m

    fabric_rows: List[Dict[str, Any]] = []
    for (fabric, fabric_type), monthly in sorted(fabric_monthly.items()):
        row: Dict[str, Any] = {'面料': fabric, '面料种类': fabric_type}
        for month in months:
            row[month] = round(monthly.get(month, 0.0), 2)
        row['累计用量(M)'] = round(sum(monthly.values()), 2)
        fabric_rows.append(row)

    spu_rows: List[Dict[str, Any]] = []
    for (spu, fabric, fabric_type, unit_usage), monthly in sorted(spu_fabric_monthly.items()):
        row = {
            '款号': spu,
            '面料': fabric,
            '面料种类': fabric_type,
            '单件用量(M)': unit_usage,
            '含损耗单件用量(M)': round(unit_usage * LOSS_FACTOR, 6),
        }
        for month in months:
            row[month] = round(monthly.get(month, 0.0), 2)
        row['累计用量(M)'] = round(sum(monthly.values()), 2)
        spu_rows.append(row)

    return fabric_rows, spu_rows


def _autosize(ws, max_width: int = 28) -> None:
    from openpyxl.utils import get_column_letter
    for idx, column_cells in enumerate(ws.columns, 1):
        width = 10
        for cell in column_cells:
            value = '' if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(value) + 2))
        ws.column_dimensions[get_column_letter(idx)].width = width


def _write_table(ws, headers: Sequence[str], rows: Sequence[Mapping[str, Any]]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    ws.append(list(headers))
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='D9EAF7')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row in rows:
        ws.append([row.get(h, '') for h in headers])
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    _autosize(ws)


def export_excel(
    output_path: Path,
    fabric_rows: Sequence[Mapping[str, Any]],
    spu_rows: Sequence[Mapping[str, Any]],
    detail_rows: Sequence[Mapping[str, Any]],
    unmapped_rows: Sequence[Mapping[str, Any]],
    months: Sequence[str],
) -> Path:
    from openpyxl import Workbook

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws1 = wb.active
    ws1.title = '面料月度用量'
    _write_table(ws1, ['面料', '面料种类', *months, '累计用量(M)'], fabric_rows)

    ws2 = wb.create_sheet('款号面料月度用量')
    _write_table(ws2, ['款号', '面料', '面料种类', '单件用量(M)', '含损耗单件用量(M)', *months, '累计用量(M)'], spu_rows)

    ws3 = wb.create_sheet('采购明细核算')
    detail_headers = [
        '采购单号', '采购明细ID', 'SKU', '款号', '款号来源', '下单时间', '采购月份',
        '采购状态', '到货状态', '实际采购量', '到货入库量', '待到货量', '面料', '面料种类',
        '单件用量(M)', '损耗系数', '含损耗单件用量(M)', '本批面料用量(M)', '到货量大于采购量',
    ]
    _write_table(ws3, detail_headers, detail_rows)

    ws4 = wb.create_sheet('未映射检查')
    unmapped_headers = [
        '采购单号', '采购明细ID', 'SKU', '款号', '款号来源', '下单时间', '采购月份',
        '实际采购量', '到货入库量', '待到货量', '异常原因',
    ]
    _write_table(ws4, unmapped_headers, unmapped_rows)

    wb.save(output_path)
    return output_path
