#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""把 SKU 级预估销量只读聚合到“17 个目标面料+当前有效颜色”的米数。

本模块是现有预测/采购管线后的核对层：

1. 从 ``预测对比表_SKU`` 读取当前月起四个月的 SKU 预估销量；
2. 优先读取预测表已存在的 ``品名/颜色编码/颜色中文名/颜色体系``；
   当前仓库主表尚未持久化这些列时，复用产品快照、SKU 颜色体系解析器和
   既有领星品名颜色解析规则补齐；
3. 复用 ``面料核价表`` 的 SPU→面料、单件用量和单件损耗；
4. 从飞书当前有效清单读取 ``面料名/颜色/领星新颜色缩写``；
5. 严格按原值中文名、缩写、清单色号、颜色体系、确定性中文核心/别名、
   历史人工确认的优先级匹配；
6. 仅在确定性规则全部失败后生成同面料模糊候选，候选永不计入已确认用量。

米数口径固定为 ``SKU 未来 4 个月预估销量 × 单件用量 × 单件损耗``。单件
用量为空或为 0 时不使用均值兜底。默认执行是 read-only dry-run：只读取
MySQL/飞书和本地人工确认 CSV，并生成本地 Excel/JSON，不回写远端。
"""
from __future__ import annotations

import argparse
import asyncio
import csv
import difflib
import json
import os
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, MutableMapping, Sequence

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from common import get_logger
from common.database import db_cursor
from common.feishu import FeishuClient
from jobs.feishu import generate_fabric_forecast as fabric_base
from jobs.feishu.color_mapping_catalog import (
    SUPPORTED_SYSTEMS,
    UNKNOWN_SYSTEM,
    ColorMappingCatalog,
)
from jobs.feishu.color_system_resolver import ColorSystemResolver, normalize_sku
from jobs.feishu.generate_procurement_report_lx_color import parse_lingxing_color

logger = get_logger("fabric_color_stocking")

PROJECT_ROOT = Path(__file__).resolve().parents[2]
FORECAST_TABLE = "预测对比表_SKU"
SNAPSHOT_TABLE = "lxpm_product_category_snapshot"

# 这些是飞书资源坐标，不是认证凭证；认证仍只从 FEISHU_APP_ID /
# FEISHU_APP_SECRET 环境变量读取。默认值复用现有面料台账同步任务。
DEFAULT_BASE_TOKEN = "XT6pbXxmdas4rdsme0XctyefnGu"
DEFAULT_CATALOG_TABLE_ID = "tblMxScMbTyLQbyj"
DEFAULT_CATALOG_VIEW_ID = "vewHPceQyu"
DEFAULT_FABRIC_NAME_FIELD = "面料名"
DEFAULT_COLOR_FIELD = "颜色"
DEFAULT_LX_CODE_FIELD = "领星新颜色缩写"
DEFAULT_LINKED_FABRIC_FIELD = "面料品名"
DEFAULT_MANUAL_MAPPING_PATH = (
    PROJECT_ROOT / "config" / "fabric_color_manual_mapping.csv"
)

TARGET_FABRICS = (
    "002双六面料",
    "013仿棉拉架",
    "013仿棉拉架-优化",
    "023内裤面料",
    "037超绒面料",
    "037超绒面料-优化",
    "208 1*1加厚罗纹",
    "255单六面料",
    "290涤双磨",
    "316棉柔丝",
    "325单六双磨",
    "381锦单面",
    "4570锦氨网",
    "7525锦双磨",
    "N7040锦仿棉",
    "云朵棉",
    "羊毛平纹",
)
TARGET_FABRIC_ORDER = {name: index for index, name in enumerate(TARGET_FABRICS)}

MATCH_NAME = "中文名直连"
MATCH_CODE = "缩写"
MATCH_CATALOG_LABEL = "清单色号解析"
MATCH_SYSTEM = "体系消歧"
MATCH_CORE_ALIAS = "确定性别名/中文核心"
MATCH_MANUAL = "历史人工确认"
MATCH_METHOD_ORDER = (
    MATCH_NAME,
    MATCH_CODE,
    MATCH_CATALOG_LABEL,
    MATCH_SYSTEM,
    MATCH_CORE_ALIAS,
    MATCH_MANUAL,
)

REASON_NO_FABRIC = "品名/SPU无面料映射"
REASON_FABRIC_INACTIVE = "面料不在当前清单"
REASON_PENDING_SYSTEM = "颜色体系待定无法消歧"
REASON_NAME_UNKNOWN = "中文名与清单无对应"
REASON_COLOR_INACTIVE = "颜色不在当前清单/疑似停用"
REASON_SYSTEM_MAPPING = "颜色体系+中文名无映射，无法消歧"
REASON_AMBIGUOUS = "当前清单一对多，无法唯一定位"
REASON_FORECAST_CONFLICT = "预估输出颜色字段冲突"
REASON_PATTERN_MISSING = "缺少对应图案色号"
REASON_MANUAL_AMBIGUOUS = "人工确认映射冲突，无法唯一定位"
REASON_USAGE_MISSING = "单件用量为空或0"

PATTERN_GROUPS = {
    "圆点/波点": ("波点", "圆点", "白点", "黑点"),
    "豹纹": ("豹纹",),
    "印花/花色": ("印花", "花色"),
    "格子": ("格子",),
    "条纹": ("条纹",),
    "撞色/拼色": ("撞色", "拼色"),
    "底色": ("黑底", "白底", "底色"),
}
INVALID_CATALOG_KEYWORDS = ("取消", "停用", "作废", "无效")
BUSINESS_SUFFIXES = ("性价比", "新版", "优化", "新")
TRADITIONAL_COLOR_TRANSLATION = str.maketrans(
    {
        "紅": "红",
        "綠": "绿",
        "藍": "蓝",
        "黃": "黄",
        "點": "点",
        "紋": "纹",
        "淺": "浅",
        "銀": "银",
        "鐵": "铁",
        "寶": "宝",
        "絲": "丝",
        "絨": "绒",
        "雲": "云",
        "瑪": "玛",
        "駝": "驼",
        "棗": "枣",
        "磚": "砖",
        "薰": "熏",
    }
)

FUZZY_REVIEW_RESULTS = (
    "通过候选1",
    "通过候选2",
    "通过候选3",
    "无合适候选",
    "新增飞书颜色",
    "暂不处理",
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
PENDING_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="D9E1F2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _quantity(value: float) -> int | float:
    value = round(float(value or 0), 4)
    return int(value) if value.is_integer() else value


def _meters(value: float) -> float:
    return round(float(value or 0), 2)


CATALOG_COLOR_PATTERNS = (
    ("数字#前缀", re.compile(r"^\d+#(.+)$")),
    ("数字-前缀", re.compile(r"^\d+[-－](.+)$")),
)


def parse_catalog_color_label(value: Any) -> tuple[str, str]:
    """显式解析飞书颜色字段中的业务色号前缀。

    仅接受两种确定性结构：

    - ``2#黑色`` -> ``黑色``
    - ``037-拿铁`` -> ``拿铁``

    不执行 strip、大小写归一、别名替换或模糊匹配。没有命中明确结构时
    保留原值。该结果使用独立匹配方式“清单色号解析”，不会伪装成原值直连。
    """
    raw = "" if value is None else str(value)
    for rule, pattern in CATALOG_COLOR_PATTERNS:
        match = pattern.fullmatch(raw)
        if match:
            return match.group(1), rule
    return raw, "原值"


def catalog_color_components(value: Any) -> tuple[str, tuple[str, ...]]:
    """返回显式色号后的中文核心及括号内别名。

    该函数只服务于第五级确定性规则和展示字段。它不会覆盖飞书原值，也不会
    删除“色”、业务后缀或做相似度判断。
    """
    parsed, _ = parse_catalog_color_label(value)
    bracketed = re.fullmatch(r"(.*?)[（(]([^()（）]+)[）)]", parsed)
    if not bracketed:
        return parsed, ()
    core = bracketed.group(1).strip()
    aliases = tuple(
        dict.fromkeys(
            part.strip()
            for part in re.split(r"[,，、/|；;]", bracketed.group(2))
            if part.strip()
        )
    )
    return core or parsed, aliases


def pattern_groups(value: Any) -> set[str]:
    text = "" if value is None else str(value)
    return {
        group
        for group, keywords in PATTERN_GROUPS.items()
        if any(keyword in text for keyword in keywords)
    }


def pattern_compatible(source: Any, target: Any) -> bool:
    source_groups = pattern_groups(source)
    return not source_groups or source_groups.issubset(pattern_groups(target))


def fuzzy_name_variants(value: Any) -> set[str]:
    """只为人工候选生成标准化名称；绝不用于自动确认。"""
    raw = "" if value is None else str(value)
    if not raw:
        return set()
    normalized = unicodedata.normalize("NFKC", raw).translate(
        TRADITIONAL_COLOR_TRANSLATION
    )
    normalized = normalized.replace("（", "(").replace("）", ")")
    normalized, _ = parse_catalog_color_label(normalized)
    bracketed = re.fullmatch(r"(.*?)\(([^()]*)\)", normalized)
    values = [normalized]
    if bracketed:
        values = [bracketed.group(1), *re.split(r"[,，、/|；;]", bracketed.group(2))]

    result: set[str] = set()
    for value_part in values:
        text = re.sub(r"[\s·•,，。；;:/|_\-—－]+", "", value_part)
        changed = True
        while changed and text:
            changed = False
            for suffix in BUSINESS_SUFFIXES:
                if text.endswith(suffix):
                    text = text[: -len(suffix)]
                    changed = True
                    break
        if text.endswith("色") and len(text) > 1:
            text = text[:-1]
        if text:
            result.add(text)
    return result


@dataclass(frozen=True)
class FuzzyMatchConfig:
    high_min_score: float = 90.0
    medium_min_score: float = 80.0
    low_min_score: float = 70.0
    high_lead_margin: float = 10.0
    max_candidates: int = 3
    name_weight: float = 75.0
    alias_weight: float = 10.0
    code_weight: float = 10.0
    system_weight: float = 5.0
    governance_weight: float = 10.0
    manual_weight: float = 10.0

    def validate(self) -> None:
        if not (
            0 <= self.low_min_score <= self.medium_min_score
            <= self.high_min_score <= 100
        ):
            raise ValueError("模糊候选阈值必须满足 0<=低<=中<=高<=100")
        if self.high_lead_margin < 0:
            raise ValueError("第一、第二候选分差阈值不能为负数")
        if not 1 <= self.max_candidates <= 3:
            raise ValueError("最多候选数必须在 1 到 3 之间")


@dataclass(frozen=True)
class ForecastSku:
    sku: str
    spu: str
    product_name: str
    color_code: str
    color_name: str
    color_system: str
    forecast_qty: float
    field_issue: str = ""


@dataclass(frozen=True)
class CatalogRow:
    fabric_name: str
    color_name: str
    lingxing_code: str
    record_ids: tuple[str, ...] = ()
    raw_row_count: int = 1

    @property
    def identity(self) -> tuple[str, str, str]:
        return self.fabric_name, self.color_name, self.lingxing_code

    @property
    def color_core(self) -> str:
        return catalog_color_components(self.color_name)[0]

    @property
    def color_aliases(self) -> tuple[str, ...]:
        return catalog_color_components(self.color_name)[1]


@dataclass(frozen=True)
class MatchDecision:
    row: CatalogRow | None
    method: str = ""
    reason_code: str = ""
    reason: str = ""
    candidates: tuple[CatalogRow, ...] = ()


@dataclass(frozen=True)
class FabricUsage:
    spu: str
    fabric_name: str
    unit_usage: float
    loss_coefficient: float

    @property
    def usage_missing(self) -> bool:
        return self.unit_usage <= 0

    def calculate_meters(self, forecast_qty: float) -> float:
        if self.usage_missing:
            return 0.0
        loss = self.loss_coefficient if self.loss_coefficient > 0 else 1.0
        return float(forecast_qty or 0) * self.unit_usage * loss


@dataclass(frozen=True)
class ManualMapping:
    fabric_name: str
    color_code: str
    color_name: str
    color_system: str
    record_id: str
    catalog_color_name: str
    reviewer: str = ""
    reviewed_at: str = ""
    enabled: bool = True
    note: str = ""

    @property
    def key(self) -> tuple[str, str, str, str]:
        return (
            self.fabric_name,
            self.color_code,
            self.color_name,
            self.color_system,
        )


@dataclass(frozen=True)
class FuzzyCandidate:
    row: CatalogRow
    score: float
    basis: str


@dataclass
class StockingResult:
    overview_rows: list[dict[str, Any]]
    color_usage_rows: list[dict[str, Any]]
    auto_merge_rows: list[dict[str, Any]]
    manual_review_rows: list[dict[str, Any]]
    usage_missing_rows: list[dict[str, Any]]
    catalog_rows: list[dict[str, Any]]
    fuzzy_review_rows: list[dict[str, Any]]
    priority_rows: list[dict[str, Any]]
    metrics: dict[str, Any]

    @property
    def main_rows(self) -> list[dict[str, Any]]:
        """兼容旧调用方；新业务表名为“飞书颜色用量”."""
        return self.color_usage_rows

    @property
    def unmatched_rows(self) -> list[dict[str, Any]]:
        """兼容旧调用方；新业务表名为“待人工确认SKU”."""
        return self.manual_review_rows


class CatalogIndex:
    """严格匹配索引，并把完全相同的飞书业务行显式折叠。"""

    def __init__(self, raw_rows: Iterable[CatalogRow], source_record_count: int | None = None):
        raw_rows = list(raw_rows)
        grouped: dict[tuple[str, str, str], list[CatalogRow]] = defaultdict(list)
        for row in raw_rows:
            if row.fabric_name and row.color_name:
                grouped[row.identity].append(row)

        rows: list[CatalogRow] = []
        for identity, duplicates in grouped.items():
            record_ids = tuple(
                sorted({rid for item in duplicates for rid in item.record_ids if rid})
            )
            rows.append(
                CatalogRow(
                    fabric_name=identity[0],
                    color_name=identity[1],
                    lingxing_code=identity[2],
                    record_ids=record_ids,
                    raw_row_count=sum(item.raw_row_count for item in duplicates),
                )
            )
        self.rows = sorted(rows, key=lambda row: row.identity)
        self.source_record_count = source_record_count if source_record_count is not None else len(raw_rows)
        self.raw_association_count = len(raw_rows)

        self.by_name: MutableMapping[tuple[str, str], list[CatalogRow]] = defaultdict(list)
        self.by_parsed_name: MutableMapping[
            tuple[str, str], list[CatalogRow]
        ] = defaultdict(list)
        self.by_core_name: MutableMapping[
            tuple[str, str], list[CatalogRow]
        ] = defaultdict(list)
        self.by_alias_name: MutableMapping[
            tuple[str, str], list[CatalogRow]
        ] = defaultdict(list)
        self.by_code: MutableMapping[tuple[str, str], list[CatalogRow]] = defaultdict(list)
        self.by_fabric: MutableMapping[str, list[CatalogRow]] = defaultdict(list)
        self.by_record_id: dict[str, CatalogRow] = {}
        self.global_names: set[str] = set()
        self.parsed_color_rule_counts: Counter[str] = Counter()
        pair_identities: MutableMapping[tuple[str, str], list[CatalogRow]] = defaultdict(list)

        for row in self.rows:
            self.by_name[(row.fabric_name, row.color_name)].append(row)

            parsed_name, parsed_rule = parse_catalog_color_label(row.color_name)
            if parsed_name and parsed_name != row.color_name:
                self.by_parsed_name[(row.fabric_name, parsed_name)].append(row)
                self.global_names.add(parsed_name)
                self.parsed_color_rule_counts[parsed_rule] += row.raw_row_count

            if row.color_core:
                self.by_core_name[(row.fabric_name, row.color_core)].append(row)
                self.global_names.add(row.color_core)
            for alias in row.color_aliases:
                self.by_alias_name[(row.fabric_name, alias)].append(row)
                self.global_names.add(alias)
            if row.lingxing_code:
                self.by_code[(row.fabric_name, row.lingxing_code)].append(row)
            for record_id in row.record_ids:
                self.by_record_id[record_id] = row
            self.by_fabric[row.fabric_name].append(row)
            self.global_names.add(row.color_name)
            pair_identities[(row.fabric_name, row.color_name)].append(row)

        self.duplicate_business_rows = [
            {
                "面料名": row.fabric_name,
                "颜色": row.color_name,
                "领星新颜色缩写": row.lingxing_code,
                "原始行数": row.raw_row_count,
                "record_ids": list(row.record_ids),
            }
            for row in self.rows
            if row.raw_row_count > 1
        ]
        self.ambiguous_pairs = [
            {
                "面料名": fabric,
                "颜色": color,
                "候选缩写": sorted({row.lingxing_code for row in candidates}),
            }
            for (fabric, color), candidates in pair_identities.items()
            if len(candidates) > 1
        ]

    def audit(self) -> dict[str, Any]:
        rows_with_code = sum(row.raw_row_count for row in self.rows if row.lingxing_code)
        return {
            "source_record_count": self.source_record_count,
            "raw_fabric_association_count": self.raw_association_count,
            "unique_business_identity_count": len(self.rows),
            "unique_fabric_color_count": len(
                {(row.fabric_name, row.color_name) for row in self.rows}
            ),
            "rows_with_lingxing_code": rows_with_code,
            "parsed_color_label_row_count": sum(
                self.parsed_color_rule_counts.values()
            ),
            "parsed_color_rule_counts": dict(
                sorted(self.parsed_color_rule_counts.items())
            ),
            "parsed_color_label_policy": (
                "仅解析数字#中文名、数字-中文名；"
                "不做strip、别名、大小写或模糊归一"
            ),
            "exact_duplicate_business_group_count": len(self.duplicate_business_rows),
            "exact_duplicate_business_rows": self.duplicate_business_rows,
            "ambiguous_fabric_color_pair_count": len(self.ambiguous_pairs),
            "ambiguous_fabric_color_pairs": self.ambiguous_pairs,
            "strict_matching": True,
            "implicit_normalization": False,
            "target_fabric_count": len(self.by_fabric),
            "target_fabrics": list(self.by_fabric),
        }


def _unique(candidates: Iterable[CatalogRow]) -> list[CatalogRow]:
    by_identity = {row.identity: row for row in candidates}
    return list(by_identity.values())


class ManualMappingCatalog:
    """按原始四字段读取人工确认，不对审核键做隐式归一。"""

    def __init__(self, entries: Iterable[ManualMapping] = ()):
        self.entries = tuple(entries)
        self.active_by_key: MutableMapping[
            tuple[str, str, str, str], list[ManualMapping]
        ] = defaultdict(list)
        for entry in self.entries:
            if entry.enabled:
                self.active_by_key[entry.key].append(entry)

    @staticmethod
    def forecast_key(
        forecast: ForecastSku, fabric_name: str
    ) -> tuple[str, str, str, str]:
        return (
            fabric_name,
            forecast.color_code,
            forecast.color_name,
            forecast.color_system,
        )

    def resolve(
        self,
        forecast: ForecastSku,
        fabric_name: str,
        index: CatalogIndex,
    ) -> list[CatalogRow]:
        rows: list[CatalogRow] = []
        for entry in self.active_by_key.get(
            self.forecast_key(forecast, fabric_name), ()
        ):
            if entry.record_id:
                row = index.by_record_id.get(entry.record_id)
                if (
                    row
                    and row.fabric_name == fabric_name
                    and (
                        not entry.catalog_color_name
                        or row.color_name == entry.catalog_color_name
                    )
                ):
                    rows.append(row)
                continue
            if entry.catalog_color_name:
                rows.extend(
                    index.by_name.get(
                        (fabric_name, entry.catalog_color_name),
                        (),
                    )
                )
        return _unique(rows)

    def supports(
        self,
        forecast: ForecastSku,
        fabric_name: str,
        row: CatalogRow,
        index: CatalogIndex,
    ) -> bool:
        return row in self.resolve(forecast, fabric_name, index)

    def audit(self) -> dict[str, Any]:
        conflicting_keys = sum(
            1
            for entries in self.active_by_key.values()
            if len(
                {
                    (entry.record_id, entry.catalog_color_name)
                    for entry in entries
                }
            )
            > 1
        )
        return {
            "read_mapping_count": len(self.entries),
            "enabled_mapping_count": sum(
                len(entries) for entries in self.active_by_key.values()
            ),
            "enabled_mapping_key_count": len(self.active_by_key),
            "conflicting_mapping_key_count": conflicting_keys,
        }


def _pattern_safe_candidates(
    forecast: ForecastSku,
    candidates: Iterable[CatalogRow],
) -> list[CatalogRow]:
    source = forecast.color_name or forecast.product_name
    return _unique(
        row
        for row in candidates
        if pattern_compatible(source, row.color_name)
    )


def _deterministic_alias_values(
    forecast: ForecastSku,
    governance_catalog: ColorMappingCatalog,
) -> set[str]:
    values: set[str] = set()
    if forecast.color_name:
        values.add(forecast.color_name)
        core, aliases = catalog_color_components(forecast.color_name)
        values.add(core)
        values.update(aliases)
    systems = (
        (forecast.color_system,)
        if forecast.color_system in SUPPORTED_SYSTEMS
        else ()
    )
    for system in systems:
        for entry in governance_catalog.entries_for_code(
            system,
            forecast.color_code,
        ):
            if entry.chinese:
                values.add(entry.chinese)
    return {value for value in values if value}


def match_catalog_row(
    forecast: ForecastSku,
    fabric_name: str,
    index: CatalogIndex,
    governance_catalog: ColorMappingCatalog,
    manual_catalog: ManualMappingCatalog | None = None,
) -> MatchDecision:
    """执行自动确认规则；模糊相似度永远不在这里参与决策。"""
    manual_catalog = manual_catalog or ManualMappingCatalog()
    if forecast.field_issue:
        return MatchDecision(
            None,
            reason_code=REASON_FORECAST_CONFLICT,
            reason=forecast.field_issue,
        )

    if fabric_name not in index.by_fabric:
        return MatchDecision(
            None,
            reason_code=REASON_FABRIC_INACTIVE,
            reason=REASON_FABRIC_INACTIVE,
        )

    name_candidates = (
        _pattern_safe_candidates(
            forecast,
            index.by_name.get((fabric_name, forecast.color_name), ()),
        )
        if forecast.color_name
        else []
    )
    if len(name_candidates) == 1:
        return MatchDecision(name_candidates[0], method=MATCH_NAME)

    code_candidates = (
        _pattern_safe_candidates(
            forecast,
            index.by_code.get((fabric_name, forecast.color_code), ()),
        )
        if forecast.color_code
        else []
    )
    if len(code_candidates) == 1:
        return MatchDecision(code_candidates[0], method=MATCH_CODE)

    parsed_name_candidates = (
        _pattern_safe_candidates(
            forecast,
            index.by_parsed_name.get(
                (fabric_name, forecast.color_name),
                (),
            )
        )
        if forecast.color_name
        else []
    )
    if len(parsed_name_candidates) == 1:
        return MatchDecision(
            parsed_name_candidates[0],
            method=MATCH_CATALOG_LABEL,
        )

    system_candidates: list[CatalogRow] = []
    mapping_found = False
    if forecast.color_system in SUPPORTED_SYSTEMS:
        mapped_by_code = governance_catalog.lookup(
            forecast.color_system, forecast.color_code
        )
        if mapped_by_code:
            mapping_found = True
            if mapped_by_code.chinese:
                system_candidates.extend(
                    index.by_name.get(
                        (fabric_name, mapped_by_code.chinese),
                        (),
                    )
                )
                system_candidates.extend(
                    index.by_parsed_name.get(
                        (fabric_name, mapped_by_code.chinese),
                        (),
                    )
                )

        if forecast.color_name:
            mapped_by_name = governance_catalog.entries_for_name(
                forecast.color_system, forecast.color_name
            )
            if mapped_by_name:
                mapping_found = True
            for entry in mapped_by_name:
                system_candidates.extend(
                    index.by_code.get((fabric_name, entry.code), ())
                )

        system_candidates = _pattern_safe_candidates(
            forecast,
            system_candidates,
        )
        if len(system_candidates) == 1:
            return MatchDecision(system_candidates[0], method=MATCH_SYSTEM)

    core_alias_candidates: list[CatalogRow] = []
    for value in _deterministic_alias_values(forecast, governance_catalog):
        core_alias_candidates.extend(
            index.by_core_name.get((fabric_name, value), ())
        )
        core_alias_candidates.extend(
            index.by_alias_name.get((fabric_name, value), ())
        )
    core_alias_candidates = _pattern_safe_candidates(
        forecast,
        core_alias_candidates,
    )
    if len(core_alias_candidates) == 1:
        return MatchDecision(
            core_alias_candidates[0],
            method=MATCH_CORE_ALIAS,
        )

    manual_candidates = manual_catalog.resolve(
        forecast,
        fabric_name,
        index,
    )
    if len(manual_candidates) == 1:
        return MatchDecision(manual_candidates[0], method=MATCH_MANUAL)

    ambiguous_candidates = _unique(
        [
            *name_candidates,
            *code_candidates,
            *parsed_name_candidates,
            *system_candidates,
            *core_alias_candidates,
            *manual_candidates,
        ]
    )
    if len(manual_candidates) > 1:
        return MatchDecision(
            None,
            reason_code=REASON_MANUAL_AMBIGUOUS,
            reason=REASON_MANUAL_AMBIGUOUS,
            candidates=tuple(manual_candidates),
        )
    if len(ambiguous_candidates) > 1:
        return MatchDecision(
            None,
            reason_code=REASON_AMBIGUOUS,
            reason=REASON_AMBIGUOUS,
            candidates=tuple(ambiguous_candidates),
        )
    source_pattern_groups = pattern_groups(
        forecast.color_name or forecast.product_name
    )
    if source_pattern_groups and not any(
        pattern_compatible(forecast.color_name or forecast.product_name, row.color_name)
        for row in index.by_fabric[fabric_name]
    ):
        return MatchDecision(
            None,
            reason_code=REASON_PATTERN_MISSING,
            reason=REASON_PATTERN_MISSING,
        )
    if forecast.color_system == UNKNOWN_SYSTEM:
        return MatchDecision(
            None,
            reason_code=REASON_PENDING_SYSTEM,
            reason=REASON_PENDING_SYSTEM,
        )
    if not mapping_found:
        return MatchDecision(
            None,
            reason_code=REASON_SYSTEM_MAPPING,
            reason=REASON_SYSTEM_MAPPING,
        )
    if forecast.color_name and forecast.color_name not in index.global_names:
        return MatchDecision(
            None,
            reason_code=REASON_NAME_UNKNOWN,
            reason=REASON_NAME_UNKNOWN,
        )
    return MatchDecision(
        None,
        reason_code=REASON_COLOR_INACTIVE,
        reason=REASON_COLOR_INACTIVE,
    )


def _candidate_governance_support(
    forecast: ForecastSku,
    row: CatalogRow,
    governance_catalog: ColorMappingCatalog,
) -> tuple[bool, bool]:
    target_variants = set()
    for value in (row.color_name, row.color_core, *row.color_aliases):
        target_variants.update(fuzzy_name_variants(value))

    systems = (
        (forecast.color_system,)
        if forecast.color_system in SUPPORTED_SYSTEMS
        else SUPPORTED_SYSTEMS
    )
    governance_supported = False
    system_supported = False
    for system in systems:
        for entry in governance_catalog.entries_for_code(
            system,
            forecast.color_code,
        ):
            entry_variants = fuzzy_name_variants(entry.chinese)
            supported = (
                bool(entry_variants & target_variants)
                or bool(entry.code and entry.code == row.lingxing_code)
            )
            governance_supported = governance_supported or supported
            if system == forecast.color_system and supported:
                system_supported = True
    return governance_supported, system_supported


def recommend_fuzzy_candidates(
    forecast: ForecastSku,
    fabric_name: str,
    index: CatalogIndex,
    governance_catalog: ColorMappingCatalog,
    manual_catalog: ManualMappingCatalog | None = None,
    config: FuzzyMatchConfig | None = None,
) -> list[FuzzyCandidate]:
    """给同面料未匹配关联生成候选；返回值只允许进入人工审核表。"""
    manual_catalog = manual_catalog or ManualMappingCatalog()
    config = config or FuzzyMatchConfig()
    config.validate()
    source_variants = fuzzy_name_variants(forecast.color_name)
    for system in SUPPORTED_SYSTEMS:
        for entry in governance_catalog.entries_for_code(
            system,
            forecast.color_code,
        ):
            source_variants.update(fuzzy_name_variants(entry.chinese))
    if not source_variants:
        return []

    source_aliases = catalog_color_components(forecast.color_name)[1]
    normalized_source_aliases = {
        alias
        for value in source_aliases
        for alias in fuzzy_name_variants(value)
    }
    result: list[FuzzyCandidate] = []
    for row in index.by_fabric.get(fabric_name, ()):
        if any(keyword in row.color_name for keyword in INVALID_CATALOG_KEYWORDS):
            continue
        if not pattern_compatible(
            forecast.color_name or forecast.product_name,
            row.color_name,
        ):
            continue

        target_variants: set[str] = set()
        for value in (row.color_name, row.color_core, *row.color_aliases):
            target_variants.update(fuzzy_name_variants(value))
        if not target_variants:
            continue
        similarity = max(
            difflib.SequenceMatcher(None, source, target).ratio()
            for source in source_variants
            for target in target_variants
        )
        score = similarity * config.name_weight
        basis = [f"标准化中文相似度{similarity * 100:.1f}"]

        normalized_target_aliases = {
            alias
            for value in row.color_aliases
            for alias in fuzzy_name_variants(value)
        }
        if normalized_source_aliases & target_variants or (
            source_variants & normalized_target_aliases
        ):
            score += config.alias_weight
            basis.append("括号内别名精确")
        if (
            forecast.color_code
            and row.lingxing_code
            and forecast.color_code == row.lingxing_code
        ):
            score += config.code_weight
            basis.append("领星缩写一致")

        governance_supported, system_supported = _candidate_governance_support(
            forecast,
            row,
            governance_catalog,
        )
        if governance_supported:
            score += config.governance_weight
            basis.append("颜色编制表支持")
        if system_supported:
            score += config.system_weight
            basis.append("颜色体系一致")
        if manual_catalog.supports(forecast, fabric_name, row, index):
            score += config.manual_weight
            basis.append("历史人工确认支持")

        score = round(min(score, 100.0), 2)
        if score >= config.low_min_score:
            result.append(
                FuzzyCandidate(
                    row=row,
                    score=score,
                    basis="；".join(basis),
                )
            )
    result.sort(
        key=lambda item: (
            -item.score,
            item.row.color_name,
            item.row.lingxing_code,
        )
    )
    return result[: config.max_candidates]


def fuzzy_confidence(
    candidates: Sequence[FuzzyCandidate],
    config: FuzzyMatchConfig,
) -> tuple[str, float | str]:
    if not candidates:
        return "", ""
    top_score = candidates[0].score
    gap: float | str = (
        round(top_score - candidates[1].score, 2)
        if len(candidates) > 1
        else ""
    )
    if top_score >= config.high_min_score and (
        gap == "" or float(gap) >= config.high_lead_margin
    ):
        return "高", gap
    if top_score >= config.medium_min_score or (
        top_score >= config.high_min_score
        and gap != ""
        and float(gap) < config.high_lead_margin
    ):
        return "中", gap
    return "低", gap


def aggregate_stocking(
    forecasts: Sequence[ForecastSku],
    fabric_usage_by_spu: Mapping[
        str,
        Sequence[FabricUsage | str | Mapping[str, Any]],
    ],
    catalog_rows: Sequence[CatalogRow],
    governance_catalog: ColorMappingCatalog,
    source_record_count: int | None = None,
    manual_catalog: ManualMappingCatalog | None = None,
    fuzzy_config: FuzzyMatchConfig | None = None,
    target_fabrics: Sequence[str] | None = None,
) -> StockingResult:
    """计算 17 面料米数，并隔离确定性匹配、模糊候选和参数缺失。"""
    manual_catalog = manual_catalog or ManualMappingCatalog()
    fuzzy_config = fuzzy_config or FuzzyMatchConfig()
    fuzzy_config.validate()

    if target_fabrics is None:
        target_fabrics = tuple(
            dict.fromkeys(row.fabric_name for row in catalog_rows)
        )
    target_fabrics = tuple(target_fabrics)
    target_set = set(target_fabrics)
    target_order = {
        fabric_name: index
        for index, fabric_name in enumerate(target_fabrics)
    }
    target_catalog_rows = [
        row for row in catalog_rows if row.fabric_name in target_set
    ]
    index = CatalogIndex(
        target_catalog_rows,
        source_record_count=source_record_count,
    )

    def coerce_usage(
        spu: str,
        item: FabricUsage | str | Mapping[str, Any],
    ) -> FabricUsage:
        if isinstance(item, FabricUsage):
            return item
        if isinstance(item, str):
            # 仅保留旧单元测试/调用方兼容；生产读取始终传入显式单耗。
            return FabricUsage(spu, item, 1.0, 1.0)
        return FabricUsage(
            spu=str(item.get("SPU") or item.get("spu") or spu),
            fabric_name=str(
                item.get("面料")
                or item.get("fabric_name")
                or ""
            ),
            unit_usage=float(
                item.get("单件用量")
                or item.get("unit_usage")
                or 0
            ),
            loss_coefficient=float(
                item.get("单件损耗")
                or item.get("loss_coefficient")
                or 1
            ),
        )

    usage_index: dict[str, list[FabricUsage]] = defaultdict(list)
    for spu, values in fabric_usage_by_spu.items():
        for item in values:
            usage = coerce_usage(str(spu), item)
            if (
                usage.fabric_name in target_set
                and usage.fabric_name
                and all(
                    existing.fabric_name != usage.fabric_name
                    for existing in usage_index[str(spu).strip()]
                )
            ):
                usage_index[str(spu).strip()].append(usage)

    buckets: dict[tuple[str, str, str], dict[str, Any]] = {
        row.identity: {
            "row": row,
            "qty": 0.0,
            "meters": 0.0,
            "skus": set(),
            "spus": set(),
            "methods": set(),
        }
        for row in index.rows
    }
    fabric_stats: dict[str, dict[str, Any]] = {
        fabric_name: {
            "catalog_color_count": len(index.by_fabric.get(fabric_name, ())),
            "assignment_count": 0,
            "forecast_qty": 0.0,
            "calculable_meters": 0.0,
            "confirmed_meters": 0.0,
            "fuzzy_meters": 0.0,
            "pending_meters": 0.0,
            "missing_usage_qty": 0.0,
        }
        for fabric_name in target_fabrics
    }

    auto_merge_rows: list[dict[str, Any]] = []
    manual_review_rows: list[dict[str, Any]] = []
    usage_missing_rows: list[dict[str, Any]] = []
    fuzzy_review_rows: list[dict[str, Any]] = []
    priority_by_sku: dict[str, dict[str, Any]] = {}
    method_counts: Counter[str] = Counter()
    method_meters: Counter[str] = Counter()
    reason_counts: Counter[str] = Counter()
    confidence_assignment_counts: Counter[str] = Counter()
    candidate_count_distribution: Counter[int] = Counter()
    score_distribution: Counter[str] = Counter()
    target_skus: set[str] = set()
    matched_skus: set[str] = set()
    failed_skus: set[str] = set()
    fuzzy_skus: set[str] = set()
    confidence_skus: MutableMapping[str, set[str]] = defaultdict(set)
    pattern_unmatched_skus: set[str] = set()
    pattern_unmatched_meters = 0.0
    matched_assignments = 0
    total_assignments = 0
    excluded_no_target_usage_skus = 0

    for forecast in forecasts:
        usages = usage_index.get(str(forecast.spu).strip(), ())
        if not usages:
            excluded_no_target_usage_skus += 1
            continue
        target_skus.add(forecast.sku)

        for usage in usages:
            total_assignments += 1
            stats = fabric_stats[usage.fabric_name]
            stats["assignment_count"] += 1
            stats["forecast_qty"] += forecast.forecast_qty
            meters = usage.calculate_meters(forecast.forecast_qty)
            if usage.usage_missing:
                stats["missing_usage_qty"] += forecast.forecast_qty
            else:
                stats["calculable_meters"] += meters

            decision = match_catalog_row(
                forecast,
                usage.fabric_name,
                index,
                governance_catalog,
                manual_catalog=manual_catalog,
            )
            match_status = (
                f"已匹配：{decision.method}"
                if decision.row
                else f"待确认：{decision.reason}"
            )
            if usage.usage_missing:
                usage_missing_rows.append(
                    {
                        "SPU": forecast.spu,
                        "SKU": forecast.sku,
                        "面料": usage.fabric_name,
                        "预估销量": _quantity(forecast.forecast_qty),
                        "单件用量": _quantity(usage.unit_usage),
                        "单件损耗": _quantity(usage.loss_coefficient),
                        "缺失原因": REASON_USAGE_MISSING,
                        "颜色匹配状态": match_status,
                    }
                )

            if decision.row:
                matched_assignments += 1
                matched_skus.add(forecast.sku)
                method_counts[decision.method] += 1
                if not usage.usage_missing:
                    stats["confirmed_meters"] += meters
                    method_meters[decision.method] += meters
                bucket = buckets[decision.row.identity]
                bucket["qty"] += forecast.forecast_qty
                bucket["meters"] += meters
                bucket["skus"].add(forecast.sku)
                bucket["spus"].add(forecast.spu)
                bucket["methods"].add(decision.method)
                auto_merge_rows.append(
                    {
                        "SKU": forecast.sku,
                        "SPU": forecast.spu,
                        "品名": forecast.product_name,
                        "面料名": usage.fabric_name,
                        "原颜色编码": forecast.color_code,
                        "原颜色中文名": forecast.color_name,
                        "颜色体系": forecast.color_system,
                        "预估销量": _quantity(forecast.forecast_qty),
                        "单件用量": _quantity(usage.unit_usage),
                        "单件损耗": _quantity(usage.loss_coefficient),
                        "预估面料用量（米）": _meters(meters),
                        "飞书颜色原值": decision.row.color_name,
                        "领星新颜色缩写": decision.row.lingxing_code,
                        "匹配方式": decision.method,
                        "匹配依据": (
                            "历史审核文件精确键"
                            if decision.method == MATCH_MANUAL
                            else f"确定性规则：{decision.method}"
                        ),
                        "飞书记录ID": "、".join(decision.row.record_ids),
                        "用量参数状态": (
                            REASON_USAGE_MISSING
                            if usage.usage_missing
                            else "正常"
                        ),
                    }
                )
                continue

            failed_skus.add(forecast.sku)
            reason_counts[decision.reason_code] += 1
            if not usage.usage_missing:
                stats["pending_meters"] += meters

            candidates = recommend_fuzzy_candidates(
                forecast,
                usage.fabric_name,
                index,
                governance_catalog,
                manual_catalog=manual_catalog,
                config=fuzzy_config,
            )
            confidence, lead_gap = fuzzy_confidence(candidates, fuzzy_config)
            if candidates:
                fuzzy_skus.add(forecast.sku)
                confidence_assignment_counts[confidence] += 1
                confidence_skus[confidence].add(forecast.sku)
                candidate_count_distribution[len(candidates)] += 1
                top_score = candidates[0].score
                if top_score >= 90:
                    score_distribution["90-100"] += 1
                elif top_score >= 80:
                    score_distribution["80-89"] += 1
                else:
                    score_distribution["70-79"] += 1
                if not usage.usage_missing:
                    stats["fuzzy_meters"] += meters

            source_is_pattern = bool(
                pattern_groups(forecast.color_name or forecast.product_name)
            )
            if source_is_pattern:
                pattern_unmatched_skus.add(forecast.sku)
                if not usage.usage_missing:
                    pattern_unmatched_meters += meters

            if source_is_pattern and not candidates:
                risk = "缺少对应图案色号；禁止推荐纯色"
            elif len(candidates) > 1 and (
                lead_gap != ""
                and float(lead_gap) < fuzzy_config.high_lead_margin
            ):
                risk = "多个相近候选，需人工确认"
            elif candidates:
                risk = "模糊候选不得计入已确认颜色用量"
            elif decision.candidates:
                risk = "确定性规则存在多个候选，禁止自动选择"
            else:
                risk = "无达到阈值的同面料候选"

            displayed_candidates = (
                [candidate.row for candidate in candidates]
                or list(decision.candidates[:3])
            )
            candidate_text = "；".join(
                f"{row.color_name}[{','.join(row.record_ids)}]"
                for row in displayed_candidates
            )
            manual_review_rows.append(
                {
                    "SKU": forecast.sku,
                    "SPU": forecast.spu,
                    "面料": usage.fabric_name,
                    "原颜色编码": forecast.color_code,
                    "原颜色中文名": forecast.color_name,
                    "颜色体系": forecast.color_system,
                    "预估销量": _quantity(forecast.forecast_qty),
                    "已算面料米数": _meters(meters),
                    "候选飞书颜色": candidate_text,
                    "未自动归并原因": decision.reason,
                }
            )

            fuzzy_row: dict[str, Any] = {
                "面料名": usage.fabric_name,
                "SKU": forecast.sku,
                "SPU": forecast.spu,
                "品名": forecast.product_name,
                "原始颜色编码": forecast.color_code,
                "原始颜色中文名": forecast.color_name,
                "颜色体系": forecast.color_system,
                "预估销量": _quantity(forecast.forecast_qty),
                "预估面料用量（米）": _meters(meters),
                "第一、第二候选分差": lead_gap,
                "置信度等级": confidence,
                "风险提示": risk,
                "人工审核结果": "",
                "人工确认飞书记录ID": "",
                "审核备注": "",
            }
            for position in range(1, 4):
                candidate = (
                    candidates[position - 1]
                    if position <= len(candidates)
                    else None
                )
                fuzzy_row[f"候选{position}飞书颜色原值"] = (
                    candidate.row.color_name if candidate else ""
                )
                fuzzy_row[f"候选{position}领星缩写"] = (
                    candidate.row.lingxing_code if candidate else ""
                )
                fuzzy_row[f"候选{position}分数"] = (
                    candidate.score if candidate else ""
                )
                fuzzy_row[f"候选{position}依据"] = (
                    candidate.basis if candidate else ""
                )
                fuzzy_row[f"候选{position}飞书记录ID"] = (
                    "、".join(candidate.row.record_ids)
                    if candidate
                    else ""
                )
            fuzzy_review_rows.append(fuzzy_row)

            if (
                forecast.color_system == UNKNOWN_SYSTEM
                and decision.reason_code == REASON_PENDING_SYSTEM
            ):
                priority_by_sku.setdefault(
                    forecast.sku,
                    {
                        "SKU": forecast.sku,
                        "品名": forecast.product_name,
                        "颜色编码": forecast.color_code,
                        "颜色中文名": forecast.color_name,
                        "预估销量": _quantity(forecast.forecast_qty),
                    },
                )

    def fabric_sort_key(row: Mapping[str, Any]) -> tuple[int, str, str]:
        fabric_name = str(row.get("面料名") or row.get("面料") or "")
        color_name = str(
            row.get("飞书颜色原值")
            or row.get("原颜色中文名")
            or ""
        )
        return (
            target_order.get(fabric_name, len(target_order)),
            fabric_name,
            color_name,
        )

    color_usage_rows: list[dict[str, Any]] = []
    catalog_output_rows: list[dict[str, Any]] = []
    for row in sorted(
        index.rows,
        key=lambda item: (
            target_order.get(item.fabric_name, len(target_order)),
            item.color_name,
            item.lingxing_code,
        ),
    ):
        bucket = buckets[row.identity]
        methods = [
            method
            for method in MATCH_METHOD_ORDER
            if method in bucket["methods"]
        ]
        base_row = {
            "面料名": row.fabric_name,
            "飞书颜色原值": row.color_name,
            "飞书颜色中文核心": row.color_core,
            "领星新颜色缩写": row.lingxing_code,
            "飞书记录ID": "、".join(row.record_ids),
            "原始重复行数": row.raw_row_count,
        }
        color_usage_rows.append(
            {
                **base_row,
                "预估销量": _quantity(bucket["qty"]),
                "预估面料用量（米）": _meters(bucket["meters"]),
                "关联SKU数": len(bucket["skus"]),
                "关联SPU数": len(bucket["spus"]),
                "匹配方式": "、".join(methods),
            }
        )
        catalog_output_rows.append(
            {
                **base_row,
                "括号内颜色别名": "、".join(row.color_aliases),
                "当前是否有需求": "是" if bucket["qty"] else "否",
            }
        )

    def coverage(numerator: float, denominator: float) -> float:
        return round(numerator / denominator * 100, 2) if denominator else 0.0

    total_stats = {
        key: sum(float(stats[key]) for stats in fabric_stats.values())
        for key in (
            "assignment_count",
            "forecast_qty",
            "calculable_meters",
            "confirmed_meters",
            "fuzzy_meters",
            "pending_meters",
            "missing_usage_qty",
        )
    }
    total_stats["catalog_color_count"] = sum(
        int(stats["catalog_color_count"])
        for stats in fabric_stats.values()
    )

    def overview_row(
        scope: str,
        fabric_name: str,
        stats: Mapping[str, Any],
    ) -> dict[str, Any]:
        calculable = float(stats["calculable_meters"])
        confirmed = float(stats["confirmed_meters"])
        fuzzy = float(stats["fuzzy_meters"])
        reasons: list[str] = []
        if scope != "总计":
            if not stats["catalog_color_count"]:
                reasons.append("飞书清单无颜色记录")
            if not stats["assignment_count"]:
                reasons.append("当前预测无对应SPU-面料需求")
            elif not calculable and stats["missing_usage_qty"]:
                reasons.append(REASON_USAGE_MISSING)
            elif calculable and not confirmed:
                reasons.append("颜色均待人工确认")
        return {
            "范围": scope,
            "面料名": fabric_name,
            "飞书颜色数（去重后）": int(stats["catalog_color_count"]),
            "可确定总用量（米）": _meters(calculable),
            "已分配到具体飞书颜色用量（米）": _meters(confirmed),
            "待人工确认颜色用量（米）": _meters(
                float(stats["pending_meters"])
            ),
            "颜色分配覆盖率": coverage(confirmed, calculable),
            "模糊候选涉及用量（米）": _meters(fuzzy),
            "模糊候选覆盖率": coverage(fuzzy, calculable),
            "潜在颜色覆盖率": coverage(confirmed + fuzzy, calculable),
            "缺少单件用量的销量": _quantity(
                float(stats["missing_usage_qty"])
            ),
            "缺失原因": "；".join(reasons),
        }

    overview_rows = [overview_row("总计", "17个目标面料", total_stats)]
    overview_rows.extend(
        overview_row("面料", fabric_name, fabric_stats[fabric_name])
        for fabric_name in target_fabrics
    )

    auto_merge_rows.sort(key=fabric_sort_key)
    manual_review_rows.sort(key=fabric_sort_key)
    usage_missing_rows.sort(key=fabric_sort_key)
    fuzzy_review_rows.sort(key=fabric_sort_key)
    priority_rows = sorted(
        priority_by_sku.values(),
        key=lambda row: (-float(row["预估销量"]), row["SKU"]),
    )
    pending_qty = sum(float(row["预估销量"]) for row in priority_rows)
    fully_matched_skus = target_skus - failed_skus
    partially_matched_skus = matched_skus & failed_skus
    audit = index.audit()
    audit["target_raw_record_count"] = len(target_catalog_rows)
    audit["expected_target_fabric_count"] = len(target_fabrics)
    audit["missing_target_fabrics"] = [
        name for name in target_fabrics if name not in index.by_fabric
    ]

    metrics = {
        "target_fabric_count": len(target_fabrics),
        "target_fabrics": list(target_fabrics),
        "target_catalog_raw_record_count": len(target_catalog_rows),
        "target_catalog_unique_color_row_count": len(index.rows),
        "input_sku_count": len(forecasts),
        "input_forecast_qty": _quantity(
            sum(row.forecast_qty for row in forecasts)
        ),
        "target_scope_sku_count": len(target_skus),
        "excluded_no_target_fabric_usage_sku_count": excluded_no_target_usage_skus,
        "fully_matched_sku_count": len(fully_matched_skus),
        "partially_matched_sku_count": len(partially_matched_skus),
        "unmatched_sku_count": len(failed_skus),
        "sku_full_match_rate_pct": coverage(
            len(fully_matched_skus),
            len(target_skus),
        ),
        "fabric_assignment_count": total_assignments,
        "matched_fabric_assignment_count": matched_assignments,
        "unmatched_fabric_assignment_count": (
            total_assignments - matched_assignments
        ),
        "fabric_assignment_match_rate_pct": coverage(
            matched_assignments,
            total_assignments,
        ),
        "catalog_fabric_name_count": len(index.by_fabric),
        "forecast_fabric_name_count": sum(
            1
            for stats in fabric_stats.values()
            if stats["assignment_count"]
        ),
        "exact_fabric_name_overlap_count": sum(
            1
            for fabric_name, stats in fabric_stats.items()
            if stats["assignment_count"] and fabric_name in index.by_fabric
        ),
        "in_scope_fabric_assignment_match_rate_pct": coverage(
            matched_assignments,
            total_assignments,
        ),
        "calculable_total_meters": _meters(
            total_stats["calculable_meters"]
        ),
        "confirmed_color_meters": _meters(
            total_stats["confirmed_meters"]
        ),
        "manual_review_color_meters": _meters(
            total_stats["pending_meters"]
        ),
        "confirmed_color_coverage_pct": coverage(
            total_stats["confirmed_meters"],
            total_stats["calculable_meters"],
        ),
        "fuzzy_candidate_sku_count": len(fuzzy_skus),
        "fuzzy_candidate_assignment_count": sum(
            confidence_assignment_counts.values()
        ),
        "fuzzy_candidate_meters": _meters(total_stats["fuzzy_meters"]),
        "fuzzy_candidate_coverage_pct": coverage(
            total_stats["fuzzy_meters"],
            total_stats["calculable_meters"],
        ),
        "potential_color_coverage_pct": coverage(
            total_stats["confirmed_meters"] + total_stats["fuzzy_meters"],
            total_stats["calculable_meters"],
        ),
        "fuzzy_confidence_sku_counts": {
            level: len(confidence_skus.get(level, set()))
            for level in ("高", "中", "低")
        },
        "fuzzy_confidence_assignment_counts": {
            level: confidence_assignment_counts.get(level, 0)
            for level in ("高", "中", "低")
        },
        "fuzzy_candidate_count_distribution": {
            str(count): candidate_count_distribution.get(count, 0)
            for count in (1, 2, 3)
        },
        "fuzzy_score_distribution": {
            bucket: score_distribution.get(bucket, 0)
            for bucket in ("70-79", "80-89", "90-100")
        },
        "fuzzy_thresholds": {
            "high_min_score": fuzzy_config.high_min_score,
            "medium_min_score": fuzzy_config.medium_min_score,
            "low_min_score": fuzzy_config.low_min_score,
            "high_lead_margin": fuzzy_config.high_lead_margin,
            "max_candidates": fuzzy_config.max_candidates,
        },
        "pattern_unmatched_sku_count": len(pattern_unmatched_skus),
        "pattern_unmatched_meters": _meters(pattern_unmatched_meters),
        "missing_usage_forecast_qty": _quantity(
            total_stats["missing_usage_qty"]
        ),
        "usage_missing_assignment_count": len(usage_missing_rows),
        "match_method_counts": {
            method: method_counts.get(method, 0)
            for method in MATCH_METHOD_ORDER
        },
        "match_method_meters": {
            method: _meters(method_meters.get(method, 0))
            for method in MATCH_METHOD_ORDER
        },
        "unmatched_reason_counts": dict(sorted(reason_counts.items())),
        "pending_system_unmatched_sku_count": len(priority_rows),
        "pending_system_unmatched_forecast_qty": _quantity(pending_qty),
        "manual_mapping_audit": manual_catalog.audit(),
        "per_fabric_coverage": overview_rows[1:],
        "color_usage_output_row_count": len(color_usage_rows),
        "auto_merge_output_row_count": len(auto_merge_rows),
        "manual_review_output_row_count": len(manual_review_rows),
        "catalog_audit": audit,
    }
    return StockingResult(
        overview_rows=overview_rows,
        color_usage_rows=color_usage_rows,
        auto_merge_rows=auto_merge_rows,
        manual_review_rows=manual_review_rows,
        usage_missing_rows=usage_missing_rows,
        catalog_rows=catalog_output_rows,
        fuzzy_review_rows=fuzzy_review_rows,
        priority_rows=priority_rows,
        metrics=metrics,
    )


def _table_columns(table_name: str) -> set[str]:
    with db_cursor() as cursor:
        cursor.execute(
            """
            SELECT COLUMN_NAME
            FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA=DATABASE() AND TABLE_NAME=%s
            """,
            (table_name,),
        )
        return {str(row.get("COLUMN_NAME") or "") for row in cursor.fetchall()}


def _month_start(value: date, delta: int) -> date:
    month_index = value.year * 12 + value.month - 1 + delta
    return date(month_index // 12, month_index % 12 + 1, 1)


def _load_snapshot_rows() -> list[dict[str, Any]]:
    if not _table_columns(SNAPSHOT_TABLE):
        logger.warning(f"{SNAPSHOT_TABLE} 不存在，无法补齐 SKU 颜色主数据")
        return []
    with db_cursor() as cursor:
        cursor.execute(
            f"""
            SELECT sku, spu, product_name, custom_fields_json
            FROM `{SNAPSHOT_TABLE}`
            WHERE sku IS NOT NULL AND CHAR_LENGTH(TRIM(CAST(sku AS CHAR))) > 0
            """
        )
        return list(cursor.fetchall())


def _snapshot_index(
    rows: Sequence[Mapping[str, Any]],
) -> tuple[dict[str, Mapping[str, Any]], set[str]]:
    result: dict[str, Mapping[str, Any]] = {}
    conflicts: set[str] = set()
    for row in rows:
        sku = normalize_sku(row.get("sku"))
        if not sku:
            continue
        previous = result.get(sku)
        if previous is None:
            result[sku] = row
            continue
        old_values = (
            str(previous.get("product_name") or ""),
            str(previous.get("custom_fields_json") or ""),
        )
        new_values = (
            str(row.get("product_name") or ""),
            str(row.get("custom_fields_json") or ""),
        )
        if old_values != new_values:
            conflicts.add(sku)
    return result, conflicts


def load_forecast_skus(
    governance_catalog: ColorMappingCatalog,
    as_of: date | None = None,
) -> tuple[list[ForecastSku], dict[str, Any]]:
    """读取并核对 SKU 预估输出字段，缺列时复用现有主数据派生。"""
    as_of = as_of or date.today()
    columns = _table_columns(FORECAST_TABLE)
    required = {"SKU", "SPU", "统计日期", "系统预测销量"}
    missing_required = sorted(required - columns)
    if missing_required:
        raise RuntimeError(
            f"{FORECAST_TABLE} 缺少核心字段: {', '.join(missing_required)}"
        )

    color_fields = ("品名", "颜色编码", "颜色中文名", "颜色体系")
    available_color_fields = [field for field in color_fields if field in columns]
    select_fields = ["SKU", "SPU", *available_color_fields]
    select_sql = ", ".join(f"`{field}`" for field in select_fields)
    group_sql = ", ".join(f"`{field}`" for field in select_fields)
    start = _month_start(as_of, 0)
    end = _month_start(as_of, 4)
    with db_cursor() as cursor:
        cursor.execute(
            f"""
            SELECT {select_sql}, SUM(`系统预测销量`) AS `预估销量`
            FROM `{FORECAST_TABLE}`
            WHERE `统计日期` >= %s AND `统计日期` < %s
              AND `系统预测销量` > 0
            GROUP BY {group_sql}
            """,
            (start, end),
        )
        rows = list(cursor.fetchall())

    snapshot_rows = _load_snapshot_rows()
    snapshot_by_sku, snapshot_conflicts = _snapshot_index(snapshot_rows)
    resolver = ColorSystemResolver(snapshot_rows)
    grouped: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for row in rows:
        sku = normalize_sku(row.get("SKU"))
        if sku:
            grouped[sku].append(row)

    forecasts: list[ForecastSku] = []
    field_conflict_count = 0
    for sku, sku_rows in grouped.items():
        snapshot = snapshot_by_sku.get(sku, {})
        identity = resolver.resolve(
            sku,
            str(sku_rows[0].get("SPU") or snapshot.get("spu") or ""),
        )
        values: dict[str, set[str]] = {
            field_name: {
                str(row.get(field_name))
                for row in sku_rows
                if row.get(field_name) not in (None, "")
            }
            for field_name in color_fields
        }
        conflicts = [name for name, candidates in values.items() if len(candidates) > 1]
        if sku in snapshot_conflicts:
            conflicts.append("产品快照")
        field_issue = ""
        if conflicts:
            field_conflict_count += 1
            field_issue = (
                f"{REASON_FORECAST_CONFLICT}：{','.join(sorted(set(conflicts)))}"
            )

        def direct(field_name: str) -> str:
            candidates = values.get(field_name) or set()
            return next(iter(candidates)) if len(candidates) == 1 else ""

        product_name = direct("品名") or str(snapshot.get("product_name") or "")
        color_code = direct("颜色编码") or identity.color_code
        color_system = direct("颜色体系")
        if color_system not in SUPPORTED_SYSTEMS:
            color_system = identity.color_system
        color_name = direct("颜色中文名") or parse_lingxing_color(product_name)
        if not color_name and color_system in SUPPORTED_SYSTEMS:
            entry = governance_catalog.lookup(color_system, color_code)
            color_name = entry.chinese if entry else ""

        forecasts.append(
            ForecastSku(
                sku=sku,
                spu=str(
                    sku_rows[0].get("SPU")
                    or snapshot.get("spu")
                    or identity.spu
                ),
                product_name=product_name,
                color_code=color_code,
                color_name=color_name,
                color_system=color_system or UNKNOWN_SYSTEM,
                forecast_qty=sum(float(row.get("预估销量") or 0) for row in sku_rows),
                field_issue=field_issue,
            )
        )

    audit = {
        "forecast_table": FORECAST_TABLE,
        "forecast_core_fields": sorted(required),
        "requested_color_fields": list(color_fields),
        "persisted_color_fields_found": available_color_fields,
        "color_field_mode": (
            "预测表直接读取"
            if len(available_color_fields) == len(color_fields)
            else "产品快照+现有解析器兼容补齐"
        ),
        "forecast_window_start": start.isoformat(),
        "forecast_window_end_exclusive": end.isoformat(),
        "forecast_grouped_sku_count": len(forecasts),
        "forecast_color_field_conflict_sku_count": field_conflict_count,
        "sample": [
            {
                "SKU": row.sku,
                "品名": row.product_name,
                "颜色编码": row.color_code,
                "颜色中文名": row.color_name,
                "颜色体系": row.color_system,
                "预估销量": _quantity(row.forecast_qty),
            }
            for row in forecasts[:5]
        ],
    }
    return sorted(forecasts, key=lambda row: row.sku), audit


def load_fabric_usage_by_spu() -> dict[str, list[FabricUsage]]:
    """复用面料核价读取器，保留每个 SPU-面料的显式单耗和损耗。"""
    usage = fabric_base.get_fabric_price_data()
    result: dict[str, list[FabricUsage]] = defaultdict(list)
    for (spu, fabric_name), values in usage.items():
        if fabric_name not in TARGET_FABRIC_ORDER:
            continue
        result[str(spu).strip()].append(
            FabricUsage(
                spu=str(spu).strip(),
                fabric_name=fabric_name,
                unit_usage=float(values.get("单件用量") or 0),
                loss_coefficient=float(values.get("单件损耗") or 1),
            )
        )
    for rows in result.values():
        rows.sort(key=lambda row: TARGET_FABRIC_ORDER[row.fabric_name])
    return dict(result)


def load_fabrics_by_spu() -> dict[str, list[str]]:
    """兼容旧调用方；新任务应使用 ``load_fabric_usage_by_spu``。"""
    return {
        spu: [row.fabric_name for row in rows]
        for spu, rows in load_fabric_usage_by_spu().items()
    }


def load_manual_mapping_catalog(
    path: Path | None = None,
) -> ManualMappingCatalog:
    """只读加载人工审核闭环 CSV；文件不存在等价于零条人工映射。"""
    path = path or DEFAULT_MANUAL_MAPPING_PATH
    if not path.exists():
        logger.info(f"人工确认映射不存在，按 0 条处理: {path}")
        return ManualMappingCatalog()

    required = {
        "面料名",
        "原始颜色编码",
        "原始颜色中文名",
        "颜色体系",
        "确认飞书记录ID",
        "确认飞书颜色原值",
        "审核人",
        "审核日期",
        "是否启用",
        "备注",
    }
    entries: list[ManualMapping] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        missing = sorted(required - set(reader.fieldnames or ()))
        if missing:
            raise RuntimeError(
                f"人工确认映射缺少字段: {', '.join(missing)}"
            )
        for row in reader:
            enabled_text = str(row.get("是否启用") or "").strip()
            enabled = not enabled_text or enabled_text.lower() in {
                "1",
                "true",
                "yes",
                "y",
                "是",
                "启用",
            }
            entry = ManualMapping(
                fabric_name=str(row.get("面料名") or "").strip(),
                color_code=str(row.get("原始颜色编码") or "").strip(),
                color_name=str(row.get("原始颜色中文名") or "").strip(),
                color_system=str(row.get("颜色体系") or "").strip(),
                record_id=str(row.get("确认飞书记录ID") or "").strip(),
                catalog_color_name=str(
                    row.get("确认飞书颜色原值") or ""
                ).strip(),
                reviewer=str(row.get("审核人") or "").strip(),
                reviewed_at=str(row.get("审核日期") or "").strip(),
                enabled=enabled,
                note=str(row.get("备注") or "").strip(),
            )
            if entry.fabric_name and (
                entry.record_id or entry.catalog_color_name
            ):
                entries.append(entry)
    return ManualMappingCatalog(entries)


def _texts(value: Any) -> list[str]:
    """提取飞书展示文本；不 strip/casefold，不把 record_id 当文本。"""
    if value is None:
        return []
    if isinstance(value, list):
        result: list[str] = []
        for item in value:
            result.extend(_texts(item))
        return result
    if isinstance(value, dict):
        for key in ("text", "name", "value"):
            if value.get(key) is not None and not isinstance(
                value.get(key), (dict, list)
            ):
                return [str(value.get(key))]
        return []
    return [str(value)]


def _link_record_ids(value: Any) -> list[str]:
    if isinstance(value, list):
        result: list[str] = []
        for item in value:
            result.extend(_link_record_ids(item))
        return result
    if isinstance(value, dict):
        for key in ("record_id", "id", "link_record_id"):
            raw = value.get(key)
            if raw:
                return [str(raw)]
    return []


def _linked_table_id(field_info: Mapping[str, Any]) -> str:
    prop = field_info.get("property") or {}
    candidates = (
        field_info.get("link_table"),
        field_info.get("link_table_id"),
        prop.get("table_id") if isinstance(prop, Mapping) else None,
        prop.get("tableId") if isinstance(prop, Mapping) else None,
    )
    return next((str(value) for value in candidates if value), "")


def parse_feishu_catalog_records(
    raw_records: Sequence[Mapping[str, Any]],
    linked_fabric_names: Mapping[str, str],
    fabric_field: str = DEFAULT_FABRIC_NAME_FIELD,
    color_field: str = DEFAULT_COLOR_FIELD,
    code_field: str = DEFAULT_LX_CODE_FIELD,
) -> tuple[list[CatalogRow], dict[str, Any]]:
    rows: list[CatalogRow] = []
    invalid_records: list[str] = []
    for record in raw_records:
        fields = record.get("fields") or {}
        record_id = str(record.get("record_id") or "")
        fabric_value = fields.get(fabric_field)
        fabric_names = _texts(fabric_value)
        if not fabric_names:
            fabric_names = [
                linked_fabric_names[record_id]
                for record_id in _link_record_ids(fabric_value)
                if record_id in linked_fabric_names
            ]
        color_names = _texts(fields.get(color_field))
        color_name = color_names[0] if color_names else ""
        code_values = _texts(fields.get(code_field))
        code = code_values[0] if code_values else ""

        if not fabric_names or not color_name:
            invalid_records.append(record_id)
            continue
        for fabric_name in dict.fromkeys(fabric_names):
            rows.append(
                CatalogRow(
                    fabric_name=fabric_name,
                    color_name=color_name,
                    lingxing_code=code,
                    record_ids=(record_id,) if record_id else (),
                )
            )
    return rows, {
        "source_record_count": len(raw_records),
        "parsed_fabric_association_count": len(rows),
        "invalid_record_count": len(invalid_records),
        "invalid_record_ids": invalid_records,
    }


async def load_catalog_from_feishu(
    base_token: str,
    table_id: str,
    view_id: str,
    fabric_field: str = DEFAULT_FABRIC_NAME_FIELD,
    color_field: str = DEFAULT_COLOR_FIELD,
    code_field: str = DEFAULT_LX_CODE_FIELD,
    linked_fabric_field: str = DEFAULT_LINKED_FABRIC_FIELD,
) -> tuple[list[CatalogRow], dict[str, Any]]:
    client = FeishuClient(
        app_token=base_token,
        table_id=table_id,
        view_id=view_id or None,
    )
    field_map = await client.get_table_fields()
    actual_fields = set(field_map.values())
    required = {fabric_field, color_field, code_field}
    missing = sorted(required - actual_fields)
    if missing:
        raise RuntimeError(f"飞书清单缺少字段: {', '.join(missing)}")

    raw_records = await client.read_records(page_size=500)
    fabric_info = await client.get_field_info(fabric_field)
    linked_table_id = (
        os.getenv("FABRIC_COLOR_LINKED_TABLE_ID")
        or _linked_table_id(fabric_info)
    )
    linked_names: dict[str, str] = {}
    unresolved_link_ids = {
        record_id
        for record in raw_records
        for record_id in _link_record_ids((record.get("fields") or {}).get(fabric_field))
    }
    if unresolved_link_ids:
        if not linked_table_id:
            raise RuntimeError(
                "面料名为关联字段但无法解析关联表；请配置 FABRIC_COLOR_LINKED_TABLE_ID"
            )
        linked_client = FeishuClient(
            app_token=base_token,
            table_id=linked_table_id,
        )
        linked_records = await linked_client.read_records(page_size=500)
        for record in linked_records:
            values = _texts((record.get("fields") or {}).get(linked_fabric_field))
            if values:
                linked_names[str(record.get("record_id") or "")] = values[0]

    rows, parse_audit = parse_feishu_catalog_records(
        raw_records,
        linked_names,
        fabric_field=fabric_field,
        color_field=color_field,
        code_field=code_field,
    )
    audit = {
        "base_token": base_token,
        "table_id": table_id,
        "view_id": view_id,
        "view_scope": "指定视图（视图过滤为空时等同全表）" if view_id else "全表",
        "actual_fields": sorted(actual_fields),
        "field_mapping": {
            "面料": fabric_field,
            "中文颜色": color_field,
            "SKU颜色编码": code_field,
            "关联面料展示字段": linked_fabric_field,
            "关联面料表": linked_table_id,
        },
        **parse_audit,
    }
    return rows, audit


def _style_header(ws: Any, headers: Sequence[str]) -> None:
    for column, header in enumerate(headers, 1):
        cell = ws.cell(1, column, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _autosize(ws: Any, minimum: int = 10, maximum: int = 50) -> None:
    for cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in cells)
        ws.column_dimensions[get_column_letter(cells[0].column)].width = min(
            max(length + 2, minimum), maximum
        )


def _write_rows(
    ws: Any,
    headers: Sequence[str],
    rows: Sequence[Mapping[str, Any]],
    pending: bool = False,
) -> None:
    _style_header(ws, headers)
    for row_index, row in enumerate(rows, 2):
        for column, header in enumerate(headers, 1):
            value = row.get(header, "")
            cell = ws.cell(row_index, column, value)
            cell.border = BORDER
            cell.alignment = RIGHT if isinstance(value, (int, float)) else LEFT
            if isinstance(value, (int, float)):
                cell.number_format = "#,##0.####"
            if pending:
                cell.fill = PENDING_FILL
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autosize(ws)


def _flatten_metrics(
    value: Any,
    prefix: str = "",
) -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = []
    if isinstance(value, Mapping):
        for key, child in value.items():
            child_prefix = f"{prefix}.{key}" if prefix else str(key)
            if isinstance(child, (Mapping, list, tuple)):
                rows.extend(_flatten_metrics(child, child_prefix))
            else:
                rows.append((child_prefix, str(child)))
    elif isinstance(value, (list, tuple)):
        for index, child in enumerate(value):
            child_prefix = f"{prefix}[{index}]"
            if isinstance(child, (Mapping, list, tuple)):
                rows.extend(_flatten_metrics(child, child_prefix))
            else:
                rows.append((child_prefix, str(child)))
    else:
        rows.append((prefix, str(value)))
    return rows


def export_workbook(
    result: StockingResult,
    output_dir: Path,
    generated_at: datetime | None = None,
) -> tuple[Path, Path]:
    generated_at = generated_at or datetime.now()
    output_dir.mkdir(parents=True, exist_ok=True)
    stem = f"面料-颜色备货对照_{generated_at.strftime('%Y%m%d_%H%M%S')}"
    workbook_path = output_dir / f"{stem}.xlsx"
    metrics_path = output_dir / f"{stem}_核对指标.json"

    workbook = Workbook()
    overview_ws = workbook.active
    overview_ws.title = "颜色用量总览"
    _write_rows(
        overview_ws,
        [
            "范围",
            "面料名",
            "飞书颜色数（去重后）",
            "可确定总用量（米）",
            "已分配到具体飞书颜色用量（米）",
            "待人工确认颜色用量（米）",
            "颜色分配覆盖率",
            "模糊候选涉及用量（米）",
            "模糊候选覆盖率",
            "潜在颜色覆盖率",
            "缺少单件用量的销量",
            "缺失原因",
        ],
        result.overview_rows,
    )

    color_ws = workbook.create_sheet("飞书颜色用量")
    _write_rows(
        color_ws,
        [
            "面料名",
            "飞书颜色原值",
            "飞书颜色中文核心",
            "领星新颜色缩写",
            "预估销量",
            "预估面料用量（米）",
            "关联SKU数",
            "关联SPU数",
            "匹配方式",
            "飞书记录ID",
            "原始重复行数",
        ],
        result.color_usage_rows,
    )

    auto_ws = workbook.create_sheet("自动归并SKU明细")
    _write_rows(
        auto_ws,
        [
            "SKU",
            "SPU",
            "品名",
            "面料名",
            "原颜色编码",
            "原颜色中文名",
            "颜色体系",
            "预估销量",
            "单件用量",
            "单件损耗",
            "预估面料用量（米）",
            "飞书颜色原值",
            "领星新颜色缩写",
            "匹配方式",
            "匹配依据",
            "飞书记录ID",
            "用量参数状态",
        ],
        result.auto_merge_rows,
    )

    manual_ws = workbook.create_sheet("待人工确认SKU")
    _write_rows(
        manual_ws,
        [
            "SKU",
            "SPU",
            "面料",
            "原颜色编码",
            "原颜色中文名",
            "颜色体系",
            "预估销量",
            "已算面料米数",
            "候选飞书颜色",
            "未自动归并原因",
        ],
        result.manual_review_rows,
        pending=True,
    )

    missing_ws = workbook.create_sheet("用量参数缺失")
    _write_rows(
        missing_ws,
        [
            "SPU",
            "SKU",
            "面料",
            "预估销量",
            "单件用量",
            "单件损耗",
            "缺失原因",
            "颜色匹配状态",
        ],
        result.usage_missing_rows,
        pending=True,
    )

    catalog_ws = workbook.create_sheet("飞书17面料颜色清单")
    _write_rows(
        catalog_ws,
        [
            "面料名",
            "飞书颜色原值",
            "飞书颜色中文核心",
            "括号内颜色别名",
            "领星新颜色缩写",
            "飞书记录ID",
            "原始重复行数",
            "当前是否有需求",
        ],
        result.catalog_rows,
    )

    fuzzy_ws = workbook.create_sheet("模糊候选审核")
    fuzzy_headers = [
        "面料名",
        "SKU",
        "SPU",
        "品名",
        "原始颜色编码",
        "原始颜色中文名",
        "颜色体系",
        "预估销量",
        "预估面料用量（米）",
    ]
    for position in range(1, 4):
        fuzzy_headers.extend(
            [
                f"候选{position}飞书颜色原值",
                f"候选{position}领星缩写",
                f"候选{position}分数",
                f"候选{position}依据",
                f"候选{position}飞书记录ID",
            ]
        )
    fuzzy_headers.extend(
        [
            "第一、第二候选分差",
            "置信度等级",
            "风险提示",
            "人工审核结果",
            "人工确认飞书记录ID",
            "审核备注",
        ]
    )
    _write_rows(
        fuzzy_ws,
        fuzzy_headers,
        result.fuzzy_review_rows,
        pending=True,
    )
    review_column = fuzzy_headers.index("人工审核结果") + 1
    review_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(FUZZY_REVIEW_RESULTS)}"',
        allow_blank=True,
    )
    review_validation.error = "请选择预设人工审核结果"
    review_validation.errorTitle = "无效审核结果"
    fuzzy_ws.add_data_validation(review_validation)
    if result.fuzzy_review_rows:
        review_validation.add(
            f"{get_column_letter(review_column)}2:"
            f"{get_column_letter(review_column)}{len(result.fuzzy_review_rows) + 1}"
        )

    priority_ws = workbook.create_sheet("优先补标清单")
    _write_rows(
        priority_ws,
        ["SKU", "品名", "颜色编码", "颜色中文名", "预估销量"],
        result.priority_rows,
        pending=True,
    )

    summary_ws = workbook.create_sheet("核对摘要")
    _style_header(summary_ws, ["指标", "值"])
    for row_index, (key, value) in enumerate(
        _flatten_metrics(result.metrics), 2
    ):
        summary_ws.cell(row_index, 1, key)
        summary_ws.cell(row_index, 2, value)
        for cell in summary_ws[row_index]:
            cell.border = BORDER
            cell.alignment = LEFT
    summary_ws.freeze_panes = "A2"
    _autosize(summary_ws, maximum=80)

    workbook.save(workbook_path)
    metrics_path.write_text(
        json.dumps(result.metrics, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    return workbook_path, metrics_path


async def run_read_only(
    output_dir: Path,
    as_of: date | None = None,
    base_token: str | None = None,
    table_id: str | None = None,
    view_id: str | None = None,
    manual_mapping_path: Path | None = None,
    fuzzy_config: FuzzyMatchConfig | None = None,
) -> tuple[StockingResult, Path, Path]:
    """执行只读 dry-run 并生成对照文件。"""
    logger.info("开始面料-颜色备货只读 dry-run（不写 MySQL/飞书）")
    governance_catalog = ColorMappingCatalog.from_runtime(strict=True)
    catalog_rows, catalog_source_audit = await load_catalog_from_feishu(
        base_token=base_token
        or os.getenv("FABRIC_COLOR_CATALOG_BASE_TOKEN", DEFAULT_BASE_TOKEN),
        table_id=table_id
        or os.getenv("FABRIC_COLOR_CATALOG_TABLE_ID", DEFAULT_CATALOG_TABLE_ID),
        view_id=(
            view_id
            if view_id is not None
            else os.getenv("FABRIC_COLOR_CATALOG_VIEW_ID", DEFAULT_CATALOG_VIEW_ID)
        ),
    )
    forecasts, forecast_audit = load_forecast_skus(
        governance_catalog, as_of=as_of
    )
    fabric_usage_by_spu = load_fabric_usage_by_spu()
    manual_catalog = load_manual_mapping_catalog(manual_mapping_path)
    result = aggregate_stocking(
        forecasts=forecasts,
        fabric_usage_by_spu=fabric_usage_by_spu,
        catalog_rows=catalog_rows,
        governance_catalog=governance_catalog,
        source_record_count=catalog_source_audit["source_record_count"],
        manual_catalog=manual_catalog,
        fuzzy_config=fuzzy_config,
        target_fabrics=TARGET_FABRICS,
    )
    result.metrics["forecast_audit"] = forecast_audit
    result.metrics["catalog_source_audit"] = catalog_source_audit
    result.metrics["fabric_mapping_spu_count"] = len(fabric_usage_by_spu)
    result.metrics["manual_mapping_path"] = str(
        manual_mapping_path or DEFAULT_MANUAL_MAPPING_PATH
    )
    result.metrics["source_adapter_disclosures"] = [
        "SKU标识复用 normalize_sku：NFKC、转大写、下划线转连字符、合并重复连字符",
        "预测表缺少颜色字段时复用 parse_lingxing_color：从品名提取“数字#中文名”并删除其中空白",
        "SPU→面料/单耗/损耗复用 get_fabric_price_data：仅 strip 首尾空白；单耗为空或0不使用平均值",
        "体系消歧复用 ColorMappingCatalog：颜色编码转大写并删除空白；匹配方式单列为“体系消歧”",
        "清单原值中文名/缩写匹配不做 strip、casefold、模糊匹配或别名归一",
        "清单颜色另显式解析“数字#中文名”“数字-中文名”；匹配方式独立标记为“清单色号解析”",
        "确定性中文核心/括号别名仅在前四级失败后使用；多候选不自动选择",
        "人工确认 CSV 对审核键执行显式首尾空白清理；命中优先级低于全部确定性清单规则、高于模糊候选",
        "模糊标准化仅生成同面料人工候选，处理NFKC符号、括号、末尾色、业务后缀、显式色号、括号别名、常见简繁与标点；永不计入已确认用量",
        "图案/复合颜色候选受关键词分组保护，不向纯色推荐",
    ]
    result.metrics["generated_at"] = datetime.now().isoformat(timespec="seconds")
    workbook_path, metrics_path = export_workbook(result, output_dir)
    logger.info(
        "dry-run 完成：可确定 %.2f 米，已确认 %.2f 米（%.2f%%），"
        "待人工确认 %.2f 米，模糊候选 %.2f 米，潜在覆盖 %.2f%%",
        result.metrics["calculable_total_meters"],
        result.metrics["confirmed_color_meters"],
        result.metrics["confirmed_color_coverage_pct"],
        result.metrics["manual_review_color_meters"],
        result.metrics["fuzzy_candidate_meters"],
        result.metrics["potential_color_coverage_pct"],
    )
    logger.info(f"Excel: {workbook_path}")
    logger.info(f"核对指标: {metrics_path}")
    return result, workbook_path, metrics_path


def _parse_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def main() -> None:
    parser = argparse.ArgumentParser(
        description="只读生成 SKU 未来4个月销量→17面料具体颜色用量（米）"
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path(
            os.getenv("PROCUREMENT_EXPORT_DIR", str(PROJECT_ROOT / "exports"))
        ),
    )
    parser.add_argument(
        "--as-of",
        type=_parse_date,
        help="预测窗口基准日 YYYY-MM-DD；默认今天",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        default=True,
        help="只读模式（当前唯一支持模式，不写数据库或飞书）",
    )
    parser.add_argument(
        "--manual-mapping",
        type=Path,
        default=DEFAULT_MANUAL_MAPPING_PATH,
        help="人工确认映射 CSV；默认 config/fabric_color_manual_mapping.csv",
    )
    parser.add_argument(
        "--fuzzy-high-score",
        type=float,
        default=90,
        help="高置信度最低分（仅人工候选）",
    )
    parser.add_argument(
        "--fuzzy-medium-score",
        type=float,
        default=80,
        help="中置信度最低分（仅人工候选）",
    )
    parser.add_argument(
        "--fuzzy-low-score",
        type=float,
        default=70,
        help="生成模糊候选的最低分",
    )
    parser.add_argument(
        "--fuzzy-lead-margin",
        type=float,
        default=10,
        help="高置信度第一、第二候选最低分差",
    )
    args = parser.parse_args()
    fuzzy_config = FuzzyMatchConfig(
        high_min_score=args.fuzzy_high_score,
        medium_min_score=args.fuzzy_medium_score,
        low_min_score=args.fuzzy_low_score,
        high_lead_margin=args.fuzzy_lead_margin,
    )
    fuzzy_config.validate()
    asyncio.run(
        run_read_only(
            output_dir=args.output_dir,
            as_of=args.as_of,
            manual_mapping_path=args.manual_mapping,
            fuzzy_config=fuzzy_config,
        )
    )


if __name__ == "__main__":
    main()
