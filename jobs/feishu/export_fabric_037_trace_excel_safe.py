#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""
安全版：导出 037超绒 面料使用情况与预估用量溯源 Excel。

修复点：
- 不使用 SQL UNION，避免不同表字段排序规则不同导致 Illegal mix of collations。
- 动态检查字段；旧表缺字段时用 0 或空值兜底。
- SKU 来源明细在 Python 内合并运营预计下单与系统预测。

运行：
  python -m jobs.feishu.export_fabric_037_trace_excel_safe
"""

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

project_root = Path(__file__).resolve().parent.parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

from common import get_logger
from common.database import db_cursor

logger = get_logger('fabric_037_trace_export_safe')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as e:
    raise SystemExit(
        "缺少 openpyxl 依赖，请先执行：\n"
        "  cd /opt/apps/pythondata\n"
        "  ./venv/bin/pip install openpyxl\n"
    ) from e

DEFAULT_KEYWORDS = ['037', '超绒']


def fetch_all(sql: str, params: Tuple[Any, ...] = ()) -> List[Dict[str, Any]]:
    with db_cursor(dictionary=True) as cursor:
        cursor.execute(sql, params)
        return list(cursor.fetchall())


def table_exists(table_name: str) -> bool:
    rows = fetch_all(
        """
        SELECT COUNT(*) AS cnt
        FROM information_schema.TABLES
        WHERE TABLE_SCHEMA = DATABASE()
          AND TABLE_NAME = %s
        """,
        (table_name,),
    )
    return bool(rows and rows[0].get('cnt'))


def get_columns(table_name: str) -> set:
    if not table_exists(table_name):
        return set()
    rows = fetch_all(
        """
        SELECT COLUMN_NAME
        FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA = DATABASE()
          AND TABLE_NAME = %s
        """,
        (table_name,),
    )
    return {r['COLUMN_NAME'] for r in rows}


def like_where(cols: Iterable[str], keywords: List[str]) -> Tuple[str, List[str]]:
    parts, params = [], []
    for col in cols:
        for kw in keywords:
            parts.append(f"`{col}` LIKE %s")
            params.append(f"%{kw}%")
    return '(' + ' OR '.join(parts) + ')', params


def col_select(cols: set, col: str, default_sql: str, alias: str = None) -> str:
    alias = alias or col
    if col in cols:
        return f"`{col}` AS `{alias}`"
    return f"{default_sql} AS `{alias}`"


def norm_date(v: Any) -> str:
    if v is None:
        return ''
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    return str(v)[:10]


def extract_spu(sku: str) -> str:
    if not sku:
        return ''
    sku = re.sub(r'\d+(?:PSC|PCS)', '', str(sku), flags=re.IGNORECASE)
    sku = re.sub(r'-+', '-', sku).strip('-')
    return sku.split('-', 1)[0] if '-' in sku else sku


def is_match(text: Any, keywords: List[str]) -> bool:
    s = str(text or '')
    return any(kw in s for kw in keywords)


def read_fabric_params(keywords: List[str]) -> List[Dict[str, Any]]:
    table = '定制面料参数'
    if not table_exists(table):
        return []
    cols = get_columns(table)
    where_cols = [c for c in ['面料', '面料编号'] if c in cols]
    if not where_cols:
        return []
    where, params = like_where(where_cols, keywords)
    selects = [
        col_select(cols, '面料', "''"),
        col_select(cols, '面料编号', "''"),
        col_select(cols, '米数每条', '0'),
        col_select(cols, '公斤数每条', '0'),
    ]
    return fetch_all(f"SELECT {', '.join(selects)} FROM `{table}` WHERE {where} ORDER BY 面料, 面料编号", tuple(params))


def read_forecast_result(keywords: List[str]) -> List[Dict[str, Any]]:
    table = '面料预估表'
    if not table_exists(table):
        return []
    cols = get_columns(table)
    where_cols = [c for c in ['面料', '面料编号', '面料颜色编号'] if c in cols]
    if not where_cols:
        return []
    where, params = like_where(where_cols, keywords)
    required = [
        ('统计类型', "''"), ('月份', "''"), ('面料', "''"), ('面料编号', "''"),
        ('颜色缩写', "''"), ('颜色', "''"), ('面料颜色编号', "''"), ('统计日期', 'NULL'),
        ('运营预计下单量', '0'), ('系统预估下单量', '0'), ('预计用量/米', '0'),
        ('系统预估用量/米', '0'), ('米数每条', '0'), ('预计用量/条', '0'),
        ('系统预估用量/条', '0'), ('库存量/条', '0'), ('库存量/米', '0'),
        ('待到货量/条', '0'), ('待到货量/米', '0'), ('预计总量/条', '0'),
        ('预计总量/米', '0'), ('用量信息缺失SPU', "''"), ('更新时间', 'NULL'),
    ]
    selects = [col_select(cols, c, d) for c, d in required]
    rows = fetch_all(
        f"SELECT {', '.join(selects)} FROM `{table}` WHERE {where} ORDER BY 统计日期, 面料, 颜色缩写",
        tuple(params),
    )
    for r in rows:
        r['统计日期'] = norm_date(r.get('统计日期'))
        r['更新时间'] = norm_date(r.get('更新时间'))
        r['系统预计采购条数'] = max(
            0,
            float(r.get('系统预估用量/条') or 0)
            - float(r.get('库存量/条') or 0)
            - float(r.get('待到货量/条') or 0),
        )
    return rows


def read_spu_fabric_usage(keywords: List[str]) -> Tuple[List[Dict[str, Any]], Dict[str, List[Dict[str, Any]]]]:
    table = '面料核价表'
    if not table_exists(table):
        return [], {}
    cols = get_columns(table)
    if 'SPU' not in cols or '面料' not in cols:
        return [], {}

    selects = [
        '`SPU` AS `SPU`',
        '`面料` AS `面料`',
        col_select(cols, '单件用量', '0'),
        col_select(cols, '单件损耗', '0'),
    ]
    all_rows = fetch_all(
        f"SELECT {', '.join(selects)} FROM `{table}` WHERE SPU IS NOT NULL AND SPU != '' AND 面料 IS NOT NULL AND 面料 != ''",
    )

    max_usage_by_spu: Dict[str, float] = {}
    by_spu: Dict[str, List[Dict[str, Any]]] = {}
    for r in all_rows:
        spu = str(r.get('SPU') or '').strip()
        usage = float(r.get('单件用量') or 0)
        by_spu.setdefault(spu, []).append(r)
        max_usage_by_spu[spu] = max(max_usage_by_spu.get(spu, 0), usage)

    matched = []
    for r in all_rows:
        if not is_match(r.get('面料'), keywords):
            continue
        spu = str(r.get('SPU') or '').strip()
        usage = float(r.get('单件用量') or 0)
        out = dict(r)
        out['主面料判定'] = '主面料' if usage == max_usage_by_spu.get(spu, 0) else '非主面料'
        matched.append(out)

    matched.sort(key=lambda x: (str(x.get('SPU') or ''), -float(x.get('单件用量') or 0)))
    return matched, by_spu


def read_op_forecast() -> Dict[Tuple[str, str], int]:
    table = '运营预计下单表'
    if not table_exists(table):
        return {}
    cols = get_columns(table)
    if not {'SKU', '统计日期', '预计下单量'} <= cols:
        return {}
    rows = fetch_all(
        """
        SELECT SKU, 统计日期, SUM(预计下单量) AS qty
        FROM `运营预计下单表`
        WHERE SKU IS NOT NULL AND SKU != '' AND 统计日期 IS NOT NULL AND 预计下单量 > 0
        GROUP BY SKU, 统计日期
        """
    )
    return {(str(r['SKU']).strip(), norm_date(r['统计日期'])): int(r.get('qty') or 0) for r in rows}


def read_sys_forecast() -> Dict[Tuple[str, str], int]:
    table = '预测对比表_SKU'
    if not table_exists(table):
        return {}
    cols = get_columns(table)
    if not {'SKU', '统计日期', '系统预测销量'} <= cols:
        return {}
    rows = fetch_all(
        """
        SELECT SKU, 统计日期, SUM(系统预测销量) AS qty
        FROM `预测对比表_SKU`
        WHERE SKU IS NOT NULL AND SKU != '' AND 统计日期 IS NOT NULL AND 系统预测销量 > 0
        GROUP BY SKU, 统计日期
        """
    )
    return {(str(r['SKU']).strip(), norm_date(r['统计日期'])): int(r.get('qty') or 0) for r in rows}


def read_sku_trace(keywords: List[str], fabric_by_spu: Dict[str, List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
    op_map = read_op_forecast()
    sys_map = read_sys_forecast()
    keys = sorted(set(op_map.keys()) | set(sys_map.keys()), key=lambda x: (x[1], x[0]))
    rows: List[Dict[str, Any]] = []

    for sku, stat_date in keys:
        spu = extract_spu(sku)
        fabric_rows = fabric_by_spu.get(spu, [])
        for f in fabric_rows:
            if not is_match(f.get('面料'), keywords):
                continue
            unit_usage = float(f.get('单件用量') or 0)
            loss = float(f.get('单件损耗') or 0)
            loss_factor = loss if loss else 1
            op_qty = op_map.get((sku, stat_date), 0)
            sys_qty = sys_map.get((sku, stat_date), 0)
            rows.append({
                'SKU': sku,
                'SPU': spu,
                '统计日期': stat_date,
                '面料': f.get('面料') or '',
                '单件用量': unit_usage,
                '单件损耗': loss,
                '主面料判定': f.get('主面料判定', ''),
                '运营预计下单量': op_qty,
                '系统预估下单量': sys_qty,
                '运营预计用量_米': round(op_qty * unit_usage * loss_factor, 2),
                '系统预估用量_米': round(sys_qty * unit_usage * loss_factor, 2),
            })
    return rows


def read_inventory(keywords: List[str]) -> List[Dict[str, Any]]:
    table = '面料库存台账'
    if not table_exists(table):
        return []
    cols = get_columns(table)
    if '面料编号颜色缩写' not in cols:
        return []
    where, params = like_where(['面料编号颜色缩写'], keywords)
    selects = [
        col_select(cols, '面料编号颜色缩写', "''"),
        col_select(cols, '库存成品数量_条', '0'),
        col_select(cols, '备货中数量_条', '0'),
        col_select(cols, '现有胚布数量_条', '0'),
        col_select(cols, '面料', "''"),
        col_select(cols, '颜色', "''"),
        col_select(cols, '更新时间', 'NULL'),
    ]
    rows = fetch_all(f"SELECT {', '.join(selects)} FROM `{table}` WHERE {where} ORDER BY 面料编号颜色缩写", tuple(params))
    for r in rows:
        r['更新时间'] = norm_date(r.get('更新时间'))
    return rows


def read_color_merge(keywords: List[str]) -> List[Dict[str, Any]]:
    table = '面料颜色归并对照'
    if not table_exists(table):
        return []
    cols = get_columns(table)
    if '面料编号' not in cols:
        return []
    where, params = like_where(['面料编号'], keywords)
    selects = [
        col_select(cols, '面料编号', "''"),
        col_select(cols, '原始颜色缩写', "''"),
        col_select(cols, '归并颜色缩写', "''"),
        col_select(cols, '是否启用', '0'),
    ]
    return fetch_all(f"SELECT {', '.join(selects)} FROM `{table}` WHERE {where} ORDER BY 面料编号, 原始颜色缩写", tuple(params))


def build_checks(params_rows, result_rows, spu_rows, sku_rows, inventory_rows, color_rows):
    checks = []

    def add(name, status, detail):
        checks.append({'检查项': name, '状态': status, '说明': detail})

    add('定制面料参数', '正常' if params_rows else '异常', f'命中 {len(params_rows)} 条')
    add('面料核价表/SPU用料关系', '正常' if spu_rows else '异常', f'命中 {len(spu_rows)} 条')
    add('SKU用量溯源', '正常' if sku_rows else '异常', f'命中 {len(sku_rows)} 条')
    add('面料预估表最终结果', '正常' if result_rows else '异常', f'命中 {len(result_rows)} 条')
    add('面料库存台账', '正常' if inventory_rows else '提醒', f'命中 {len(inventory_rows)} 条')
    add('颜色归并', '正常' if color_rows else '提醒', f'命中 {len(color_rows)} 条，没有不一定异常')
    add('单件用量为0', '异常' if any(float(r.get('单件用量') or 0) <= 0 for r in spu_rows) else '正常', '')
    add('米数每条为0', '异常' if any(float(r.get('米数每条') or 0) <= 0 for r in params_rows) else '正常', '')
    add('用量信息缺失SPU', '异常' if any(r.get('用量信息缺失SPU') for r in result_rows) else '正常', '')
    return checks


def safe_sheet_name(name: str) -> str:
    for ch in r'[]:*?/\\':
        name = name.replace(ch, '_')
    return name[:31]


def write_sheet(wb: Workbook, title: str, rows: List[Dict[str, Any]]):
    ws = wb.create_sheet(safe_sheet_name(title))
    if not rows:
        ws.append(['提示'])
        ws.append(['无数据'])
        ws.freeze_panes = 'A2'
        return

    headers = list(rows[0].keys())
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h) for h in headers])

    fill = PatternFill('solid', fgColor='1F4E78')
    font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='D9E2F3')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in ws[1]:
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical='center')

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    if ws.max_row >= 2:
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        display = re.sub(r'[^A-Za-z0-9_]', '_', f'tbl_{title}')[:250]
        if not re.match(r'^[A-Za-z_]', display):
            display = 'T_' + display
        table = Table(displayName=display, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True, showColumnStripes=False)
        try:
            ws.add_table(table)
        except Exception:
            pass
    for idx, col in enumerate(ws.columns, start=1):
        width = max(10, min(42, max(len(str(c.value or '')) for c in col) + 2))
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_summary(wb: Workbook, keywords, counts, checks):
    ws = wb.active
    ws.title = '汇总说明'
    ws.append(['037超绒面料使用情况与预估用量溯源'])
    ws.append(['生成时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    ws.append(['关键词', ', '.join(keywords)])
    ws.append(['说明', '安全版脚本：不使用SQL UNION，兼容旧表结构。'])
    ws.append([])
    ws.append(['数据量汇总'])
    for k, v in counts.items():
        ws.append([k, v])
    ws.append([])
    ws.append(['检查项', '状态', '说明'])
    for r in checks:
        ws.append([r['检查项'], r['状态'], r['说明']])
    ws.merge_cells('A1:C1')
    ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='1F4E78')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 90
    ws.freeze_panes = 'A2'


def export_excel(output_path: str, keywords: List[str]):
    logger.info('开始读取037超绒相关数据（安全版）...')
    params_rows = read_fabric_params(keywords)
    result_rows = read_forecast_result(keywords)
    spu_rows, fabric_by_spu = read_spu_fabric_usage(keywords)
    sku_rows = read_sku_trace(keywords, fabric_by_spu)
    inventory_rows = read_inventory(keywords)
    color_rows = read_color_merge(keywords)
    checks = build_checks(params_rows, result_rows, spu_rows, sku_rows, inventory_rows, color_rows)
    counts = {
        '面料预估结果': len(result_rows),
        'SPU用料关系': len(spu_rows),
        'SKU用量溯源': len(sku_rows),
        '库存台账': len(inventory_rows),
        '颜色归并': len(color_rows),
        '定制面料参数': len(params_rows),
    }

    wb = Workbook()
    write_summary(wb, keywords, counts, checks)
    write_sheet(wb, '01_面料预估结果', result_rows)
    write_sheet(wb, '02_SPU用料关系', spu_rows)
    write_sheet(wb, '03_SKU用量溯源', sku_rows)
    write_sheet(wb, '04_库存台账', inventory_rows)
    write_sheet(wb, '05_颜色归并', color_rows)
    write_sheet(wb, '06_定制面料参数', params_rows)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    logger.info(f'✓ Excel 已生成: {out}')
    logger.info('数据量汇总: ' + ', '.join(f'{k}={v}' for k, v in counts.items()))


def parse_args():
    p = argparse.ArgumentParser(description='安全导出037超绒面料溯源Excel')
    p.add_argument('--fabric-keyword', action='append', default=None, help='匹配关键词，可重复。默认：037、超绒')
    p.add_argument('--output', default=None, help='输出Excel路径')
    return p.parse_args()


def main():
    args = parse_args()
    keywords = args.fabric_keyword or DEFAULT_KEYWORDS
    output = args.output or f"exports/037超绒面料溯源_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    export_excel(output, keywords)


if __name__ == '__main__':
    main()
