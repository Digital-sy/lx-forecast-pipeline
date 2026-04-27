#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""
生成领星费用单导入Excel文件
从数据库读取本月（或指定月份）累计利润报表数据，
生成符合领星导入模板格式的Excel文件，每文件最多4800行。

使用方式：
  python generate_fee_excel.py                      # 生成本月累计数据
  python generate_fee_excel.py --month 2026-03      # 生成指定月份数据
  python generate_fee_excel.py --output /tmp/fees   # 指定输出目录
"""

import os
import sys
import argparse
from datetime import datetime, date
from typing import List, Dict, Any, Tuple
from calendar import monthrange

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from common import get_logger
from common.database import db_cursor

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl --break-system-packages")
    sys.exit(1)

logger = get_logger('generate_fee_excel')

# ── 常量 ────────────────────────────────────────────────────────────────────
ROWS_PER_FILE   = 4800          # 每个文件最大行数（领星建议上限）
HEADER_ROW = [                  # 与领星官方模板完全一致
    '*店铺', '*MSKU', '*币种', '*金额', '*费用类型',
    '*费用分摊日期', '*入账方式', '备注', '*应用报表类型'
]
FIXED_CURRENCY      = 'CNY'
FIXED_ACCOUNT_TYPE  = '仅分摊'
FIXED_REPORT_TYPE   = '全选'

# 无效MSKU（领星商品库中不存在，跳过）
INVALID_MSKUS = {'Amazon.Found.B0CJCDHZYN', 'Amazon.Found.B0CJCFLXJW', 'UNKNOW'}

# 费用字段 → 领星费用类型名称
FEE_FIELD_MAP = {
    '商品成本附加费': '商品成本附加费',
    '头程成本附加费': '头程成本附加费',
    '录入费用单头程': '头程费用',
    '汇损':          '汇损',
}


# ── 数据库查询 ────────────────────────────────────────────────────────────────
def fetch_fee_data(start_date: str, end_date: str) -> List[Dict[str, Any]]:
    """
    读取利润报表按月汇总数据（只取4个费用字段非零的记录）

    Args:
        start_date: 月份第一天，格式 Y-m-d
        end_date:   月份最后一天，格式 Y-m-d

    Returns:
        List[Dict]: 每条记录包含 店铺、MSKU、各费用金额
    """
    logger.info(f"📊 从数据库读取费用数据（{start_date} ~ {end_date}）...")

    with db_cursor() as cursor:
        sql = """
            SELECT
                `店铺`,
                `MSKU`,
                SUM(`商品成本附加费`) AS `商品成本附加费`,
                SUM(`头程成本附加费`) AS `头程成本附加费`,
                SUM(`录入费用单头程`) AS `录入费用单头程`,
                SUM(`汇损`)          AS `汇损`
            FROM `利润报表`
            WHERE `统计日期` >= %s
              AND `统计日期` <= %s
              AND (
                  (`商品成本附加费` IS NOT NULL AND `商品成本附加费` != 0) OR
                  (`头程成本附加费` IS NOT NULL AND `头程成本附加费` != 0) OR
                  (`录入费用单头程` IS NOT NULL AND `录入费用单头程` != 0) OR
                  (`汇损`          IS NOT NULL AND `汇损`          != 0)
              )
            GROUP BY `店铺`, `MSKU`
            ORDER BY `店铺`, `MSKU`
        """
        cursor.execute(sql, (start_date, end_date))
        records = cursor.fetchall()

    logger.info(f"✅ 查询到 {len(records)} 条 (店铺, MSKU) 组合")
    return records


# ── 展开为Excel行 ─────────────────────────────────────────────────────────────
def expand_to_rows(
    records: List[Dict[str, Any]],
    fee_month: str          # 格式 YYYY-MM，写入「费用分摊日期」列
) -> List[List[Any]]:
    """
    将每条数据库记录展开为最多4行（每个非零费用类型一行）
    同时过滤无效MSKU
    """
    rows = []
    skipped = 0

    for rec in records:
        msku  = (rec.get('MSKU') or '').strip()
        shop  = (rec.get('店铺') or '').strip()

        if not msku or not shop:
            skipped += 1
            continue

        if msku in INVALID_MSKUS:
            logger.debug(f"  跳过无效MSKU: {msku}")
            skipped += 1
            continue

        for db_field, fee_type_name in FEE_FIELD_MAP.items():
            amount = float(rec.get(db_field) or 0)
            if amount == 0:
                continue

            rows.append([
                shop,
                msku,
                FIXED_CURRENCY,
                round(amount, 4),
                fee_type_name,
                fee_month,
                FIXED_ACCOUNT_TYPE,
                '',            # 备注（留空）
                FIXED_REPORT_TYPE,
            ])

    logger.info(f"  展开后共 {len(rows)} 行，跳过 {skipped} 条无效记录")
    return rows


# ── 写入单个Excel文件 ─────────────────────────────────────────────────────────
def write_excel(rows: List[List[Any]], filepath: str) -> None:
    """将数据行写入符合领星模板格式的Excel文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'MSKU'

    # ── 表头样式 ──
    header_fill = PatternFill('solid', start_color='BDD7EE')
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(HEADER_ROW)
    for col_idx, _ in enumerate(HEADER_ROW, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_alignment
        cell.border    = border

    # ── 数据行 ──
    for row in rows:
        ws.append(row)

    # ── 列宽 ──
    col_widths = [15, 40, 8, 16, 18, 16, 12, 20, 16]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── hiddenSelect1 / hiddenSelect5（领星模板需要这两个隐藏Sheet）──
    ws2 = wb.create_sheet('hiddenSelect1')
    for shop in ['CY-US','MT-US','MT-CA','SY-US','JQ-US','JQ-CA','RKZ-US',
                 'RR-UK','RR-IT','RR-DE','RR-FR','RR-ES','RR-PL','RR-SE']:
        ws2.append([shop])

    ws3 = wb.create_sheet('hiddenSelect5')
    for ft in ['头程费用','汇损','其他费用分摊','头程成本附加费','商品成本附加费',
               '多平台成本','销毁费用','广告费用减免']:
        ws3.append([ft])

    wb.save(filepath)
    logger.info(f"  💾 已保存: {filepath}  ({len(rows)} 行数据)")


# ── 主逻辑 ────────────────────────────────────────────────────────────────────
def generate(month: str, output_dir: str) -> List[str]:
    """
    生成指定月份的费用单Excel文件

    Args:
        month:      格式 YYYY-MM
        output_dir: 输出目录

    Returns:
        List[str]: 生成的文件路径列表
    """
    # 解析月份，得到日期范围
    try:
        year, mon = map(int, month.split('-'))
    except ValueError:
        logger.error(f"❌ 月份格式错误，应为 YYYY-MM，实际：{month}")
        return []

    last_day = monthrange(year, mon)[1]
    start_date = f"{year:04d}-{mon:02d}-01"
    end_date   = f"{year:04d}-{mon:02d}-{last_day:02d}"
    fee_month  = f"{year:04d}-{mon:02d}"   # 写入Excel的「费用分摊日期」

    logger.info("=" * 70)
    logger.info(f"🗂️   生成费用单Excel  |  月份：{fee_month}")
    logger.info(f"    数据范围：{start_date} ~ {end_date}")
    logger.info(f"    输出目录：{output_dir}")
    logger.info("=" * 70)

    os.makedirs(output_dir, exist_ok=True)

    # 1. 从数据库读取
    records = fetch_fee_data(start_date, end_date)
    if not records:
        logger.warning("⚠️  没有需要写入的费用数据")
        return []

    # 2. 展开为Excel行
    all_rows = expand_to_rows(records, fee_month)
    if not all_rows:
        logger.warning("⚠️  展开后没有有效行")
        return []

    # 3. 按4800行分批写文件
    total_files = (len(all_rows) + ROWS_PER_FILE - 1) // ROWS_PER_FILE
    logger.info(f"📁 共 {len(all_rows)} 行，将生成 {total_files} 个文件")

    generated_files = []
    for i in range(total_files):
        chunk = all_rows[i * ROWS_PER_FILE : (i + 1) * ROWS_PER_FILE]
        filename = f"fee_{fee_month}_{i+1:02d}of{total_files:02d}.xlsx"
        filepath = os.path.join(output_dir, filename)
        write_excel(chunk, filepath)
        generated_files.append(filepath)

    logger.info("")
    logger.info(f"✅ 生成完毕，共 {len(generated_files)} 个文件：")
    for f in generated_files:
        logger.info(f"   {f}")

    return generated_files


# ── 入口 ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='生成领星费用单导入Excel')
    parser.add_argument(
        '--month', type=str, default=None,
        help='目标月份，格式：YYYY-MM，默认：本月'
    )
    parser.add_argument(
        '--output', type=str,
        default=os.path.join(os.path.dirname(os.path.abspath(__file__)), 'fee_excel_output'),
        help='输出目录，默认：./fee_excel_output'
    )
    args = parser.parse_args()

    target_month = args.month or datetime.now().strftime('%Y-%m')

    try:
        files = generate(month=target_month, output_dir=args.output)
        if not files:
            sys.exit(1)
    except KeyboardInterrupt:
        logger.warning("⚠️  用户中断")
        sys.exit(1)
    except Exception as e:
        logger.error(f"❌ 执行失败: {e}", exc_info=True)
        sys.exit(1)
