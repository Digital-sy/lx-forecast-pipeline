#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""
RKZ-US FBA库存治理 MVP
======================

用途：
- 仅查询 RKZ-US（sid=11550，店号=7店）
- 以 ods_db.ods_lx_fba_warehouse_detail 为库存主表
- 拼接近30天销量、产品管理、店铺、员工组织信息
- 计算 SKU / SPU 可售天数
- 颜色/尺码复用项目现有 SKU 解析逻辑
- 暂不计算未来仓储费、超龄库存附加费、弃置费等无法可靠生成的字段，保持空值
- 导出为 Excel，便于MVP核对

运行方式：
    python tmp/rkz_fba_inventory_mvp.py

可选：
    python tmp/rkz_fba_inventory_mvp.py --output /tmp/rkz_fba_inventory_mvp.xlsx
    python tmp/rkz_fba_inventory_mvp.py --sample 20

说明：
1. 当前FBA总库存口径按业务确认：
   FBA可售 + FBA预留 + 待调仓 + 标发在途 + 入库中
2. 当前FBA可用库存口径按业务确认：
   FBA可售 + FBA预留 + 待调仓 + 入库中
3. 目前尚未验证 afn_reserved_quantity 是否已经包含 reserved_fc_transfers；
   MVP阶段保留业务确认公式，正式落表前需与领星页面抽样核对。
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# 允许从仓库根目录直接运行：python tmp/rkz_fba_inventory_mvp.py
PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from common import get_logger  # noqa: E402
from common.database import db_cursor  # noqa: E402
from jobs.feishu.color_system_resolver import (  # noqa: E402
    extract_raw_color_code,
    normalize_sku,
)

logger = get_logger("rkz_fba_inventory_mvp")

RKZ_SID = 11550
RKZ_STORE_NAME = "RKZ-US"
RKZ_STORE_NO = "7店"

FBA_TABLE = "ods_db.ods_lx_fba_warehouse_detail"
PERF_TABLE = "ods_db.ods_lx_product_performance_asin"
PRODUCT_TABLE = "ods_db.ods_lx_product_management"
STORE_TABLE = "ods_db.ods_lx_store_lists"
EMPLOYEE_TABLE = "ods_db.ods_fs_employee_info"

OUTPUT_COLUMNS = [
    "ASIN",
    "FNSKU",
    "MSKU",
    "SKU",
    "FBA总库存",
    "FBA可用库存",
    "30天内库龄",
    "31-60天库龄",
    "61-90天库龄",
    "91-180天库龄",
    "181-270天库龄",
    "270天以上库龄",
    "款号",
    "店铺",
    "站点",
    "店号",
    "部门",
    "组别",
    "运营",
    "季节",
    "颜色",
    "尺码",
    "30天销售",
    "日均销量",
    "该SKU的FBA总库存可售天数",
    "该SKU的FBA可用库存SKU可售天数",
    "该款号的FBA总库存款号可售天数",
    "该款号的FBA可用库存款号可售天数",
    "负责人",
    "未来6个月月度仓储费",
    "未来6个月超龄库存附加费",
    "未来6个月仓储费合计",
    "一次性弃置费用",
    "延迟6个月处理总成本",
    "延迟处理新增成本",
    "币种",
    "存放2个月月度仓储费",
    "存放2个月超龄库存附加费",
    "存放2个月仓储费合计",
    "品类",
]

REQUIRED_COLUMNS = {
    "ods_lx_fba_warehouse_detail": {
        "sid",
        "asin",
        "fnsku",
        "seller_sku",
        "sku",
        "afn_fulfillable_quantity",
        "afn_reserved_quantity",
        "reserved_fc_transfers",
        "afn_inbound_shipped_quantity",
        "afn_inbound_receiving_quantity",
        "inv_age_0_to_30_days",
        "inv_age_31_to_60_days",
        "inv_age_61_to_90_days",
        "inv_age_91_to_180_days",
        "inv_age_181_to_270_days",
        "inv_age_271_to_330_days",
        "inv_age_331_to_365_days",
        "inv_age_365_plus_days",
    },
    "ods_lx_product_performance_asin": {
        "dt",
        "sid",
        "store_name",
        "asin",
        "msku",
        "sku",
        "country",
        "principal_names",
        "currency_code",
        "volume",
        "delete_flag",
    },
    "ods_lx_product_management": {
        "product_id",
        "sku",
        "spu",
        "season",
        "develop_year",
        "product_category",
        "update_time",
        "etl_load_time",
    },
    "ods_lx_store_lists": {
        "sid",
        "store_name",
        "region",
        "delete_flag",
    },
    "ods_fs_employee_info": {
        "record_id",
        "employee",
        "dept",
        "is_resigned",
        "delete_flag",
        "modify_time",
    },
}


def nz(value: Any) -> float:
    """None/空字符串安全转数字。"""
    if value in (None, "", "None"):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def first_owner_name(value: Any) -> str:
    """principal_names 若意外包含多人，只取第一个用于组织匹配。"""
    s = text(value)
    if not s:
        return ""
    parts = re.split(r"[,，/、;；]+", s)
    return parts[0].strip() if parts else s


def parse_size_from_sku(sku: Any) -> str:
    """尺码取标准化 SKU 最后一段。"""
    normalized = normalize_sku(sku)
    if not normalized or "-" not in normalized:
        return ""
    return normalized.split("-")[-1]


def combine_season(develop_year: Any, season: Any) -> Optional[str]:
    """
    沿用现有季节组合方向：
    - develop_year=历史 -> 历史-原season
    - develop_year有值但season为空 -> NULL
    - 其他 -> YY-season
    - develop_year为空 -> 原season
    """
    year = text(develop_year)
    season_text = text(season)

    if year == "历史":
        return f"历史-{season_text}" if season_text else "历史"

    if year and not season_text:
        return None

    if not year:
        return season_text or None

    if year.isdigit() and len(year) >= 4:
        year_short = year[-2:]
    else:
        year_short = year[-2:] if len(year) > 2 else year

    if season_text.startswith(f"{year_short}-"):
        return season_text

    return f"{year_short}-{season_text}" if season_text else None


def days_of_supply(inventory: float, sales_30d: float) -> str:
    """0销量显示 ∞；否则库存 / (30天销量/30)。"""
    if sales_30d <= 0:
        return "∞"
    return f"{inventory / (sales_30d / 30.0):.1f}"


def split_schema_table(full_name: str) -> Tuple[str, str]:
    schema, table = full_name.split(".", 1)
    return schema, table


def validate_schema() -> None:
    """运行前检查关键字段，避免字段变更后静默跑错。"""
    problems: List[str] = []

    with db_cursor() as cursor:
        for table_name, required in REQUIRED_COLUMNS.items():
            cursor.execute(
                """
                SELECT COLUMN_NAME
                FROM information_schema.columns
                WHERE TABLE_SCHEMA = 'ods_db'
                  AND TABLE_NAME = %s
                """,
                (table_name,),
            )
            actual = {row["COLUMN_NAME"] for row in cursor.fetchall()}
            missing = sorted(required - actual)
            if missing:
                problems.append(f"{table_name} 缺字段: {', '.join(missing)}")

    if problems:
        raise RuntimeError("数据库字段自检失败:\n- " + "\n- ".join(problems))

    logger.info("数据库字段自检通过")


def fetch_fba_rows() -> List[Dict[str, Any]]:
    sql = f"""
        SELECT
            sid,
            asin,
            fnsku,
            seller_sku,
            sku,
            afn_fulfillable_quantity,
            afn_reserved_quantity,
            reserved_fc_transfers,
            afn_inbound_shipped_quantity,
            afn_inbound_receiving_quantity,
            inv_age_0_to_30_days,
            inv_age_31_to_60_days,
            inv_age_61_to_90_days,
            inv_age_91_to_180_days,
            inv_age_181_to_270_days,
            inv_age_271_to_330_days,
            inv_age_331_to_365_days,
            inv_age_365_plus_days
        FROM {FBA_TABLE}
        WHERE sid = %s
    """
    with db_cursor() as cursor:
        cursor.execute(sql, (RKZ_SID,))
        rows = list(cursor.fetchall())
    logger.info(f"FBA库存读取完成: {len(rows)} 行")
    return rows


def fetch_perf_30d() -> Dict[Tuple[str, str], Dict[str, Any]]:
    """按 asin+msku 汇总最新日期往前30天销量。"""
    with db_cursor() as cursor:
        cursor.execute(
            f"""
            SELECT MAX(dt) AS max_dt
            FROM {PERF_TABLE}
            WHERE delete_flag = 0
              AND sid = %s
            """,
            (str(RKZ_SID),),
        )
        max_dt_row = cursor.fetchone()
        max_dt = max_dt_row.get("max_dt") if max_dt_row else None

        if not max_dt:
            logger.warning("产品表现表未找到 RKZ-US 数据")
            return {}

        cursor.execute(
            f"""
            SELECT
                asin,
                msku,
                MAX(sku) AS sku,
                MAX(store_name) AS store_name,
                MAX(country) AS country,
                MAX(NULLIF(TRIM(principal_names), '')) AS principal_names,
                MAX(NULLIF(TRIM(currency_code), '')) AS currency_code,
                SUM(COALESCE(volume, 0)) AS sales_30d
            FROM {PERF_TABLE}
            WHERE delete_flag = 0
              AND sid = %s
              AND dt BETWEEN DATE_SUB(%s, INTERVAL 29 DAY) AND %s
            GROUP BY asin, msku
            """,
            (str(RKZ_SID), max_dt, max_dt),
        )
        rows = list(cursor.fetchall())

    result: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for row in rows:
        key = (text(row.get("asin")), text(row.get("msku")))
        result[key] = row

    logger.info(f"近30天销量汇总完成: 截止 {max_dt}, {len(result)} 个 asin+msku")
    return result


def fetch_product_map() -> Dict[str, Dict[str, Any]]:
    """每个SKU保留产品管理表更新时间最新的一条。"""
    sql = f"""
        WITH ranked AS (
            SELECT
                product_id,
                sku,
                spu,
                season,
                develop_year,
                product_category,
                update_time,
                etl_load_time,
                ROW_NUMBER() OVER (
                    PARTITION BY sku
                    ORDER BY update_time DESC, etl_load_time DESC, product_id DESC
                ) AS rn
            FROM {PRODUCT_TABLE}
            WHERE sku IS NOT NULL
              AND TRIM(sku) <> ''
        )
        SELECT *
        FROM ranked
        WHERE rn = 1
    """
    with db_cursor() as cursor:
        cursor.execute(sql)
        rows = list(cursor.fetchall())

    result = {text(row.get("sku")): row for row in rows if text(row.get("sku"))}
    logger.info(f"产品管理映射完成: {len(result)} 个SKU")
    return result


def fetch_store_info() -> Dict[str, Any]:
    with db_cursor() as cursor:
        cursor.execute(
            f"""
            SELECT sid, store_name, region
            FROM {STORE_TABLE}
            WHERE sid = %s
              AND COALESCE(delete_flag, 0) = 0
            ORDER BY id DESC
            LIMIT 1
            """,
            (RKZ_SID,),
        )
        row = cursor.fetchone() or {}
    return row


def fetch_employee_map() -> Dict[str, Dict[str, Any]]:
    sql = f"""
        WITH ranked AS (
            SELECT
                record_id,
                employee,
                dept,
                modify_time,
                ROW_NUMBER() OVER (
                    PARTITION BY TRIM(employee)
                    ORDER BY modify_time DESC, record_id DESC
                ) AS rn
            FROM {EMPLOYEE_TABLE}
            WHERE COALESCE(delete_flag, 0) = 0
              AND COALESCE(is_resigned, 0) = 0
              AND employee IS NOT NULL
              AND TRIM(employee) <> ''
        )
        SELECT employee, dept
        FROM ranked
        WHERE rn = 1
    """
    with db_cursor() as cursor:
        cursor.execute(sql)
        rows = list(cursor.fetchall())

    return {
        text(row.get("employee")): row
        for row in rows
        if text(row.get("employee"))
    }


def build_rows() -> List[Dict[str, Any]]:
    fba_rows = fetch_fba_rows()
    perf_map = fetch_perf_30d()
    product_map = fetch_product_map()
    store_info = fetch_store_info()
    employee_map = fetch_employee_map()

    built: List[Dict[str, Any]] = []

    for fba in fba_rows:
        asin = text(fba.get("asin"))
        fnsku = text(fba.get("fnsku"))
        msku = text(fba.get("seller_sku"))
        sku = text(fba.get("sku"))

        perf = perf_map.get((asin, msku), {})
        pm = product_map.get(sku, {})

        operator_name = text(perf.get("principal_names"))
        employee_name = first_owner_name(operator_name)
        employee = employee_map.get(employee_name, {})
        dept = text(employee.get("dept"))

        fba_total_inventory = (
            nz(fba.get("afn_fulfillable_quantity"))
            + nz(fba.get("afn_reserved_quantity"))
            + nz(fba.get("reserved_fc_transfers"))
            + nz(fba.get("afn_inbound_shipped_quantity"))
            + nz(fba.get("afn_inbound_receiving_quantity"))
        )

        fba_available_inventory = (
            nz(fba.get("afn_fulfillable_quantity"))
            + nz(fba.get("afn_reserved_quantity"))
            + nz(fba.get("reserved_fc_transfers"))
            + nz(fba.get("afn_inbound_receiving_quantity"))
        )

        age_270_plus = (
            nz(fba.get("inv_age_271_to_330_days"))
            + nz(fba.get("inv_age_331_to_365_days"))
            + nz(fba.get("inv_age_365_plus_days"))
        )

        spu = text(pm.get("spu"))
        if not spu:
            normalized_sku = normalize_sku(sku)
            spu = normalized_sku.split("-", 1)[0] if normalized_sku else ""

        sales_30d = nz(perf.get("sales_30d"))

        row = {
            "ASIN": asin,
            "FNSKU": fnsku,
            "MSKU": msku,
            "SKU": sku,
            "FBA总库存": fba_total_inventory,
            "FBA可用库存": fba_available_inventory,
            "30天内库龄": nz(fba.get("inv_age_0_to_30_days")),
            "31-60天库龄": nz(fba.get("inv_age_31_to_60_days")),
            "61-90天库龄": nz(fba.get("inv_age_61_to_90_days")),
            "91-180天库龄": nz(fba.get("inv_age_91_to_180_days")),
            "181-270天库龄": nz(fba.get("inv_age_181_to_270_days")),
            "270天以上库龄": age_270_plus,
            "款号": spu,
            "店铺": text(store_info.get("store_name")) or RKZ_STORE_NAME,
            "站点": text(store_info.get("region")),
            "店号": RKZ_STORE_NO,
            "部门": dept,
            "组别": dept,
            "运营": operator_name,
            "季节": combine_season(pm.get("develop_year"), pm.get("season")),
            "颜色": extract_raw_color_code(sku),
            "尺码": parse_size_from_sku(sku),
            "30天销售": sales_30d,
            "日均销量": round(sales_30d / 30.0, 2),
            "该SKU的FBA总库存可售天数": days_of_supply(fba_total_inventory, sales_30d),
            "该SKU的FBA可用库存SKU可售天数": days_of_supply(fba_available_inventory, sales_30d),
            "该款号的FBA总库存款号可售天数": None,
            "该款号的FBA可用库存款号可售天数": None,
            "负责人": None,
            "未来6个月月度仓储费": None,
            "未来6个月超龄库存附加费": None,
            "未来6个月仓储费合计": None,
            "一次性弃置费用": None,
            "延迟6个月处理总成本": None,
            "延迟处理新增成本": None,
            "币种": text(perf.get("currency_code")),
            "存放2个月月度仓储费": None,
            "存放2个月超龄库存附加费": None,
            "存放2个月仓储费合计": None,
            "品类": text(pm.get("product_category")),
        }
        built.append(row)

    # 款号级汇总：按 店铺+SPU 计算，不做SKU天数平均
    spu_agg: Dict[Tuple[str, str], Dict[str, float]] = defaultdict(
        lambda: {"total": 0.0, "available": 0.0, "sales": 0.0}
    )

    for row in built:
        key = (text(row.get("店铺")), text(row.get("款号")))
        spu_agg[key]["total"] += nz(row.get("FBA总库存"))
        spu_agg[key]["available"] += nz(row.get("FBA可用库存"))
        spu_agg[key]["sales"] += nz(row.get("30天销售"))

    for row in built:
        key = (text(row.get("店铺")), text(row.get("款号")))
        agg = spu_agg[key]
        row["该款号的FBA总库存款号可售天数"] = days_of_supply(
            agg["total"], agg["sales"]
        )
        row["该款号的FBA可用库存款号可售天数"] = days_of_supply(
            agg["available"], agg["sales"]
        )

    built.sort(
        key=lambda r: (
            text(r.get("款号")),
            text(r.get("颜色")),
            text(r.get("尺码")),
            text(r.get("SKU")),
        )
    )

    logger.info(f"MVP结果构建完成: {len(built)} 行")
    return built


def export_excel(rows: Iterable[Dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "RKZ-US库存MVP"

    ws.append(OUTPUT_COLUMNS)

    row_count = 0
    for item in rows:
        ws.append([item.get(col) for col in OUTPUT_COLUMNS])
        row_count += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # 基础列宽，仅为便于MVP核对
    for idx, column_name in enumerate(OUTPUT_COLUMNS, start=1):
        if column_name in {"MSKU", "SKU"}:
            width = 30
        elif column_name in {"ASIN", "FNSKU"}:
            width = 18
        elif "可售天数" in column_name:
            width = 24
        elif "仓储费" in column_name or "成本" in column_name or "弃置" in column_name:
            width = 22
        else:
            width = max(12, min(20, len(column_name) * 2 + 2))
        ws.column_dimensions[get_column_letter(idx)].width = width

    wb.save(output_path)
    logger.info(f"Excel已输出: {output_path}，共 {row_count} 行")


def print_sample(rows: List[Dict[str, Any]], sample: int) -> None:
    if sample <= 0:
        return

    show_columns = [
        "ASIN",
        "FNSKU",
        "MSKU",
        "SKU",
        "FBA总库存",
        "FBA可用库存",
        "款号",
        "运营",
        "部门",
        "颜色",
        "尺码",
        "30天销售",
        "该SKU的FBA总库存可售天数",
    ]

    print("\n=== RKZ-US MVP SAMPLE ===")
    for row in rows[:sample]:
        print(" | ".join(f"{col}={row.get(col)}" for col in show_columns))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="RKZ-US FBA库存治理 MVP")
    parser.add_argument(
        "--output",
        default=None,
        help="输出Excel路径；默认写入 tmp/rkz_fba_inventory_mvp_时间戳.xlsx",
    )
    parser.add_argument(
        "--sample",
        type=int,
        default=10,
        help="控制台打印前N行关键字段，默认10；传0关闭",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    if args.output:
        output_path = Path(args.output).expanduser().resolve()
    else:
        output_path = (
            PROJECT_ROOT
            / "tmp"
            / f"rkz_fba_inventory_mvp_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        )

    logger.info(
        f"开始执行 RKZ-US FBA库存MVP: sid={RKZ_SID}, 店号={RKZ_STORE_NO}"
    )

    validate_schema()
    rows = build_rows()

    if not rows:
        logger.warning("RKZ-US 未查询到FBA库存数据，不生成Excel")
        return

    print_sample(rows, args.sample)
    export_excel(rows, output_path)

    logger.info("RKZ-US FBA库存MVP执行完成")


if __name__ == "__main__":
    main()
