import json
import tempfile
import unittest
from datetime import date
from pathlib import Path
from unittest.mock import AsyncMock, patch

from openpyxl import load_workbook

from jobs.feishu.color_mapping_catalog import ColorMappingCatalog
from jobs.feishu import fabric_color_stocking as stocking
from jobs.feishu.fabric_color_stocking import (
    MATCH_CATALOG_LABEL,
    MATCH_CODE,
    MATCH_CORE_ALIAS,
    MATCH_MANUAL,
    MATCH_NAME,
    MATCH_SYSTEM,
    REASON_AMBIGUOUS,
    REASON_PATTERN_MISSING,
    REASON_PENDING_SYSTEM,
    CatalogIndex,
    CatalogRow,
    FabricUsage,
    ForecastSku,
    FuzzyMatchConfig,
    ManualMapping,
    ManualMappingCatalog,
    TARGET_FABRICS,
    aggregate_stocking,
    export_workbook,
    match_catalog_row,
    parse_feishu_catalog_records,
    recommend_fuzzy_candidates,
)


def forecast(**overrides):
    values = {
        "sku": "SP1-BK-S",
        "spu": "SP1",
        "product_name": "示例品名-1#黑色",
        "color_code": "BK",
        "color_name": "1#黑色",
        "color_system": "A2023",
        "forecast_qty": 100,
    }
    values.update(overrides)
    return ForecastSku(**values)


def usage(
    spu="SP1",
    fabric_name="面料A",
    unit_usage=0.5,
    loss=1.1,
):
    return FabricUsage(spu, fabric_name, unit_usage, loss)


class FabricColorMatchingTests(unittest.TestCase):
    def setUp(self):
        self.governance = ColorMappingCatalog(
            [
                (1, "BK", "Black", "黑色", "", "A2023"),
                (2, "BK", "Black Onyx", "黑玛瑙", "", "B2024"),
                (3, "WH", "White", "白色", "", "A2023"),
            ]
        )

    def test_chinese_name_has_priority_over_conflicting_code(self):
        rows = [
            CatalogRow("面料A", "1#黑色", "NEW-BK"),
            CatalogRow("面料A", "别的颜色", "BK"),
        ]
        decision = match_catalog_row(
            forecast(), "面料A", CatalogIndex(rows), self.governance
        )
        self.assertEqual(decision.method, MATCH_NAME)
        self.assertEqual(decision.row.color_name, "1#黑色")

    def test_code_is_second_priority(self):
        rows = [CatalogRow("面料A", "黑色现用名", "BK")]
        decision = match_catalog_row(
            forecast(color_name="旧中文名"),
            "面料A",
            CatalogIndex(rows),
            self.governance,
        )
        self.assertEqual(decision.method, MATCH_CODE)

    def test_catalog_number_prefix_is_explicit_match_method(self):
        rows = [CatalogRow("面料A", "2#黑色", "NEW-BK")]
        decision = match_catalog_row(
            forecast(
                color_name="黑色",
                color_code="OLD",
                color_system="待定",
            ),
            "面料A",
            CatalogIndex(rows),
            self.governance,
        )
        self.assertEqual(decision.method, MATCH_CATALOG_LABEL)
        self.assertEqual(decision.row.color_name, "2#黑色")

    def test_catalog_dash_prefix_is_explicit_match_method(self):
        rows = [CatalogRow("面料A", "037-拿铁", "")]
        decision = match_catalog_row(
            forecast(
                color_name="拿铁",
                color_code="OLD",
                color_system="待定",
            ),
            "面料A",
            CatalogIndex(rows),
            self.governance,
        )
        self.assertEqual(decision.method, MATCH_CATALOG_LABEL)

    def test_system_mapping_resolves_code_to_current_chinese(self):
        rows = [CatalogRow("面料A", "黑色", "NEW-BK")]
        decision = match_catalog_row(
            forecast(color_name="旧中文名"),
            "面料A",
            CatalogIndex(rows),
            self.governance,
        )
        self.assertEqual(decision.method, MATCH_SYSTEM)

    def test_deterministic_parenthetical_alias_is_unique_only(self):
        rows = [CatalogRow("面料A", "1#黑玛瑙（黑色）", "ONYX")]
        decision = match_catalog_row(
            forecast(
                color_code="OLD",
                color_name="黑玛瑙",
                color_system="待定",
            ),
            "面料A",
            CatalogIndex(rows),
            self.governance,
        )
        self.assertEqual(decision.method, MATCH_CORE_ALIAS)

    def test_color_compilation_historical_chinese_alias_is_supported(self):
        governance = ColorMappingCatalog(
            [
                (1, "X1", "Old", "旧名称", "", "A2023"),
                (2, "X1", "Alias", "历史别名", "", "A2023"),
            ]
        )
        rows = [CatalogRow("面料A", "1#历史别名", "NEW-X1")]
        decision = match_catalog_row(
            forecast(
                color_code="X1",
                color_name="无法直连",
                color_system="A2023",
            ),
            "面料A",
            CatalogIndex(rows),
            governance,
        )
        self.assertEqual(decision.method, MATCH_CORE_ALIAS)

    def test_same_chinese_core_with_multiple_color_numbers_is_ambiguous(self):
        rows = [
            CatalogRow("面料A", "1#黑玛瑙（黑色）", ""),
            CatalogRow("面料A", "2#黑玛瑙（黑色）", ""),
        ]
        decision = match_catalog_row(
            forecast(
                color_code="OLD",
                color_name="黑玛瑙",
                color_system="待定",
            ),
            "面料A",
            CatalogIndex(rows),
            self.governance,
        )
        self.assertEqual(decision.reason_code, REASON_AMBIGUOUS)
        self.assertEqual(len(decision.candidates), 2)

    def test_pending_system_does_not_use_cross_system_code_normalization(self):
        rows = [CatalogRow("面料A", "1#黑色", "BK")]
        decision = match_catalog_row(
            forecast(
                color_name="1# 黑色",
                color_code="bk",
                color_system="待定",
            ),
            "面料A",
            CatalogIndex(rows),
            self.governance,
        )
        self.assertEqual(decision.reason_code, REASON_PENDING_SYSTEM)

    def test_manual_mapping_is_below_exact_match_and_above_fuzzy(self):
        rows = [
            CatalogRow("面料A", "1#黑色", "BK", ("rec-black",)),
            CatalogRow("面料A", "2#深黑", "DB", ("rec-dark",)),
        ]
        manual = ManualMappingCatalog(
            [
                ManualMapping(
                    "面料A",
                    "OLD",
                    "旧黑",
                    "待定",
                    "rec-dark",
                    "2#深黑",
                ),
                ManualMapping(
                    "面料A",
                    "BK",
                    "1#黑色",
                    "A2023",
                    "rec-dark",
                    "2#深黑",
                ),
            ]
        )
        manual_decision = match_catalog_row(
            forecast(
                color_code="OLD",
                color_name="旧黑",
                color_system="待定",
            ),
            "面料A",
            CatalogIndex(rows),
            self.governance,
            manual,
        )
        exact_decision = match_catalog_row(
            forecast(),
            "面料A",
            CatalogIndex(rows),
            self.governance,
            manual,
        )
        self.assertEqual(manual_decision.method, MATCH_MANUAL)
        self.assertEqual(manual_decision.row.color_name, "2#深黑")
        self.assertEqual(exact_decision.method, MATCH_NAME)
        self.assertEqual(exact_decision.row.color_name, "1#黑色")

    def test_fuzzy_candidates_stay_in_same_fabric_and_are_not_auto(self):
        rows = [
            CatalogRow("面料A", "1#黑色", "BK"),
            CatalogRow("面料B", "黑色新版", "BK2"),
        ]
        item = forecast(
            color_code="OLD",
            color_name="黑色新版",
            color_system="待定",
        )
        index = CatalogIndex(rows)
        decision = match_catalog_row(
            item, "面料A", index, self.governance
        )
        candidates = recommend_fuzzy_candidates(
            item,
            "面料A",
            index,
            self.governance,
            config=FuzzyMatchConfig(low_min_score=70),
        )
        self.assertIsNone(decision.row)
        self.assertEqual([candidate.row.fabric_name for candidate in candidates], ["面料A"])

    def test_pattern_color_is_not_forced_to_pure_color(self):
        rows = [
            CatalogRow("面料A", "1#黑色", "BK"),
            CatalogRow("面料A", "2#白色", "WH"),
        ]
        item = forecast(
            color_code="OLD",
            color_name="黑底白点",
            product_name="黑底白点内裤",
            color_system="待定",
        )
        index = CatalogIndex(rows)
        decision = match_catalog_row(
            item, "面料A", index, self.governance
        )
        candidates = recommend_fuzzy_candidates(
            item, "面料A", index, self.governance
        )
        self.assertEqual(decision.reason_code, REASON_PATTERN_MISSING)
        self.assertEqual(candidates, [])

    def test_exact_duplicate_catalog_rows_are_consolidated_with_ids(self):
        rows = [
            CatalogRow("面料A", "1#黑色", "BK", ("rec1",)),
            CatalogRow("面料A", "1#黑色", "BK", ("rec2",)),
        ]
        index = CatalogIndex(rows, source_record_count=2)
        self.assertEqual(len(index.rows), 1)
        self.assertEqual(index.rows[0].raw_row_count, 2)
        self.assertEqual(index.rows[0].record_ids, ("rec1", "rec2"))
        self.assertEqual(
            index.audit()["exact_duplicate_business_group_count"], 1
        )

    def test_feishu_link_ids_are_resolved_to_fabric_names(self):
        raw = [
            {
                "record_id": "rec-color",
                "fields": {
                    "面料名": [{"record_id": "rec-fabric"}],
                    "颜色": "1#黑色",
                    "领星新颜色缩写": "BK",
                },
            }
        ]
        rows, audit = parse_feishu_catalog_records(
            raw, {"rec-fabric": "面料A"}
        )
        self.assertEqual(rows[0].fabric_name, "面料A")
        self.assertEqual(rows[0].record_ids, ("rec-color",))
        self.assertEqual(audit["invalid_record_count"], 0)


class FabricUsageAggregationTests(unittest.TestCase):
    def setUp(self):
        self.governance = ColorMappingCatalog(
            [(1, "BK", "Black", "黑色", "", "A2023")]
        )
        self.catalog = [
            CatalogRow("面料A", "1#黑色", "BK", ("rec-black",))
        ]

    def aggregate(self, forecasts, usages, **kwargs):
        return aggregate_stocking(
            forecasts,
            usages,
            self.catalog,
            self.governance,
            **kwargs,
        )

    def test_forecast_times_unit_usage_times_loss(self):
        result = self.aggregate(
            [forecast(forecast_qty=100)],
            {"SP1": [usage(unit_usage=0.5, loss=1.1)]},
        )
        self.assertEqual(result.metrics["calculable_total_meters"], 55.0)
        self.assertEqual(result.metrics["confirmed_color_meters"], 55.0)
        self.assertEqual(
            result.color_usage_rows[0]["预估面料用量（米）"], 55.0
        )

    def test_missing_unit_usage_never_uses_other_spu_average(self):
        forecasts = [
            forecast(sku="MISSING", forecast_qty=100),
            forecast(sku="NORMAL", spu="SP2", forecast_qty=10),
        ]
        usages = {
            "SP1": [usage(unit_usage=0, loss=1.2)],
            "SP2": [usage(spu="SP2", unit_usage=1, loss=1)],
        }
        result = self.aggregate(forecasts, usages)
        self.assertEqual(result.metrics["calculable_total_meters"], 10.0)
        self.assertEqual(result.metrics["missing_usage_forecast_qty"], 100)
        self.assertEqual(len(result.usage_missing_rows), 1)
        self.assertEqual(
            result.usage_missing_rows[0]["缺失原因"], "单件用量为空或0"
        )

    def test_unmatched_color_still_calculates_meters_for_manual_review(self):
        result = self.aggregate(
            [
                forecast(
                    color_code="OLD",
                    color_name="不存在颜色",
                    color_system="待定",
                )
            ],
            {"SP1": [usage(unit_usage=0.5, loss=1.1)]},
        )
        self.assertEqual(result.metrics["calculable_total_meters"], 55.0)
        self.assertEqual(result.metrics["confirmed_color_meters"], 0.0)
        self.assertEqual(result.metrics["manual_review_color_meters"], 55.0)
        self.assertEqual(result.manual_review_rows[0]["已算面料米数"], 55.0)

    def test_multi_candidate_never_auto_selects(self):
        rows = [
            CatalogRow("面料A", "1#黑色", "", ("rec1",)),
            CatalogRow("面料A", "2#黑色", "", ("rec2",)),
        ]
        result = aggregate_stocking(
            [
                forecast(
                    color_code="OLD",
                    color_name="黑色",
                    color_system="待定",
                )
            ],
            {"SP1": [usage()]},
            rows,
            self.governance,
        )
        self.assertEqual(result.auto_merge_rows, [])
        self.assertEqual(len(result.manual_review_rows), 1)
        self.assertIn(
            "一对多", result.manual_review_rows[0]["未自动归并原因"]
        )

    def test_all_target_fabrics_always_appear_in_overview(self):
        result = self.aggregate(
            [],
            {},
            target_fabrics=TARGET_FABRICS,
        )
        fabric_names = [
            row["面料名"]
            for row in result.overview_rows
            if row["范围"] == "面料"
        ]
        self.assertEqual(fabric_names, list(TARGET_FABRICS))
        self.assertTrue(
            all(row["可确定总用量（米）"] == 0 for row in result.overview_rows[1:])
        )

    def test_zero_demand_feishu_color_is_still_output(self):
        result = self.aggregate([], {})
        self.assertEqual(len(result.color_usage_rows), 1)
        self.assertEqual(result.color_usage_rows[0]["预估销量"], 0)
        self.assertEqual(result.color_usage_rows[0]["预估面料用量（米）"], 0.0)

    def test_confirmed_and_fuzzy_coverages_are_separate(self):
        rows = [
            CatalogRow("面料A", "1#黑色", "BK"),
            CatalogRow("面料A", "2#蓝色", "BL"),
        ]
        forecasts = [
            forecast(sku="EXACT", forecast_qty=50),
            forecast(
                sku="FUZZY",
                color_code="OLD",
                color_name="蓝色新版",
                color_system="待定",
                forecast_qty=50,
            ),
        ]
        result = aggregate_stocking(
            forecasts,
            {"SP1": [usage(unit_usage=1, loss=1)]},
            rows,
            self.governance,
        )
        self.assertEqual(result.metrics["confirmed_color_coverage_pct"], 50.0)
        self.assertEqual(result.metrics["fuzzy_candidate_coverage_pct"], 50.0)
        self.assertEqual(result.metrics["potential_color_coverage_pct"], 100.0)
        self.assertEqual(
            sum(row["预估面料用量（米）"] for row in result.color_usage_rows),
            50.0,
        )

    def test_export_contains_all_required_sheets_and_zero_rows(self):
        result = self.aggregate(
            [forecast()],
            {"SP1": [usage()]},
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path, metrics_path = export_workbook(
                result, Path(temp_dir)
            )
            workbook = load_workbook(workbook_path, read_only=True)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    [
                        "颜色用量总览",
                        "飞书颜色用量",
                        "自动归并SKU明细",
                        "待人工确认SKU",
                        "用量参数缺失",
                        "飞书17面料颜色清单",
                        "模糊候选审核",
                        "优先补标清单",
                        "核对摘要",
                    ],
                )
                self.assertIn(
                    "预估面料用量（米）",
                    [
                        cell.value
                        for cell in next(workbook["飞书颜色用量"].rows)
                    ],
                )
            finally:
                workbook.close()
            metrics = json.loads(metrics_path.read_text(encoding="utf-8"))
            self.assertEqual(metrics["confirmed_color_meters"], 55.0)


class DryRunBoundaryTests(unittest.IsolatedAsyncioTestCase):
    async def test_dry_run_orchestration_has_no_remote_write_path(self):
        governance = ColorMappingCatalog(
            [(1, "BK", "Black", "黑色", "", "A2023")]
        )
        catalog = [
            CatalogRow("002双六面料", "1#黑色", "BK", ("rec1",))
        ]
        forecasts = [
            forecast(
                spu="SP1",
                forecast_qty=10,
            )
        ]
        usages = {
            "SP1": [
                usage(
                    fabric_name="002双六面料",
                    unit_usage=1,
                    loss=1,
                )
            ]
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            with (
                patch.object(
                    stocking.ColorMappingCatalog,
                    "from_runtime",
                    return_value=governance,
                ),
                patch.object(
                    stocking,
                    "load_catalog_from_feishu",
                    new=AsyncMock(
                        return_value=(
                            catalog,
                            {"source_record_count": 1},
                        )
                    ),
                ),
                patch.object(
                    stocking,
                    "load_forecast_skus",
                    return_value=(forecasts, {}),
                ),
                patch.object(
                    stocking,
                    "load_fabric_usage_by_spu",
                    return_value=usages,
                ),
                patch.object(
                    stocking,
                    "load_manual_mapping_catalog",
                    return_value=ManualMappingCatalog(),
                ),
                patch.object(stocking, "db_cursor") as db_cursor_mock,
                patch.object(stocking, "FeishuClient") as feishu_client_mock,
            ):
                result, workbook_path, metrics_path = (
                    await stocking.run_read_only(
                        Path(temp_dir),
                        as_of=date(2026, 7, 28),
                    )
                )
            db_cursor_mock.assert_not_called()
            feishu_client_mock.assert_not_called()
            self.assertEqual(result.metrics["confirmed_color_meters"], 10.0)
            self.assertTrue(workbook_path.exists())
            self.assertTrue(metrics_path.exists())


if __name__ == "__main__":
    unittest.main()
