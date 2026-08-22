import unittest
from datetime import date

from jobs.feishu.historical_fabric_usage import (
    LOSS_FACTOR,
    aggregate_usage,
    expand_purchase_orders,
    export_excel,
    month_range,
    normalize_fabric_map,
)


class HistoricalFabricUsageTests(unittest.TestCase):
    def setUp(self):
        self.fabric_map = normalize_fabric_map([
            {'SPU': 'BX402', '面料': '037超绒', '单件用量': 0.82},
            {'SPU': 'BX402', '面料': '罗纹A', '单件用量': 0.12},
        ])

    def test_month_range(self):
        self.assertEqual(
            month_range(date(2025, 6, 1), date(2025, 8, 6)),
            ['2025-06', '2025-07', '2025-08'],
        )

    def test_fixed_loss_and_entry_quantity(self):
        orders = [{
            'order_sn': 'PO1', 'order_time': '2025-08-12 10:00:00', 'status': 9,
            'status_text': '已完成', 'status_shipped_text': '部分到货',
            'item_list': [{
                'id': 1, 'sku': 'BX402-BK-S', 'spu': 'BX402', 'is_delete': 0,
                'quantity_real': 100, 'quantity_entry': 60, 'quantity_receive': 40,
            }],
        }]
        detail, unmapped = expand_purchase_orders(
            orders, date(2025, 6, 1), date(2025, 8, 31), self.fabric_map, {'037超绒'}
        )
        self.assertEqual(len(unmapped), 0)
        self.assertEqual(len(detail), 2)
        rows = {r['面料']: r for r in detail}
        self.assertAlmostEqual(rows['037超绒']['本批面料用量(M)'], 60 * 0.82 * LOSS_FACTOR)
        self.assertAlmostEqual(rows['罗纹A']['本批面料用量(M)'], 60 * 0.12 * LOSS_FACTOR)
        self.assertEqual(rows['037超绒']['面料种类'], '定制')
        self.assertEqual(rows['罗纹A']['面料种类'], '现货')

    def test_filter_by_order_time_and_void_deleted(self):
        orders = [
            {'order_sn': 'BEFORE', 'order_time': '2025-05-31 23:59:59', 'status': 9,
             'item_list': [{'id': 1, 'sku': 'BX402-BK-S', 'spu': 'BX402', 'quantity_entry': 10}]},
            {'order_sn': 'VOID', 'order_time': '2025-06-01 00:00:00', 'status': -1,
             'item_list': [{'id': 2, 'sku': 'BX402-BK-S', 'spu': 'BX402', 'quantity_entry': 10}]},
            {'order_sn': 'DELETED', 'order_time': '2025-06-01 00:00:00', 'status': 9,
             'item_list': [{'id': 3, 'sku': 'BX402-BK-S', 'spu': 'BX402', 'is_delete': 1, 'quantity_entry': 10}]},
        ]
        detail, unmapped = expand_purchase_orders(
            orders, date(2025, 6, 1), date(2025, 8, 31), self.fabric_map, {'037超绒'}
        )
        self.assertEqual(detail, [])
        self.assertEqual(unmapped, [])

    def test_unmapped_is_visible(self):
        orders = [{
            'order_sn': 'PO2', 'order_time': '2025-06-15 10:00:00', 'status': 9,
            'item_list': [{'id': 4, 'sku': 'UNKNOWN-BK-S', 'spu': 'UNKNOWN', 'quantity_entry': 20}],
        }]
        detail, unmapped = expand_purchase_orders(
            orders, date(2025, 6, 1), date(2025, 8, 31), self.fabric_map, set()
        )
        self.assertEqual(detail, [])
        self.assertEqual(len(unmapped), 1)
        self.assertEqual(unmapped[0]['异常原因'], '面料核价表无款号映射')

    def test_aggregate(self):
        detail = [
            {'采购月份': '2025-08', '面料': '037超绒', '面料种类': '定制', '款号': 'BX402',
             '单件用量(M)': 0.82, '本批面料用量(M)': 90.2},
            {'采购月份': '2025-08', '面料': '037超绒', '面料种类': '定制', '款号': 'BX402',
             '单件用量(M)': 0.82, '本批面料用量(M)': 45.1},
        ]
        fabrics, spus = aggregate_usage(detail, ['2025-06', '2025-07', '2025-08'])
        self.assertEqual(len(fabrics), 1)
        self.assertEqual(fabrics[0]['2025-08'], 135.3)
        self.assertEqual(spus[0]['2025-08'], 135.3)

    def test_export_excel_four_sheets(self):
        from pathlib import Path
        from tempfile import TemporaryDirectory
        from openpyxl import load_workbook

        with TemporaryDirectory() as td:
            out = Path(td) / 'out.xlsx'
            export_excel(
                out,
                [{'面料': '037超绒', '面料种类': '定制', '2025-06': 10.0, '累计用量(M)': 10.0}],
                [{'款号': 'BX402', '面料': '037超绒', '面料种类': '定制', '单件用量(M)': 0.82,
                  '含损耗单件用量(M)': 0.902, '2025-06': 10.0, '累计用量(M)': 10.0}],
                [], [], ['2025-06']
            )
            wb = load_workbook(out, read_only=True)
            self.assertEqual(wb.sheetnames, ['面料月度用量', '款号面料月度用量', '采购明细核算', '未映射检查'])
            self.assertEqual(wb['面料月度用量']['A2'].value, '037超绒')


if __name__ == '__main__':
    unittest.main()
