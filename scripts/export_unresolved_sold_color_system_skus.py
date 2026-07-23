#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""导出“有销量但颜色体系待定”的 SKU 清单。"""
from __future__ import annotations

import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from jobs.feishu.color_mapping_catalog import ColorMappingCatalog
from jobs.feishu.color_system_resolver import ColorSystemResolver, UNKNOWN_SYSTEM
from scripts import analyze_unresolved_color_system_sales as base

OUTPUT_DIR = ROOT / "exports"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws, minimum: int = 10, maximum: int = 42) -> None:
    for cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in cells)
        ws.column_dimensions[get_column_letter(cells[0].column)].width = min(
            max(length + 2, minimum), maximum
        )


def style_header(ws, row: int, headers: list[str]) -> None:
    for index, header in enumerate(headers, 1):
        cell = ws.cell(row, index, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def main() -> Path:
    today = datetime.now().date()
    snapshot = base.load_snapshot()
    resolver = ColorSystemResolver(snapshot.values())
    catalog = ColorMappingCatalog.from_runtime()
    identities = {
        sku: resolver.resolve(sku, str(row.get("spu") or ""))
        for sku, row in snapshot.items()
    }
    unresolved = {
        sku for sku, identity in identities.items()
        if identity.color_system == UNKNOWN_SYSTEM
    }

    sales_by_sku = defaultdict(lambda: {period: 0.0 for period in base.PERIODS})
    for row in base.load_sales():
        sku = base.normalize_sku(row.get("SKU"))
        stat_date = base.parse_date(row.get("统计日期"))
        qty = float(row.get("销量") or 0)
        if not sku or sku not in snapshot or stat_date is None or qty == 0:
            continue
        for period in base.included_periods(stat_date, today):
            sales_by_sku[sku][period] += qty

    sold_unresolved = [
        sku for sku in unresolved if sales_by_sku[sku]["全历史"] > 0
    ]
    sold_unresolved.sort(
        key=lambda sku: (
            -sales_by_sku[sku]["当年累计"],
            -sales_by_sku[sku]["近12个月"],
            -sales_by_sku[sku]["全历史"],
            sku,
        )
    )

    workbook = Workbook()
    summary = workbook.active
    summary.title = "汇总"
    summary_rows = [
        ("统计日期", str(today)),
        ("当前产品SKU总数", len(snapshot)),
        ("颜色体系待定SKU", len(unresolved)),
        ("待定且有历史销量SKU", len(sold_unresolved)),
        ("待定有销量SKU占全盘", len(sold_unresolved) / len(snapshot) if snapshot else 0),
        ("待定有销量SKU占待定SKU", len(sold_unresolved) / len(unresolved) if unresolved else 0),
        ("全历史销量", sum(sales_by_sku[x]["全历史"] for x in sold_unresolved)),
        ("当年累计销量", sum(sales_by_sku[x]["当年累计"] for x in sold_unresolved)),
        ("近12个月销量", sum(sales_by_sku[x]["近12个月"] for x in sold_unresolved)),
    ]
    style_header(summary, 1, ["指标", "数值"])
    for row_index, (name, value) in enumerate(summary_rows, 2):
        summary.cell(row_index, 1, name)
        summary.cell(row_index, 2, value)
        if "占" in name:
            summary.cell(row_index, 2).number_format = "0.0000%"
        elif isinstance(value, (int, float)):
            summary.cell(row_index, 2).number_format = "#,##0.00"
    autosize(summary)

    detail = workbook.create_sheet("有销量待定SKU清单")
    headers = [
        "SKU", "SPU", "产品名称", "颜色代码",
        "A2023中文候选", "B2024中文候选",
        "A2023英文候选", "B2024英文候选",
        "A2023潘通", "B2024潘通",
        "全历史销量", "当年累计销量", "近12个月销量",
        "当前识别来源", "人工确认建议",
    ]
    style_header(detail, 1, headers)
    for row_index, sku in enumerate(sold_unresolved, 2):
        identity = identities[sku]
        candidates = catalog.candidates(identity.color_code)
        a = candidates["A2023"]
        b = candidates["B2024"]
        if a and b and a.chinese != b.chinese:
            advice = "按产品实物/开发年份确认A2023或B2024"
        elif a and b:
            advice = "A/B中文同名，仍需按开发年份确认体系"
        elif a or b:
            advice = "仅一套颜色表收录，建议人工核对后补体系"
        else:
            advice = "颜色代码未收录，需先补颜色映射"
        values = [
            sku,
            identity.spu,
            snapshot[sku].get("product_name") or "",
            identity.color_code,
            a.chinese if a else "",
            b.chinese if b else "",
            a.english if a else "",
            b.english if b else "",
            a.pantone if a else "",
            b.pantone if b else "",
            round(sales_by_sku[sku]["全历史"], 2),
            round(sales_by_sku[sku]["当年累计"], 2),
            round(sales_by_sku[sku]["近12个月"], 2),
            identity.source,
            advice,
        ]
        for column, value in enumerate(values, 1):
            detail.cell(row_index, column, value)
        for column in (11, 12, 13):
            detail.cell(row_index, column).number_format = "#,##0.00"
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions
    autosize(detail)

    by_code: dict[str, dict[str, Any]] = {}
    for sku in sold_unresolved:
        code = identities[sku].color_code
        bucket = by_code.setdefault(code, {
            "SKU数": 0, "全历史销量": 0.0, "当年累计销量": 0.0, "近12个月销量": 0.0
        })
        bucket["SKU数"] += 1
        for period in base.PERIODS:
            bucket[period + "销量"] += sales_by_sku[sku][period]
    code_sheet = workbook.create_sheet("颜色代码汇总")
    code_headers = [
        "颜色代码", "A2023中文候选", "B2024中文候选", "SKU数",
        "全历史销量", "当年累计销量", "近12个月销量",
    ]
    style_header(code_sheet, 1, code_headers)
    sorted_codes = sorted(
        by_code,
        key=lambda code: (-by_code[code]["当年累计销量"], code),
    )
    for row_index, code in enumerate(sorted_codes, 2):
        a = catalog.lookup("A2023", code)
        b = catalog.lookup("B2024", code)
        bucket = by_code[code]
        values = [
            code, a.chinese if a else "", b.chinese if b else "", bucket["SKU数"],
            bucket["全历史销量"], bucket["当年累计销量"], bucket["近12个月销量"],
        ]
        for column, value in enumerate(values, 1):
            code_sheet.cell(row_index, column, value)
    code_sheet.freeze_panes = "A2"
    code_sheet.auto_filter.ref = code_sheet.dimensions
    autosize(code_sheet)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output = OUTPUT_DIR / f"有销量但颜色体系待定SKU清单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    workbook.save(output)
    print(f"有销量待定SKU：{len(sold_unresolved):,}")
    print(f"输出文件：{output}")
    return output


if __name__ == "__main__":
    main()
